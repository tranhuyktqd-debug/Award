import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.drawing.image import Image

class ChiaPhongThi:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Hệ thống chia phòng thi tự động")
        
        # Tự động điều chỉnh kích thước theo độ phân giải màn hình
        self.thiet_lap_kich_thuoc_man_hinh()
        
        # Biến lưu trữ dữ liệu
        self.df_goc = None
        self.df_da_chia = None
        self.cac_cot_duoc_chon = []
        self.thu_tu_cot_duoc_chon = []  # Lưu thứ tự click chọn cột
        self.so_thi_sinh_moi_phong = 30
        self.cau_hinh_phong = {}  # Lưu cấu hình số lượng thí sinh cho từng phòng
        
        # Biến lưu trữ thông tin về file Excel và sheets
        self.duong_dan_file_excel = None
        self.danh_sach_sheet = []
        self.sheet_hien_tai = None
        
        # Biến cho chế độ hybrid
        self.cot_che_do_thi = None  # Cột chứa thông tin chế độ thi (Offline/Online)
        
        # Biến lưu trữ thông tin bổ sung
        self.ban_to_chuc = ""
        self.diem_thi = ""
        self.mon_thi = "Toán"
        self.thoi_gian_co_mat = ""
        
        # Biến lưu trữ chế độ thi
        self.che_do_thi = "offline"  # offline, online, hybrid
        
        # Biến lưu trữ đường dẫn ảnh
        self.duong_dan_anh_trai = ""
        self.duong_dan_anh_phai = ""
        
        # Biến lưu trữ đường dẫn ảnh cho danh sách trường
        self.duong_dan_anh_trai_truong = ""
        self.duong_dan_anh_phai_truong = ""
        
        # Biến lưu trữ chiều cao ảnh tùy chỉnh (đơn vị: cm)
        self.chieu_cao_anh = 2.0  # Mặc định 2cm
        
        # Load danh sách địa điểm thi từ file Excel
        self.danh_sach_dia_diem, self.ma_dia_diem = self.load_dia_diem_thi()
        
        # Load thông tin giờ thi từ file Excel
        self.thong_tin_gio_thi = self.load_gio_thi()
        
        self.tao_giao_dien()
    
    def thiet_lap_kich_thuoc_man_hinh(self):
        """Tự động điều chỉnh kích thước cửa sổ theo độ phân giải màn hình"""
        try:
            # Lấy kích thước màn hình
            screen_width = self.root.winfo_screenwidth()
            screen_height = self.root.winfo_screenheight()
            
            print(f"🖥️ Độ phân giải màn hình: {screen_width}x{screen_height}")
            
            # Xác định loại màn hình và thiết lập kích thước phù hợp
            if screen_width >= 2560:  # 2K hoặc 4K
                # Màn hình 2K/4K - cửa sổ lớn hơn
                window_width = min(1600, int(screen_width * 0.7))
                window_height = min(1000, int(screen_height * 0.8))
                min_width = 1200
                min_height = 700
                print("📺 Phát hiện màn hình 2K/4K - sử dụng kích thước lớn")
                
            elif screen_width >= 1920:  # Full HD
                # Màn hình Full HD - kích thước trung bình
                window_width = min(1400, int(screen_width * 0.8))
                window_height = min(900, int(screen_height * 0.85))
                min_width = 1000
                min_height = 600
                print("📺 Phát hiện màn hình Full HD - sử dụng kích thước trung bình")
                
            elif screen_width >= 1366:  # HD
                # Màn hình HD - kích thước nhỏ hơn
                window_width = min(1200, int(screen_width * 0.9))
                window_height = min(800, int(screen_height * 0.9))
                min_width = 900
                min_height = 500
                print("📺 Phát hiện màn hình HD - sử dụng kích thước nhỏ")
                
            else:  # Màn hình nhỏ
                # Màn hình nhỏ - tối ưu không gian
                window_width = min(1000, int(screen_width * 0.95))
                window_height = min(700, int(screen_height * 0.95))
                min_width = 800
                min_height = 500
                print("📺 Phát hiện màn hình nhỏ - tối ưu không gian")
            
            # Thiết lập kích thước cửa sổ
            self.root.geometry(f"{window_width}x{window_height}")
            self.root.minsize(min_width, min_height)
            
            # Căn giữa cửa sổ trên màn hình
            x = (screen_width - window_width) // 2
            y = (screen_height - window_height) // 2
            self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
            
            print(f"✅ Kích thước cửa sổ: {window_width}x{window_height}")
            print(f"✅ Kích thước tối thiểu: {min_width}x{min_height}")
            
        except Exception as e:
            print(f"❌ Lỗi khi thiết lập kích thước màn hình: {str(e)}")
            # Fallback về kích thước mặc định
            self.root.geometry("1200x800")
            self.root.minsize(1000, 600)
    
    def load_dia_diem_thi(self):
        """Load danh sách địa điểm thi từ file Excel với mã phòng thi"""
        try:
            # Lấy đường dẫn thư mục chứa script
            script_dir = os.path.dirname(os.path.abspath(__file__))
            file_dia_diem = os.path.join(script_dir, "1 ĐỊA ĐIỂM THI.xlsx")
            
            # Kiểm tra file có tồn tại không
            if not os.path.exists(file_dia_diem):
                print(f"⚠️ Không tìm thấy file: {file_dia_diem}")
                return [], {}
            
            # Đọc file Excel (bỏ qua dòng 1, đọc từ dòng 2)
            df_dia_diem = pd.read_excel(file_dia_diem, engine='openpyxl', header=None, skiprows=1)
            
            # Lấy cột đầu tiên (địa điểm thi)
            dia_diem_list = df_dia_diem[0].dropna().tolist()
            dia_diem_list = [str(item).strip() for item in dia_diem_list if str(item).strip()]
            
            # Load mã địa điểm (cột 2 - mã viết tắt của địa điểm)
            ma_dia_diem_dict = {}
            for index, row in df_dia_diem.iterrows():
                dia_diem = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                ma_dia_diem = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ""
                
                if dia_diem:
                    # Nếu không có mã địa điểm trong file, tự động tạo từ tên
                    if not ma_dia_diem:
                        # Lấy chữ cái đầu của mỗi từ (VD: "BAN MAI" -> "BM")
                        ma_dia_diem = ''.join([word[0].upper() for word in dia_diem.split() if word])
                    
                    ma_dia_diem_dict[dia_diem] = ma_dia_diem
            
            print(f"✅ Đã load {len(dia_diem_list)} địa điểm thi từ file")
            return dia_diem_list, ma_dia_diem_dict
            
        except Exception as e:
            print(f"❌ Lỗi khi load file địa điểm thi: {str(e)}")
            import traceback
            traceback.print_exc()
            return [], {}
    
    def load_gio_thi(self):
        """Load thông tin giờ thi từ file Excel theo cấu trúc: Ban tổ chức | Môn thi | Giờ thi"""
        try:
            # Lấy đường dẫn thư mục chứa script
            script_dir = os.path.dirname(os.path.abspath(__file__))
            file_gio_thi = os.path.join(script_dir, "2 GIỜ THI.xlsx")
            
            # Kiểm tra file có tồn tại không
            if not os.path.exists(file_gio_thi):
                print(f"⚠️ Không tìm thấy file: {file_gio_thi}")
                print(f"💡 Tạo file mẫu '2 GIỜ THI.xlsx' với cấu trúc:")
                print(f"   Cột A: Ban tổ chức (ASMO VIỆT NAM, SEAMO VIỆT NAM, ...)")
                print(f"   Cột B: Môn thi (Toán, Khoa học, Tiếng Anh)")
                print(f"   Cột C: Giờ thi (8h30, 13h00, 15h00, ...)")
                print(f"   Dòng 1: Tiêu đề (sẽ bỏ qua)")
                return {}
            
            # Đọc file Excel (bỏ qua dòng 1 - tiêu đề, đọc từ dòng 2)
            df_gio_thi = pd.read_excel(file_gio_thi, engine='openpyxl', header=None, skiprows=1)
            
            # Kiểm tra có ít nhất 3 cột không
            if len(df_gio_thi.columns) < 3:
                print(f"⚠️ File giờ thi cần có ít nhất 3 cột (Ban tổ chức | Môn thi | Giờ thi)")
                return {}
            
            # Tạo dictionary lưu trữ: {(ban_to_chuc, mon_thi): [gio_thi1, gio_thi2, ...]}
            thong_tin_gio = {}
            
            for index, row in df_gio_thi.iterrows():
                ban_to_chuc = str(row[0]).strip() if pd.notna(row[0]) else ""
                mon_thi = str(row[1]).strip() if pd.notna(row[1]) else ""
                gio_thi = str(row[2]).strip() if pd.notna(row[2]) else ""
                
                # Bỏ qua dòng trống
                if not ban_to_chuc or not mon_thi or not gio_thi:
                    continue
                
                # Tạo key dạng (ban_to_chuc, mon_thi)
                key = (ban_to_chuc, mon_thi)
                
                # Thêm giờ thi vào danh sách
                if key not in thong_tin_gio:
                    thong_tin_gio[key] = []
                
                if gio_thi not in thong_tin_gio[key]:  # Tránh trùng lặp
                    thong_tin_gio[key].append(gio_thi)
            
            print(f"✅ Đã load {len(thong_tin_gio)} cấu hình giờ thi từ file")
            for key, gio_list in thong_tin_gio.items():
                print(f"   {key[0]} - {key[1]}: {', '.join(gio_list)}")
            
            return thong_tin_gio
            
        except Exception as e:
            print(f"❌ Lỗi khi load file giờ thi: {str(e)}")
            import traceback
            traceback.print_exc()
            return {}
        
    def tao_giao_dien(self):
        # Tạo canvas và scrollbar cho toàn bộ giao diện
        canvas = tk.Canvas(self.root)
        scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Frame container chứa main_frame và kết quả
        container_frame = ttk.Frame(scrollable_frame)
        container_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Frame chính (bên trái - 75% màn hình)
        main_frame = ttk.Frame(container_frame, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Frame kết quả (bên phải - 25% màn hình)
        self.frame_ket_qua = ttk.LabelFrame(container_frame, text="📊 Tiến trình thực hiện", padding="10")
        self.frame_ket_qua.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(10, 0))
        
        # Cấu hình grid - phân chia không gian 75%-25%
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        container_frame.columnconfigure(0, weight=3)  # 75% cho main_frame
        container_frame.columnconfigure(1, weight=1)  # 25% cho kết quả
        container_frame.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)  # Cột 0 chiếm toàn bộ
        main_frame.columnconfigure(1, weight=1)  # Cột 1 chiếm toàn bộ
        
        # Pack canvas và scrollbar
        canvas.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # Thêm khả năng cuộn bằng chuột
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # 1. Chọn file Excel
        ttk.Label(main_frame, text="1. Chọn file Excel danh sách thí sinh:", font=("Arial", 12, "bold")).grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        ttk.Button(main_frame, text="Chọn file Excel", command=self.chon_file_excel).grid(row=2, column=0, sticky=tk.W, pady=(0, 10))
        self.lbl_file_path = ttk.Label(main_frame, text="Chưa chọn file", foreground="red")
        self.lbl_file_path.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(0, 10))
        
        # Frame chọn trang tính (sheet) - ẩn ban đầu
        self.frame_chon_sheet = ttk.Frame(main_frame)
        self.frame_chon_sheet.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(5, 10))
        self.frame_chon_sheet.grid_remove()  # Ẩn ban đầu
        
        ttk.Label(self.frame_chon_sheet, text="📋 Chọn trang tính cần xử lý:", 
                 font=("Arial", 10, "bold")).grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        
        self.combo_chon_sheet = ttk.Combobox(self.frame_chon_sheet, state="readonly", width=40)
        self.combo_chon_sheet.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0))
        self.combo_chon_sheet.bind('<<ComboboxSelected>>', self.chon_sheet)
        
        # 2. Chọn chế độ thi
        ttk.Label(main_frame, text="2. Chọn chế độ thi:", font=("Arial", 12, "bold")).grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 5))
        
        # Frame cho chọn chế độ thi
        frame_che_do = ttk.Frame(main_frame)
        frame_che_do.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Radio buttons cho chế độ thi
        self.var_che_do = tk.StringVar(value="offline")
        
        ttk.Radiobutton(frame_che_do, text="📚 Thi Offline (truyền thống)", variable=self.var_che_do, 
                       value="offline", command=self.cap_nhat_che_do_thi).grid(row=0, column=0, sticky=tk.W, padx=(0, 20))
        
        ttk.Radiobutton(frame_che_do, text="💻 Thi Online", variable=self.var_che_do, 
                       value="online", command=self.cap_nhat_che_do_thi).grid(row=0, column=1, sticky=tk.W, padx=(0, 20))
        
        ttk.Radiobutton(frame_che_do, text="🔄 Thi Hỗn hợp (Offline + Online)", variable=self.var_che_do, 
                       value="hybrid", command=self.cap_nhat_che_do_thi).grid(row=0, column=2, sticky=tk.W)
        
        # Label mô tả chế độ thi
        self.lbl_mo_ta_che_do = ttk.Label(frame_che_do, text="💡 Thi Offline: Có giám thị, chữ ký thủ công", 
                                         font=("Arial", 9), foreground="blue")
        self.lbl_mo_ta_che_do.grid(row=1, column=0, columnspan=3, sticky=tk.W, pady=(5, 0))
        
        # Frame chọn cột chế độ thi (chỉ hiện khi chọn hybrid)
        self.frame_chon_cot_che_do = ttk.Frame(main_frame)
        self.frame_chon_cot_che_do.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(5, 10))
        self.frame_chon_cot_che_do.grid_remove()  # Ẩn ban đầu
        
        ttk.Label(self.frame_chon_cot_che_do, text="Chọn cột chứa thông tin chế độ thi (Offline/Online):", 
                 font=("Arial", 10, "bold")).grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        
        self.combo_cot_che_do = ttk.Combobox(self.frame_chon_cot_che_do, state="readonly", width=30)
        self.combo_cot_che_do.grid(row=1, column=0, sticky=(tk.W, tk.E), padx=(0, 10))
        self.combo_cot_che_do.bind('<<ComboboxSelected>>', self.cap_nhat_cot_che_do_thi)
        
        ttk.Label(self.frame_chon_cot_che_do, text="💡 Cột này phải chứa giá trị 'Offline' hoặc 'Online'", 
                 font=("Arial", 9), foreground="blue").grid(row=2, column=0, sticky=tk.W, pady=(5, 0))
        
        # 3. Chọn cột dữ liệu
        ttk.Label(main_frame, text="3. Chọn các cột dữ liệu cần thiết:", font=("Arial", 12, "bold")).grid(row=6, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 5))
        
        # Frame chính cho việc chọn cột và xem trước
        frame_chon_cot_chinh = ttk.Frame(main_frame)
        frame_chon_cot_chinh.grid(row=7, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        frame_chon_cot_chinh.columnconfigure(0, weight=1)
        frame_chon_cot_chinh.columnconfigure(1, weight=1)
        
        # Frame bên trái cho checkbox
        frame_checkbox_container = ttk.Frame(frame_chon_cot_chinh)
        frame_checkbox_container.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 10))
        frame_checkbox_container.columnconfigure(0, weight=1)
        
        self.frame_checkbox = ttk.Frame(frame_checkbox_container)
        self.frame_checkbox.grid(row=0, column=0, sticky=(tk.W, tk.E))
        
        # Frame bên phải cho xem trước
        frame_xem_truoc = ttk.LabelFrame(frame_chon_cot_chinh, text="Xem trước dữ liệu", padding="5")
        frame_xem_truoc.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S))
        frame_xem_truoc.columnconfigure(0, weight=1)
        frame_xem_truoc.rowconfigure(1, weight=1)
        
        # Label thứ tự cột
        self.lbl_thu_tu_cot = ttk.Label(frame_xem_truoc, text="Thứ tự cột:", font=("Arial", 10, "bold"))
        self.lbl_thu_tu_cot.grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        
        # Text widget cho xem trước
        self.text_xem_truoc = tk.Text(frame_xem_truoc, height=8, width=30, wrap=tk.WORD)
        scrollbar_xem_truoc = ttk.Scrollbar(frame_xem_truoc, orient="vertical", command=self.text_xem_truoc.yview)
        self.text_xem_truoc.configure(yscrollcommand=scrollbar_xem_truoc.set)
        
        self.text_xem_truoc.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar_xem_truoc.grid(row=1, column=1, sticky=(tk.N, tk.S))
        
        # 4. Thông tin bổ sung cho form phòng thi
        ttk.Label(main_frame, text="4. Thông tin bổ sung cho form phòng thi:", font=("Arial", 12, "bold")).grid(row=8, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 5))
        
        # Frame cho thông tin bổ sung
        frame_thong_tin = ttk.Frame(main_frame)
        frame_thong_tin.grid(row=9, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        frame_thong_tin.columnconfigure(1, weight=1)
        frame_thong_tin.columnconfigure(3, weight=1)
        
        # Dòng 1: Ban tổ chức và Điểm thi
        ttk.Label(frame_thong_tin, text="Ban tổ chức kỳ thi:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10), pady=2)
        self.combo_ban_to_chuc = ttk.Combobox(frame_thong_tin, values=["ASMO VIỆT NAM", "SEAMO VIỆT NAM", "IKSC VIỆT NAM", "IKLC VIỆT NAM"], state="readonly", width=27)
        self.combo_ban_to_chuc.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 20), pady=2)
        self.combo_ban_to_chuc.bind('<<ComboboxSelected>>', self.cap_nhat_thoi_gian_theo_ban_to_chuc)
        
        ttk.Label(frame_thong_tin, text="Điểm thi:").grid(row=0, column=2, sticky=tk.W, padx=(0, 10), pady=2)
        self.combo_diem_thi = ttk.Combobox(frame_thong_tin, values=self.danh_sach_dia_diem, width=27)
        self.combo_diem_thi.grid(row=0, column=3, sticky=(tk.W, tk.E), pady=2)
        self.combo_diem_thi.bind('<<ComboboxSelected>>', self.cap_nhat_ma_phong_theo_dia_diem)
        
        # Dòng 2: Môn thi và Thời gian có mặt
        ttk.Label(frame_thong_tin, text="Môn thi:").grid(row=1, column=0, sticky=tk.W, padx=(0, 10), pady=2)
        self.combo_mon_thi = ttk.Combobox(frame_thong_tin, values=["Toán", "Khoa học", "Tiếng Anh"], state="readonly", width=27)
        self.combo_mon_thi.set("Toán")
        self.combo_mon_thi.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(0, 20), pady=2)
        self.combo_mon_thi.bind('<<ComboboxSelected>>', self.cap_nhat_theo_mon_thi)
        
        ttk.Label(frame_thong_tin, text="Thí sinh có mặt lúc:").grid(row=1, column=2, sticky=tk.W, padx=(0, 10), pady=2)
        self.combo_thoi_gian = ttk.Combobox(frame_thong_tin, values=[], state="readonly", width=27)
        self.combo_thoi_gian.grid(row=1, column=3, sticky=(tk.W, tk.E), pady=2)
        
        # 5. Chọn ảnh cho danh sách
        ttk.Label(main_frame, text="5. Chọn ảnh cho danh sách:", font=("Arial", 12, "bold")).grid(row=10, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 5))
        
        # Frame chọn ảnh cho mục 5 - Danh sách để in
        frame_anh_de_in = ttk.LabelFrame(main_frame, text="🖼️ Ảnh cho danh sách để in", padding="10")
        frame_anh_de_in.grid(row=11, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        frame_anh_de_in.columnconfigure(1, weight=1)
        frame_anh_de_in.columnconfigure(3, weight=1)
        
        # Ảnh bên trái
        ttk.Label(frame_anh_de_in, text="Ảnh góc trái:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10), pady=2)
        self.lbl_anh_trai = ttk.Label(frame_anh_de_in, text="Chưa chọn ảnh", foreground="red")
        self.lbl_anh_trai.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 20), pady=2)
        ttk.Button(frame_anh_de_in, text="Chọn ảnh trái", command=self.chon_anh_trai).grid(row=0, column=2, padx=(0, 10), pady=2)
        
        # Ảnh bên phải
        ttk.Label(frame_anh_de_in, text="Ảnh góc phải:").grid(row=0, column=3, sticky=tk.W, padx=(0, 10), pady=2)
        self.lbl_anh_phai = ttk.Label(frame_anh_de_in, text="Chưa chọn ảnh", foreground="red")
        self.lbl_anh_phai.grid(row=0, column=4, sticky=(tk.W, tk.E), pady=2)
        ttk.Button(frame_anh_de_in, text="Chọn ảnh phải", command=self.chon_anh_phai).grid(row=0, column=5, padx=(0, 10), pady=2)
        
        # Nút xóa ảnh
        ttk.Button(frame_anh_de_in, text="🗑️ Xóa tất cả ảnh", command=self.xoa_tat_ca_anh).grid(row=1, column=0, columnspan=6, pady=(10, 0))
        
        # Điều chỉnh chiều cao ảnh
        ttk.Label(frame_anh_de_in, text="Chiều cao ảnh (cm):").grid(row=2, column=0, sticky=tk.W, padx=(0, 10), pady=(10, 0))
        self.entry_chieu_cao_anh = ttk.Entry(frame_anh_de_in, width=10)
        self.entry_chieu_cao_anh.insert(0, "2.0")
        self.entry_chieu_cao_anh.grid(row=2, column=1, sticky=tk.W, padx=(0, 20), pady=(10, 0))
        self.entry_chieu_cao_anh.bind('<KeyRelease>', self.cap_nhat_chieu_cao_anh)
        
        ttk.Label(frame_anh_de_in, text="💡 Mặc định: 2.0cm. Phạm vi: 1.0 - 5.0cm", font=("Arial", 9), foreground="blue").grid(row=2, column=2, columnspan=4, sticky=tk.W, padx=(10, 0), pady=(10, 0))
        
        # Frame chọn ảnh cho danh sách trường (chỉ offline)
        self.frame_anh_truong = ttk.LabelFrame(main_frame, text="🖼️ Ảnh cho danh sách trường (chỉ offline)", padding="10")
        self.frame_anh_truong.grid(row=12, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        self.frame_anh_truong.columnconfigure(1, weight=1)
        self.frame_anh_truong.columnconfigure(3, weight=1)
        
        # Ảnh bên trái cho trường
        ttk.Label(self.frame_anh_truong, text="Ảnh góc trái:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10), pady=2)
        self.lbl_anh_trai_truong = ttk.Label(self.frame_anh_truong, text="Chưa chọn ảnh", foreground="red")
        self.lbl_anh_trai_truong.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 20), pady=2)
        ttk.Button(self.frame_anh_truong, text="Chọn ảnh trái", command=self.chon_anh_trai_truong).grid(row=0, column=2, padx=(0, 10), pady=2)
        
        # Ảnh bên phải cho trường
        ttk.Label(self.frame_anh_truong, text="Ảnh góc phải:").grid(row=0, column=3, sticky=tk.W, padx=(0, 10), pady=2)
        self.lbl_anh_phai_truong = ttk.Label(self.frame_anh_truong, text="Chưa chọn ảnh", foreground="red")
        self.lbl_anh_phai_truong.grid(row=0, column=4, sticky=(tk.W, tk.E), padx=(0, 10), pady=2)
        ttk.Button(self.frame_anh_truong, text="Chọn ảnh phải", command=self.chon_anh_phai_truong).grid(row=0, column=5, padx=(0, 10), pady=2)
        
        # Nút xóa ảnh cho trường
        ttk.Button(self.frame_anh_truong, text="🗑️ Xóa tất cả ảnh", command=self.xoa_tat_ca_anh_truong).grid(row=1, column=0, columnspan=6, pady=(10, 0))
        
        # 6. Cài đặt chia phòng
        ttk.Label(main_frame, text="6. Cài đặt chia phòng thi:", font=("Arial", 12, "bold")).grid(row=13, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 5))
        
        # Frame cho cài đặt chia phòng
        frame_cai_dat = ttk.Frame(main_frame)
        frame_cai_dat.grid(row=14, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        
        # Dòng 1: Cài đặt cơ bản
        ttk.Label(frame_cai_dat, text="Số thí sinh mỗi phòng (mặc định):").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        self.entry_so_hs = ttk.Entry(frame_cai_dat, width=10)
        self.entry_so_hs.insert(0, "30")
        self.entry_so_hs.grid(row=0, column=1, sticky=tk.W, padx=(0, 20))
        
        ttk.Label(frame_cai_dat, text="Tên phòng bắt đầu:").grid(row=0, column=2, sticky=tk.W, padx=(0, 10))
        self.entry_ten_phong = ttk.Entry(frame_cai_dat, width=15)
        self.entry_ten_phong.insert(0, "Phòng")
        self.entry_ten_phong.grid(row=0, column=3, sticky=tk.W)
        
        # Dòng 2: Nút quản lý cấu hình phòng
        ttk.Button(frame_cai_dat, text="⚙️ Thiết lập số lượng từng phòng", command=self.thiet_lap_phong_rieng).grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=(10, 0))
        ttk.Label(frame_cai_dat, text="(Tùy chọn: Thiết lập số lượng thí sinh khác nhau cho từng phòng)", font=("Arial", 9), foreground="blue").grid(row=1, column=2, columnspan=2, sticky=tk.W, padx=(10, 0), pady=(10, 0))
        
        # Dòng 3: Nút lưu/tải cấu hình
        frame_cau_hinh = ttk.Frame(frame_cai_dat)
        frame_cau_hinh.grid(row=2, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=(5, 0))
        
        ttk.Button(frame_cau_hinh, text="💾 Lưu cấu hình phòng", command=self.luu_cau_hinh_file).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(frame_cau_hinh, text="📂 Tải cấu hình phòng", command=self.tai_cau_hinh_file).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(frame_cau_hinh, text="📋 Nhập từ dữ liệu", command=self.nhap_cau_hinh_tu_du_lieu).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(frame_cau_hinh, text="🗑️ Xóa cấu hình", command=self.xoa_cau_hinh).pack(side=tk.LEFT)
        
        # Hiển thị đường dẫn thư mục config
        config_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config")
        ttk.Label(frame_cau_hinh, text=f"📁 Thư mục cấu hình: {config_dir}", 
                 font=("Arial", 9), foreground="gray").pack(side=tk.RIGHT, padx=(10, 0))
        
        # 7. Thực hiện chia phòng thi
        ttk.Label(main_frame, text="7. Thực hiện chia phòng thi:", font=("Arial", 12, "bold")).grid(row=15, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 5))
        
        frame_nut_chinh = ttk.Frame(main_frame)
        frame_nut_chinh.grid(row=16, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=10)
        
        # Tạo 3 cột chính
        frame_cot1 = ttk.LabelFrame(frame_nut_chinh, text="📋 Chia phòng", padding="10")
        frame_cot1.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N), padx=(0, 10))
        
        frame_cot2 = ttk.LabelFrame(frame_nut_chinh, text="📊 Xuất dữ liệu", padding="10")
        frame_cot2.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N), padx=(0, 10))
        
        frame_cot3 = ttk.LabelFrame(frame_nut_chinh, text="🖨️ In ấn", padding="10")
        frame_cot3.grid(row=0, column=2, sticky=(tk.W, tk.E, tk.N))
        
        # Cột 1: Chia phòng (1, 1b, 1c)
        ttk.Button(frame_cot1, text="1. Thực hiện chia phòng thi", command=self.thuc_hien_chia_phong, width=30).pack(fill=tk.X, pady=2)
        ttk.Button(frame_cot1, text="1b. Bổ sung thí sinh", command=self.bo_sung_thi_sinh, width=30).pack(fill=tk.X, pady=2)
        ttk.Button(frame_cot1, text="1c. Chia lại từ phòng X", command=self.chia_lai_tu_phong_x, width=30).pack(fill=tk.X, pady=2)
        
        # Cột 2: Xuất dữ liệu (2, 3, 4)
        ttk.Button(frame_cot2, text="2. Preview Danh Sách", command=self.preview_form_phong_thi, width=30).pack(fill=tk.X, pady=2)
        ttk.Button(frame_cot2, text="3. Xuất file Excel", command=self.xuat_file_ket_qua, width=30).pack(fill=tk.X, pady=2)
        ttk.Button(frame_cot2, text="4. Xuất DS chia phòng", command=self.xuat_danh_sach_chia_phong, width=30).pack(fill=tk.X, pady=2)
        
        # Cột 3: In ấn (5, 6)
        ttk.Button(frame_cot3, text="5. Xuất DS để in", command=self.xuat_form_phong_thi, width=30).pack(fill=tk.X, pady=2)
        self.btn_xuat_cho_truong = ttk.Button(frame_cot3, text="6. Xuất DS cho trường (offline)", command=self.xuat_danh_sach_cho_truong, width=30)
        self.btn_xuat_cho_truong.pack(fill=tk.X, pady=2)
        
        # Hướng dẫn thứ tự sử dụng
        ttk.Label(main_frame, text="💡 Thứ tự: 1 → (1b hoặc 1c nếu cần) → 2 → 3 → 4 → 5 → 6 (offline)", 
                 font=("Arial", 9), foreground="blue").grid(row=17, column=0, columnspan=2, pady=(5, 10))
        
        # Text widget hiển thị kết quả (bên phải)
        self.text_ket_qua = tk.Text(self.frame_ket_qua, height=40, wrap=tk.WORD, font=("Consolas", 9))
        scrollbar_ket_qua = ttk.Scrollbar(self.frame_ket_qua, orient="vertical", command=self.text_ket_qua.yview)
        self.text_ket_qua.configure(yscrollcommand=scrollbar_ket_qua.set)
        
        self.text_ket_qua.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar_ket_qua.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # Cấu hình grid cho text widget kết quả
        self.frame_ket_qua.columnconfigure(0, weight=1)
        self.frame_ket_qua.rowconfigure(0, weight=1)
    
    def cap_nhat_che_do_thi(self):
        """Cập nhật chế độ thi và mô tả"""
        self.che_do_thi = self.var_che_do.get()
        
        if self.che_do_thi == "offline":
            self.lbl_mo_ta_che_do.config(text="💡 Thi Offline: Có giám thị, chữ ký thủ công")
            self.frame_chon_cot_che_do.grid_remove()  # Ẩn frame chọn cột chế độ thi
            # Hiện nút xuất danh sách cho trường
            if hasattr(self, 'btn_xuat_cho_truong'):
                self.btn_xuat_cho_truong.grid(row=0, column=5)
            # Hiện frame chọn ảnh cho trường
            if hasattr(self, 'frame_anh_truong'):
                self.frame_anh_truong.grid(row=17, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
        elif self.che_do_thi == "online":
            self.lbl_mo_ta_che_do.config(text="💡 Thi Online: Thời gian gọi, điểm danh, nộp bài online")
            self.frame_chon_cot_che_do.grid_remove()  # Ẩn frame chọn cột chế độ thi
            # Ẩn nút xuất danh sách cho trường
            if hasattr(self, 'btn_xuat_cho_truong'):
                self.btn_xuat_cho_truong.grid_remove()
            # Ẩn frame chọn ảnh cho trường
            if hasattr(self, 'frame_anh_truong'):
                self.frame_anh_truong.grid_remove()
        elif self.che_do_thi == "hybrid":
            self.lbl_mo_ta_che_do.config(text="💡 Thi Hỗn hợp: Kết hợp cả offline và online")
            self.frame_chon_cot_che_do.grid()  # Hiện frame chọn cột chế độ thi
            # Ẩn nút xuất danh sách cho trường
            if hasattr(self, 'btn_xuat_cho_truong'):
                self.btn_xuat_cho_truong.grid_remove()
            # Ẩn frame chọn ảnh cho trường
            if hasattr(self, 'frame_anh_truong'):
                self.frame_anh_truong.grid_remove()
            # Cập nhật danh sách cột cho combo box
            self.cap_nhat_danh_sach_cot_che_do()
        
        # Cập nhật panel xem trước nếu có dữ liệu
        if hasattr(self, 'text_xem_truoc'):
            self.cap_nhat_panel_xem_truoc()
    
    def cap_nhat_thoi_gian_theo_ban_to_chuc(self, event=None):
        """Cập nhật danh sách thời gian có mặt và môn thi dựa trên ban tổ chức được chọn"""
        ban_to_chuc = self.combo_ban_to_chuc.get()
        
        # Nếu có dữ liệu từ file giờ thi
        if self.thong_tin_gio_thi:
            # Lấy danh sách môn thi từ ban tổ chức này
            mon_thi_set = set()
            for (ban, mon), gio_list in self.thong_tin_gio_thi.items():
                if ban == ban_to_chuc:
                    mon_thi_set.add(mon)
            
            # Cập nhật combo môn thi
            if mon_thi_set:
                self.combo_mon_thi['values'] = sorted(list(mon_thi_set))
                # Tự động chọn môn đầu tiên
                self.combo_mon_thi.set(sorted(list(mon_thi_set))[0])
            else:
                # Không có dữ liệu từ file, dùng mặc định
                self.combo_mon_thi['values'] = ["Toán", "Khoa học", "Tiếng Anh"]
                self.combo_mon_thi.set("Toán")
            
            # Cập nhật giờ thi theo môn đã chọn
            self.cap_nhat_thoi_gian_theo_mon_thi()
        else:
            # Fallback về cách cũ nếu không có file giờ thi
            # Định nghĩa thời gian cho từng ban tổ chức
            thoi_gian_map = {
                "ASMO VIỆT NAM": ["8h30", "13h00", "15h00"],
                "SEAMO VIỆT NAM": ["8h30"],
                "IKSC VIỆT NAM": ["7h30", "9h30"],
                "IKLC VIỆT NAM": ["7h30", "9h30"]
            }
            
            # Định nghĩa môn thi cho từng ban tổ chức
            mon_thi_map = {
                "ASMO VIỆT NAM": ["Toán", "Khoa học", "Tiếng Anh"],
                "SEAMO VIỆT NAM": ["Toán"],
                "IKSC VIỆT NAM": ["Khoa học"],
                "IKLC VIỆT NAM": ["Tiếng Anh"]
            }
            
            # Cập nhật danh sách thời gian
            if ban_to_chuc in thoi_gian_map:
                self.combo_thoi_gian['values'] = thoi_gian_map[ban_to_chuc]
                # Tự động chọn thời gian đầu tiên
                if len(thoi_gian_map[ban_to_chuc]) > 0:
                    self.combo_thoi_gian.set(thoi_gian_map[ban_to_chuc][0])
            else:
                self.combo_thoi_gian['values'] = []
                self.combo_thoi_gian.set('')
            
            # Cập nhật danh sách môn thi
            if ban_to_chuc in mon_thi_map:
                self.combo_mon_thi['values'] = mon_thi_map[ban_to_chuc]
                # Tự động chọn môn đầu tiên
                if len(mon_thi_map[ban_to_chuc]) > 0:
                    self.combo_mon_thi.set(mon_thi_map[ban_to_chuc][0])
            else:
                self.combo_mon_thi['values'] = ["Toán", "Khoa học", "Tiếng Anh"]
                self.combo_mon_thi.set("Toán")
    
    def cap_nhat_theo_mon_thi(self, event=None):
        """Cập nhật giờ thi và mã phòng khi chọn môn thi"""
        # Cập nhật giờ thi
        self.cap_nhat_thoi_gian_theo_mon_thi(event)
        
        # Cập nhật mã phòng tự động
        self.cap_nhat_ma_phong_theo_dia_diem(event)
    
    def cap_nhat_ma_phong_theo_dia_diem(self, event=None):
        """Cập nhật mã phòng tự động khi chọn địa điểm thi hoặc môn thi"""
        # Kiểm tra các widget cần thiết đã tồn tại
        if not hasattr(self, 'combo_diem_thi') or not hasattr(self, 'combo_mon_thi') or not hasattr(self, 'entry_ten_phong'):
            return
        
        dia_diem = self.combo_diem_thi.get()
        mon_thi = self.combo_mon_thi.get()
        
        # Nếu đã chọn cả địa điểm và môn thi
        if dia_diem and mon_thi and hasattr(self, 'ma_dia_diem') and self.ma_dia_diem:
            if dia_diem in self.ma_dia_diem:
                ma_dia_diem = self.ma_dia_diem[dia_diem]
                
                # Tạo mã môn thi
                ma_mon = ""
                if mon_thi == "Toán":
                    ma_mon = "PT"
                elif mon_thi == "Khoa học":
                    ma_mon = "PKH"
                elif mon_thi == "Tiếng Anh":
                    ma_mon = "PTA"
                
                # Tạo mã phòng theo format: [Mã môn]-[Mã địa điểm]
                ma_phong = f"{ma_mon}-{ma_dia_diem}"
                
                # Cập nhật tên phòng với mã phòng
                self.entry_ten_phong.delete(0, tk.END)
                self.entry_ten_phong.insert(0, ma_phong)
                print(f"✅ Tự động cập nhật mã phòng: {ma_phong} cho {dia_diem} - {mon_thi}")
    
    def cap_nhat_thoi_gian_theo_mon_thi(self, event=None):
        """Cập nhật danh sách giờ thi khi chọn môn thi"""
        ban_to_chuc = self.combo_ban_to_chuc.get()
        mon_thi = self.combo_mon_thi.get()
        
        # Nếu có dữ liệu từ file giờ thi
        if self.thong_tin_gio_thi:
            # Tìm giờ thi tương ứng
            key = (ban_to_chuc, mon_thi)
            if key in self.thong_tin_gio_thi:
                gio_thi_list = self.thong_tin_gio_thi[key]
                self.combo_thoi_gian['values'] = gio_thi_list
                # Tự động chọn giờ đầu tiên
                if gio_thi_list:
                    self.combo_thoi_gian.set(gio_thi_list[0])
            else:
                # Không tìm thấy giờ thi cho tổ hợp này
                self.combo_thoi_gian['values'] = []
                self.combo_thoi_gian.set('')
                print(f"⚠️ Không tìm thấy giờ thi cho: {ban_to_chuc} - {mon_thi}")
        else:
            # Không có file giờ thi, giữ nguyên giờ hiện tại
            pass
    
    def cap_nhat_chieu_cao_anh(self, event=None):
        """Cập nhật chiều cao ảnh từ entry với validation"""
        try:
            # Lấy giá trị từ entry
            value = self.entry_chieu_cao_anh.get().strip()
            
            if value:
                chieu_cao = float(value)
                
                # Kiểm tra phạm vi hợp lệ
                if 1.0 <= chieu_cao <= 5.0:
                    self.chieu_cao_anh = chieu_cao
                    print(f"✅ Đã cập nhật chiều cao ảnh: {chieu_cao}cm")
                else:
                    # Nếu nằm ngoài phạm vi, đặt về giá trị mặc định và cập nhật entry
                    self.chieu_cao_anh = 2.0
                    self.entry_chieu_cao_anh.delete(0, tk.END)
                    self.entry_chieu_cao_anh.insert(0, "2.0")
                    print(f"⚠️ Chiều cao ảnh phải trong phạm vi 1.0 - 5.0cm. Đã đặt về mặc định: 2.0cm")
            else:
                # Nếu để trống, dùng giá trị mặc định
                self.chieu_cao_anh = 2.0
                self.entry_chieu_cao_anh.delete(0, tk.END)
                self.entry_chieu_cao_anh.insert(0, "2.0")
                print(f"✅ Đã đặt lại chiều cao ảnh về mặc định: 2.0cm")
                
        except ValueError:
            # Nếu không phải số, đặt về giá trị mặc định và cập nhật entry
            self.chieu_cao_anh = 2.0
            self.entry_chieu_cao_anh.delete(0, tk.END)
            self.entry_chieu_cao_anh.insert(0, "2.0")
            print(f"⚠️ Vui lòng nhập số hợp lệ cho chiều cao ảnh. Đã đặt về mặc định: 2.0cm")
    
    def cap_nhat_danh_sach_cot_che_do(self):
        """Cập nhật danh sách cột cho combo box chế độ thi"""
        if self.df_goc is not None:
            self.combo_cot_che_do['values'] = list(self.df_goc.columns)
            if self.cot_che_do_thi and self.cot_che_do_thi in self.df_goc.columns:
                self.combo_cot_che_do.set(self.cot_che_do_thi)
            else:
                self.combo_cot_che_do.set("")
        else:
            self.combo_cot_che_do['values'] = []
            self.combo_cot_che_do.set("")
    
    def cap_nhat_cot_che_do_thi(self, event=None):
        """Cập nhật cột chế độ thi được chọn"""
        self.cot_che_do_thi = self.combo_cot_che_do.get()
        print(f"🔧 Đã chọn cột chế độ thi: {self.cot_che_do_thi}")
        
        # Cập nhật panel xem trước
        if hasattr(self, 'text_xem_truoc'):
            self.cap_nhat_panel_xem_truoc()
        
    def thiet_lap_phong_rieng(self):
        """Mở cửa sổ thiết lập số lượng thí sinh cho từng phòng riêng biệt"""
        if self.df_goc is None:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn file Excel trước!")
            return
        
        # Tạo cửa sổ thiết lập phòng
        window_phong = tk.Toplevel(self.root)
        window_phong.title("Thiết lập số lượng thí sinh cho từng phòng")
        window_phong.geometry("800x600")
        window_phong.grab_set()  # Modal window
        
        # Frame chính
        main_frame_phong = ttk.Frame(window_phong, padding="10")
        main_frame_phong.pack(fill=tk.BOTH, expand=True)
        
        # Tiêu đề
        ttk.Label(main_frame_phong, text="THIẾT LẬP SỐ LƯỢNG THÍ SINH CHO TỪNG PHÒNG", font=("Arial", 14, "bold")).pack(pady=(0, 10))
        
        # Hướng dẫn
        ttk.Label(main_frame_phong, text="💡 Hướng dẫn: Nhập số lượng thí sinh cho từng phòng. Để trống sẽ dùng số mặc định.", 
                 font=("Arial", 10), foreground="blue").pack(pady=(0, 10))
        
        # Frame cho danh sách phòng
        frame_danh_sach = ttk.LabelFrame(main_frame_phong, text="Danh sách phòng", padding="10")
        frame_danh_sach.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Tạo canvas và scrollbar cho danh sách phòng
        canvas_phong = tk.Canvas(frame_danh_sach)
        scrollbar_phong = ttk.Scrollbar(frame_danh_sach, orient="vertical", command=canvas_phong.yview)
        scrollable_frame_phong = ttk.Frame(canvas_phong)
        
        scrollable_frame_phong.bind(
            "<Configure>",
            lambda e: canvas_phong.configure(scrollregion=canvas_phong.bbox("all"))
        )
        
        canvas_phong.create_window((0, 0), window=scrollable_frame_phong, anchor="nw")
        canvas_phong.configure(yscrollcommand=scrollbar_phong.set)
        
        # Tính toán số phòng cần thiết
        so_thi_sinh_mac_dinh = int(self.entry_so_hs.get() or 30)
        ten_phong_goc = self.entry_ten_phong.get().strip() or "Phòng"
        so_phong_can_thiet = (len(self.df_goc) + so_thi_sinh_mac_dinh - 1) // so_thi_sinh_mac_dinh
        
        # Tạo các entry cho từng phòng
        self.entry_phong_vars = {}
        for i in range(so_phong_can_thiet):
            phong_so = i + 1
            ten_phong = f"{ten_phong_goc} {phong_so}"
            
            # Frame cho mỗi phòng
            frame_phong = ttk.Frame(scrollable_frame_phong)
            frame_phong.pack(fill=tk.X, pady=2)
            
            # Label tên phòng
            ttk.Label(frame_phong, text=f"{ten_phong}:", width=15, anchor="w").pack(side=tk.LEFT, padx=(0, 10))
            
            # Entry số lượng thí sinh
            var_phong = tk.StringVar()
            # Nếu đã có cấu hình trước đó, sử dụng lại
            if ten_phong in self.cau_hinh_phong:
                var_phong.set(str(self.cau_hinh_phong[ten_phong]))
            else:
                var_phong.set("")  # Để trống = dùng mặc định
            
            entry_phong = ttk.Entry(frame_phong, textvariable=var_phong, width=10)
            entry_phong.pack(side=tk.LEFT, padx=(0, 10))
            
            # Label đơn vị
            ttk.Label(frame_phong, text="thí sinh").pack(side=tk.LEFT, padx=(0, 20))
            
            # Label gợi ý
            ttk.Label(frame_phong, text=f"(Mặc định: {so_thi_sinh_mac_dinh})", 
                     font=("Arial", 9), foreground="gray").pack(side=tk.LEFT)
            
            self.entry_phong_vars[ten_phong] = var_phong
        
        # Pack canvas và scrollbar
        canvas_phong.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_phong.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Frame nút điều khiển
        frame_nut_phong = ttk.Frame(main_frame_phong)
        frame_nut_phong.pack(fill=tk.X, pady=(10, 0))
        
        # Nút lưu cấu hình
        ttk.Button(frame_nut_phong, text="💾 Lưu cấu hình", 
                  command=lambda: self.luu_cau_hinh_phong(window_phong)).pack(side=tk.LEFT, padx=(0, 10))
        
        # Nút reset về mặc định
        ttk.Button(frame_nut_phong, text="🔄 Reset về mặc định", 
                  command=self.reset_cau_hinh_phong).pack(side=tk.LEFT, padx=(0, 10))
        
        # Nút đóng
        ttk.Button(frame_nut_phong, text="❌ Đóng", 
                  command=window_phong.destroy).pack(side=tk.RIGHT)
    
    def luu_cau_hinh_phong(self, window):
        """Lưu cấu hình phòng thi"""
        try:
            # Xóa cấu hình cũ
            self.cau_hinh_phong = {}
            
            # Lưu cấu hình mới
            for ten_phong, var in self.entry_phong_vars.items():
                gia_tri = var.get().strip()
                if gia_tri:  # Nếu có nhập giá trị
                    try:
                        so_luong = int(gia_tri)
                        if so_luong > 0:
                            self.cau_hinh_phong[ten_phong] = so_luong
                        else:
                            messagebox.showwarning("Cảnh báo", f"Số lượng thí sinh phải lớn hơn 0 cho {ten_phong}")
                            return
                    except ValueError:
                        messagebox.showwarning("Cảnh báo", f"Số lượng thí sinh phải là số nguyên cho {ten_phong}")
                        return
            
            # Hiển thị thông báo thành công
            so_phong_da_cau_hinh = len(self.cau_hinh_phong)
            if so_phong_da_cau_hinh > 0:
                messagebox.showinfo("Thành công", f"Đã lưu cấu hình cho {so_phong_da_cau_hinh} phòng!\nCác phòng khác sẽ dùng số lượng mặc định.")
            else:
                messagebox.showinfo("Thông báo", "Tất cả phòng sẽ dùng số lượng mặc định.")
            
            # Đóng cửa sổ
            window.destroy()
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể lưu cấu hình: {str(e)}")
    
    def reset_cau_hinh_phong(self):
        """Reset cấu hình phòng về mặc định"""
        for var in self.entry_phong_vars.values():
            var.set("")
        self.cau_hinh_phong = {}
        messagebox.showinfo("Thông báo", "Đã reset cấu hình về mặc định!")
    
    def chia_phong_voi_so_luong_khac_nhau(self, tong_so_hoc_sinh, ten_phong_goc):
        """Chia phòng với số lượng khác nhau cho từng phòng"""
        phong_thi = []
        vi_tri_hien_tai = 0
        so_phong = 1
        
        # Debug: In ra cấu hình hiện tại
        print("=== DEBUG: Cấu hình phòng ===")
        print(f"Tên phòng gốc: '{ten_phong_goc}'")
        print(f"Số lượng cấu hình: {len(self.cau_hinh_phong)}")
        
        # Tự động phát hiện format tên phòng từ cấu hình
        ten_phong_mau = None
        co_so_0_dau = False
        if self.cau_hinh_phong:
            # Lấy tên phòng đầu tiên làm mẫu
            ten_phong_mau = list(self.cau_hinh_phong.keys())[0]
            print(f"Tên phòng mẫu: '{ten_phong_mau}'")
            
            # Phát hiện format số (01, 02 hay 1, 2)
            if "01" in ten_phong_mau or "02" in ten_phong_mau:
                co_so_0_dau = True
                print("Format số: 01, 02, 03...")
            else:
                print("Format số: 1, 2, 3...")
            
            # Trích xuất prefix (phần trước số)
            import re
            match = re.match(r'(.+?)\s*(\d+)$', ten_phong_mau)
            if match:
                ten_phong_goc = match.group(1).strip()
                print(f"Prefix tự động: '{ten_phong_goc}'")
        
        for phong, sl in self.cau_hinh_phong.items():
            print(f"  '{phong}': {sl}")
        print("="*40)
        
        while vi_tri_hien_tai < tong_so_hoc_sinh:
            # Tạo tên phòng với format phù hợp
            if co_so_0_dau:
                ten_phong = f"{ten_phong_goc} {so_phong:02d}"
            else:
                ten_phong = f"{ten_phong_goc} {so_phong}"
            
            # Lấy số lượng thí sinh cho phòng này
            if ten_phong in self.cau_hinh_phong:
                so_hoc_sinh_phong = self.cau_hinh_phong[ten_phong]
                print(f"✅ Phòng '{ten_phong}': Dùng cấu hình {so_hoc_sinh_phong}")
            else:
                so_hoc_sinh_phong = self.so_thi_sinh_moi_phong
                print(f"⚠️ Phòng '{ten_phong}': Không tìm thấy cấu hình, dùng mặc định {so_hoc_sinh_phong}")
                # Debug: In ra các key trong cấu hình để so sánh
                if self.cau_hinh_phong and so_phong <= 3:  # Chỉ in 3 phòng đầu
                    print(f"   DEBUG: Tìm kiếm '{ten_phong}' (len={len(ten_phong)})")
                    for key in list(self.cau_hinh_phong.keys())[:3]:
                        print(f"   - Key: '{key}' (len={len(key)}) - Match: {key == ten_phong}")
                        # So sánh từng ký tự
                        if len(key) == len(ten_phong):
                            for i, (c1, c2) in enumerate(zip(key, ten_phong)):
                                if c1 != c2:
                                    print(f"     Khác biệt tại vị trí {i}: '{c1}' (ord={ord(c1)}) vs '{c2}' (ord={ord(c2)})")
            
            # Tính số thí sinh thực tế cho phòng này
            so_hoc_sinh_con_lai = tong_so_hoc_sinh - vi_tri_hien_tai
            so_hoc_sinh_thuc_te = min(so_hoc_sinh_phong, so_hoc_sinh_con_lai)
            
            # Gán phòng cho các thí sinh
            for i in range(so_hoc_sinh_thuc_te):
                phong_thi.append(ten_phong)
            
            vi_tri_hien_tai += so_hoc_sinh_thuc_te
            so_phong += 1
        
        return phong_thi
    
    def luu_cau_hinh_file(self):
        """Lưu cấu hình phòng thi ra file JSON"""
        if not self.cau_hinh_phong:
            messagebox.showwarning("Cảnh báo", "Chưa có cấu hình phòng nào để lưu!")
            return
        
        # Tạo thư mục config nếu chưa có
        config_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config")
        if not os.path.exists(config_dir):
            os.makedirs(config_dir)
        
        # Đường dẫn mặc định trong thư mục config
        default_path = os.path.join(config_dir, "cau_hinh_phong_thi.json")
        
        file_path = filedialog.asksaveasfilename(
            title="Lưu cấu hình phòng thi",
            defaultextension=".json",
            initialfile=default_path,
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                import json
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(self.cau_hinh_phong, f, ensure_ascii=False, indent=2)
                messagebox.showinfo("Thành công", f"Đã lưu cấu hình phòng thi: {os.path.basename(file_path)}")
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể lưu cấu hình: {str(e)}")
    
    def tai_cau_hinh_file(self):
        """Tải cấu hình phòng thi từ file JSON"""
        # Kiểm tra thư mục config trước
        config_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config")
        initial_dir = config_dir if os.path.exists(config_dir) else os.path.dirname(os.path.abspath(__file__))
        
        file_path = filedialog.askopenfilename(
            title="Tải cấu hình phòng thi",
            initialdir=initial_dir,
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                import json
                with open(file_path, 'r', encoding='utf-8') as f:
                    self.cau_hinh_phong = json.load(f)
                
                # Hiển thị thông báo thành công
                so_phong = len(self.cau_hinh_phong)
                messagebox.showinfo("Thành công", f"Đã tải cấu hình cho {so_phong} phòng từ: {os.path.basename(file_path)}")
                
                # Hiển thị cấu hình đã tải
                self.text_ket_qua.delete(1.0, tk.END)
                self.text_ket_qua.insert(tk.END, "=== CẤU HÌNH PHÒNG THI ĐÃ TẢI ===\n\n")
                for phong, so_hs in self.cau_hinh_phong.items():
                    self.text_ket_qua.insert(tk.END, f"{phong}: {so_hs} thí sinh\n")
                
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể tải cấu hình: {str(e)}")
    
    def xoa_cau_hinh(self):
        """Xóa cấu hình phòng thi hiện tại"""
        if not self.cau_hinh_phong:
            messagebox.showinfo("Thông báo", "Chưa có cấu hình phòng nào để xóa!")
            return
        
        if messagebox.askyesno("Xác nhận", "Bạn có chắc chắn muốn xóa cấu hình phòng thi hiện tại?"):
            self.cau_hinh_phong = {}
            messagebox.showinfo("Thành công", "Đã xóa cấu hình phòng thi!")
            
            # Cập nhật hiển thị
            self.text_ket_qua.delete(1.0, tk.END)
            self.text_ket_qua.insert(tk.END, "Đã xóa cấu hình phòng thi. Tất cả phòng sẽ dùng số lượng mặc định.")
    
    def nhap_cau_hinh_tu_du_lieu(self):
        """Nhập cấu hình phòng từ dữ liệu (text paste hoặc file)"""
        # Tạo cửa sổ mới với kích thước lớn hơn
        window = tk.Toplevel(self.root)
        window.title("Nhập cấu hình phòng từ dữ liệu")
        window.geometry("800x650")
        
        # Frame chính với scrollbar
        main_canvas = tk.Canvas(window)
        scrollbar_main = ttk.Scrollbar(window, orient="vertical", command=main_canvas.yview)
        main_frame = ttk.Frame(main_canvas, padding="10")
        
        main_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )
        
        main_canvas.create_window((0, 0), window=main_frame, anchor="nw")
        main_canvas.configure(yscrollcommand=scrollbar_main.set)
        
        main_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_main.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Hướng dẫn
        huong_dan = """📋 HƯỚNG DẪN NHẬP DỮ LIỆU:

1. Dán trực tiếp từ ảnh (OCR hoặc copy):
   Phòng thi số 01    25
   Phòng thi số 02    22
   Phòng thi số 03    23
   ...

2. Hoặc chọn file Excel/CSV có 2 cột: [Tên phòng] [Số lượng]

3. Nhấn "💾 Lưu cấu hình" để lưu"""
        
        ttk.Label(main_frame, text=huong_dan, font=("Arial", 10), justify=tk.LEFT, foreground="blue").pack(anchor=tk.W, pady=(0, 10))
        
        # Frame cho nút chọn file
        frame_chon_file = ttk.Frame(main_frame)
        frame_chon_file.pack(fill=tk.X, pady=(0, 10))
        
        # Khởi tạo text_input trước để dùng trong lambda
        text_input = None
        
        ttk.Button(frame_chon_file, text="📂 Chọn file Excel/CSV", 
                  command=lambda: self.chon_file_cau_hinh(text_input) if text_input else None).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Label(frame_chon_file, text="hoặc dán dữ liệu vào ô bên dưới:", font=("Arial", 9)).pack(side=tk.LEFT)
        
        # Text widget cho nhập liệu
        ttk.Label(main_frame, text="Dữ liệu cấu hình phòng:", font=("Arial", 10, "bold")).pack(anchor=tk.W, pady=(0, 5))
        
        frame_text = ttk.Frame(main_frame)
        frame_text.pack(fill=tk.X, pady=(0, 10))
        
        text_input = tk.Text(frame_text, height=12, width=80, wrap=tk.WORD)
        scrollbar_text = ttk.Scrollbar(frame_text, orient="vertical", command=text_input.yview)
        text_input.configure(yscrollcommand=scrollbar_text.set)
        
        text_input.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_text.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Separator
        ttk.Separator(main_frame, orient='horizontal').pack(fill=tk.X, pady=10)
        
        # Frame nút điều khiển - QUAN TRỌNG: Đặt ở cuối
        frame_nut = ttk.Frame(main_frame)
        frame_nut.pack(fill=tk.X, pady=(10, 0))
        
        # Các nút với kích thước rõ ràng
        btn_luu = ttk.Button(frame_nut, text="💾 Lưu cấu hình", 
                  command=lambda: self.xu_ly_du_lieu_cau_hinh(text_input.get("1.0", tk.END), window),
                  width=20)
        btn_luu.pack(side=tk.LEFT, padx=(0, 10))
        
        btn_xoa = ttk.Button(frame_nut, text="🗑️ Xóa", 
                  command=lambda: text_input.delete("1.0", tk.END),
                  width=15)
        btn_xoa.pack(side=tk.LEFT, padx=(0, 10))
        
        btn_huy = ttk.Button(frame_nut, text="❌ Hủy", 
                  command=window.destroy,
                  width=10)
        btn_huy.pack(side=tk.RIGHT)
        
        # Thêm ghi chú
        ttk.Label(main_frame, text="💡 Nhấn 'Lưu cấu hình' để lưu và đóng cửa sổ này", 
                 font=("Arial", 9, "bold"), foreground="green").pack(pady=(10, 0))
        
        # Đảm bảo cửa sổ luôn nằm trên cùng
        window.transient(self.root)
        window.grab_set()
        
        # Focus vào text input
        text_input.focus_set()
    
    def chon_file_cau_hinh(self, text_widget):
        """Chọn file Excel/CSV để nhập cấu hình"""
        file_path = filedialog.askopenfilename(
            title="Chọn file cấu hình phòng",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("CSV files", "*.csv"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                # Đọc file
                if file_path.endswith('.csv'):
                    df = pd.read_csv(file_path, header=None)
                else:
                    df = pd.read_excel(file_path, header=None, engine='openpyxl')
                
                # Chuyển đổi thành text
                text_widget.delete("1.0", tk.END)
                for _, row in df.iterrows():
                    if len(row) >= 2:
                        ten_phong = str(row[0]).strip()
                        so_luong = str(row[1]).strip()
                        text_widget.insert(tk.END, f"{ten_phong}\t{so_luong}\n")
                
                messagebox.showinfo("Thành công", f"Đã tải dữ liệu từ file: {os.path.basename(file_path)}")
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể đọc file: {str(e)}")
    
    def xu_ly_du_lieu_cau_hinh(self, du_lieu, window):
        """Xử lý dữ liệu cấu hình phòng từ text"""
        try:
            if not du_lieu.strip():
                messagebox.showwarning("Cảnh báo", "Vui lòng nhập dữ liệu cấu hình!")
                return
            
            # Xóa cấu hình cũ
            self.cau_hinh_phong = {}
            
            # Xử lý từng dòng
            dong_loi = []
            so_phong_thanh_cong = 0
            
            for i, dong in enumerate(du_lieu.strip().split('\n'), 1):
                dong = dong.strip()
                if not dong:
                    continue
                
                # Loại bỏ dấu hai chấm nếu có
                dong = dong.replace(':', ' ')
                
                # Thử nhiều cách tách dữ liệu
                # Cách 1: Tách bằng tab
                parts = dong.split('\t')
                if len(parts) < 2:
                    # Cách 2: Tách bằng nhiều space
                    parts = dong.split()
                
                if len(parts) >= 2:
                    try:
                        # Lấy số ở cuối (số lượng thí sinh)
                        so_luong = int(parts[-1])
                        # Phần còn lại là tên phòng
                        ten_phong = ' '.join(parts[:-1]).strip()
                        
                        if so_luong > 0:
                            self.cau_hinh_phong[ten_phong] = so_luong
                            so_phong_thanh_cong += 1
                            print(f"✅ Đã nhập: '{ten_phong}' = {so_luong}")
                        else:
                            dong_loi.append(f"Dòng {i}: Số lượng phải > 0")
                    except ValueError:
                        dong_loi.append(f"Dòng {i}: Không thể chuyển đổi số lượng")
                else:
                    dong_loi.append(f"Dòng {i}: Định dạng không hợp lệ")
            
            # Hiển thị kết quả
            if so_phong_thanh_cong > 0:
                thong_bao = f"✅ Đã nhập cấu hình cho {so_phong_thanh_cong} phòng!"
                
                if dong_loi:
                    thong_bao += f"\n\n⚠️ Có {len(dong_loi)} dòng bị lỗi:\n" + "\n".join(dong_loi[:5])
                    if len(dong_loi) > 5:
                        thong_bao += f"\n... và {len(dong_loi) - 5} lỗi khác"
                
                messagebox.showinfo("Kết quả", thong_bao)
                
                # Hiển thị cấu hình đã nhập
                self.text_ket_qua.delete(1.0, tk.END)
                self.text_ket_qua.insert(tk.END, "=== CẤU HÌNH PHÒNG ĐÃ NHẬP ===\n\n")
                for phong, so_hs in self.cau_hinh_phong.items():
                    self.text_ket_qua.insert(tk.END, f"{phong}: {so_hs} thí sinh\n")
                
                window.destroy()
            else:
                messagebox.showerror("Lỗi", "Không thể nhập được phòng nào!\n\n" + "\n".join(dong_loi[:10]))
                
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể xử lý dữ liệu: {str(e)}")


    def chon_anh_trai(self):
        """Chọn ảnh góc trái"""
        file_path = filedialog.askopenfilename(
            title="Chọn ảnh góc trái",
            filetypes=[("Image files", "*.jpg *.jpeg *.png *.bmp *.gif"), ("All files", "*.*")]
        )
        
        if file_path:
            # Kiểm tra file có tồn tại không
            if not os.path.exists(file_path):
                messagebox.showerror("Lỗi", f"File ảnh không tồn tại: {file_path}")
                return
            
            self.duong_dan_anh_trai = file_path
            self.lbl_anh_trai.config(text=f"Đã chọn: {os.path.basename(file_path)}", foreground="green")
            messagebox.showinfo("Thành công", f"Đã chọn ảnh góc trái: {os.path.basename(file_path)}")
    
    def chon_anh_phai(self):
        """Chọn ảnh góc phải"""
        file_path = filedialog.askopenfilename(
            title="Chọn ảnh góc phải",
            filetypes=[("Image files", "*.jpg *.jpeg *.png *.bmp *.gif"), ("All files", "*.*")]
        )
        
        if file_path:
            # Kiểm tra file có tồn tại không
            if not os.path.exists(file_path):
                messagebox.showerror("Lỗi", f"File ảnh không tồn tại: {file_path}")
                return
            
            self.duong_dan_anh_phai = file_path
            self.lbl_anh_phai.config(text=f"Đã chọn: {os.path.basename(file_path)}", foreground="green")
            messagebox.showinfo("Thành công", f"Đã chọn ảnh góc phải: {os.path.basename(file_path)}")
    
    def xoa_tat_ca_anh(self):
        """Xóa tất cả ảnh đã chọn"""
        if not self.duong_dan_anh_trai and not self.duong_dan_anh_phai:
            messagebox.showinfo("Thông báo", "Chưa có ảnh nào để xóa!")
            return
        
        if messagebox.askyesno("Xác nhận", "Bạn có chắc chắn muốn xóa tất cả ảnh đã chọn?"):
            self.duong_dan_anh_trai = ""
            self.duong_dan_anh_phai = ""
            self.lbl_anh_trai.config(text="Chưa chọn ảnh", foreground="red")
            self.lbl_anh_phai.config(text="Chưa chọn ảnh", foreground="red")
            messagebox.showinfo("Thành công", "Đã xóa tất cả ảnh!")
    
    def chon_anh_trai_truong(self):
        """Chọn ảnh góc trái cho danh sách trường"""
        file_path = filedialog.askopenfilename(
            title="Chọn ảnh góc trái cho danh sách trường",
            filetypes=[("Image files", "*.jpg *.jpeg *.png *.bmp *.gif"), ("All files", "*.*")]
        )
        
        if file_path:
            # Kiểm tra file có tồn tại không
            if not os.path.exists(file_path):
                messagebox.showerror("Lỗi", f"File ảnh không tồn tại: {file_path}")
                return
            
            self.duong_dan_anh_trai_truong = file_path
            self.lbl_anh_trai_truong.config(text=f"Đã chọn: {os.path.basename(file_path)}", foreground="green")
            messagebox.showinfo("Thành công", f"Đã chọn ảnh góc trái cho danh sách trường: {os.path.basename(file_path)}")
    
    def chon_anh_phai_truong(self):
        """Chọn ảnh góc phải cho danh sách trường"""
        file_path = filedialog.askopenfilename(
            title="Chọn ảnh góc phải cho danh sách trường",
            filetypes=[("Image files", "*.jpg *.jpeg *.png *.bmp *.gif"), ("All files", "*.*")]
        )
        
        if file_path:
            # Kiểm tra file có tồn tại không
            if not os.path.exists(file_path):
                messagebox.showerror("Lỗi", f"File ảnh không tồn tại: {file_path}")
                return
            
            self.duong_dan_anh_phai_truong = file_path
            self.lbl_anh_phai_truong.config(text=f"Đã chọn: {os.path.basename(file_path)}", foreground="green")
            messagebox.showinfo("Thành công", f"Đã chọn ảnh góc phải cho danh sách trường: {os.path.basename(file_path)}")
    
    def xoa_tat_ca_anh_truong(self):
        """Xóa tất cả ảnh đã chọn cho danh sách trường"""
        if not self.duong_dan_anh_trai_truong and not self.duong_dan_anh_phai_truong:
            messagebox.showinfo("Thông báo", "Chưa có ảnh nào để xóa!")
            return
        
        if messagebox.askyesno("Xác nhận", "Bạn có chắc chắn muốn xóa tất cả ảnh đã chọn cho danh sách trường?"):
            self.duong_dan_anh_trai_truong = ""
            self.duong_dan_anh_phai_truong = ""
            self.lbl_anh_trai_truong.config(text="Chưa chọn ảnh", foreground="red")
            self.lbl_anh_phai_truong.config(text="Chưa chọn ảnh", foreground="red")
            messagebox.showinfo("Thành công", "Đã xóa tất cả ảnh cho danh sách trường!")

    def chon_file_excel(self):
        file_path = filedialog.askopenfilename(
            title="Chọn file Excel",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                # Kiểm tra file có tồn tại không
                if not os.path.exists(file_path):
                    messagebox.showerror("Lỗi", f"File không tồn tại: {file_path}")
                    return
                
                # Kiểm tra kích thước file
                file_size = os.path.getsize(file_path)
                if file_size == 0:
                    messagebox.showerror("Lỗi", "File Excel trống!")
                    return
                
                # Lưu đường dẫn file
                self.duong_dan_file_excel = file_path
                
                # Đọc danh sách tất cả các sheet trong file Excel
                try:
                    excel_file = pd.ExcelFile(file_path, engine='openpyxl')
                    self.danh_sach_sheet = excel_file.sheet_names
                except Exception as e1:
                    try:
                        # Thử với engine xlrd cho file .xls cũ
                        excel_file = pd.ExcelFile(file_path, engine='xlrd')
                        self.danh_sach_sheet = excel_file.sheet_names
                    except Exception as e2:
                        raise Exception(f"Không thể đọc file với openpyxl: {str(e1)}\nKhông thể đọc file với xlrd: {str(e2)}")
                
                self.lbl_file_path.config(text=f"Đã chọn: {os.path.basename(file_path)}", foreground="green")
                
                # Hiển thị thông tin file
                self.text_ket_qua.delete(1.0, tk.END)
                self.text_ket_qua.insert(tk.END, f"✅ Đã tải file thành công!\n")
                self.text_ket_qua.insert(tk.END, f"📁 File: {os.path.basename(file_path)}\n")
                self.text_ket_qua.insert(tk.END, f"📋 Số trang tính: {len(self.danh_sach_sheet)}\n")
                self.text_ket_qua.insert(tk.END, f"📝 Các trang tính: {', '.join(self.danh_sach_sheet)}\n\n")
                
                # Nếu chỉ có 1 sheet, tự động load luôn
                if len(self.danh_sach_sheet) == 1:
                    self.sheet_hien_tai = self.danh_sach_sheet[0]
                    self.text_ket_qua.insert(tk.END, f"🔄 Tự động load trang tính: {self.sheet_hien_tai}\n\n")
                    self.load_du_lieu_tu_sheet(self.sheet_hien_tai)
                else:
                    # Hiển thị combobox để chọn sheet
                    self.frame_chon_sheet.grid()
                    self.combo_chon_sheet['values'] = self.danh_sach_sheet
                    self.combo_chon_sheet.set('')  # Reset
                    self.text_ket_qua.insert(tk.END, "⚠️ Vui lòng chọn trang tính cần xử lý!\n\n")
                    
            except Exception as e:
                error_msg = f"Không thể đọc file Excel: {str(e)}"
                print(f"❌ {error_msg}")
                messagebox.showerror("Lỗi", error_msg)
                self.lbl_file_path.config(text="Lỗi khi đọc file", foreground="red")
                import traceback
                traceback.print_exc()
    
    def chon_sheet(self, event=None):
        """Xử lý khi user chọn sheet từ combobox"""
        sheet_name = self.combo_chon_sheet.get()
        if sheet_name:
            self.sheet_hien_tai = sheet_name
            self.text_ket_qua.insert(tk.END, f"🔄 Đang load trang tính: {sheet_name}...\n")
            self.load_du_lieu_tu_sheet(sheet_name)
    
    def load_du_lieu_tu_sheet(self, sheet_name):
        """Load dữ liệu từ sheet được chọn"""
        try:
            # Đọc dữ liệu từ sheet cụ thể, giữ định dạng text cho tất cả cột
            if self.duong_dan_file_excel.endswith('.xlsx'):
                self.df_goc = pd.read_excel(self.duong_dan_file_excel, sheet_name=sheet_name, engine='openpyxl', dtype=str)
            else:
                self.df_goc = pd.read_excel(self.duong_dan_file_excel, sheet_name=sheet_name, engine='xlrd', dtype=str)
            
            # Kiểm tra dữ liệu có rỗng không
            if self.df_goc.empty:
                messagebox.showwarning("Cảnh báo", f"Trang tính '{sheet_name}' không có dữ liệu!")
                return
            
            # Kiểm tra có ít nhất 1 cột không
            if len(self.df_goc.columns) == 0:
                messagebox.showwarning("Cảnh báo", f"Trang tính '{sheet_name}' không có cột dữ liệu!")
                return
            
            # Hiển thị các cột để chọn
            self.hien_thi_cac_cot()
            
            # Hiển thị thông tin sheet
            self.text_ket_qua.insert(tk.END, f"✅ Đã load trang tính '{sheet_name}' thành công!\n")
            self.text_ket_qua.insert(tk.END, f"📊 Số dòng dữ liệu: {len(self.df_goc)}\n")
            self.text_ket_qua.insert(tk.END, f"📋 Số cột: {len(self.df_goc.columns)}\n")
            self.text_ket_qua.insert(tk.END, f"📝 Các cột có sẵn: {', '.join(self.df_goc.columns.tolist())}\n\n")
            
        except Exception as e:
            error_msg = f"Không thể load trang tính '{sheet_name}': {str(e)}"
            print(f"❌ {error_msg}")
            messagebox.showerror("Lỗi", error_msg)
            import traceback
            traceback.print_exc()
                
    def hien_thi_cac_cot(self):
        # Xóa các checkbox cũ
        for widget in self.frame_checkbox.winfo_children():
            widget.destroy()
            
        # Tạo checkbox cho mỗi cột
        self.checkbox_vars = {}
        self.cac_cot_duoc_chon = []
        self.thu_tu_cot_duoc_chon = []  # Reset thứ tự cột
        
        # Danh sách cột ưu tiên hiển thị đầu tiên
        cot_uu_tien = ["FULL NAME", "SBD", "DOB", "KHỐI", "TRƯỜNG"]
        
        # Sắp xếp lại: Cột ưu tiên trước, sau đó các cột còn lại
        cot_sap_xep = []
        for cot in cot_uu_tien:
            if cot in self.df_goc.columns:
                cot_sap_xep.append(cot)
        
        # Thêm các cột còn lại (không nằm trong danh sách ưu tiên)
        for cot in self.df_goc.columns:
            if cot not in cot_uu_tien:
                cot_sap_xep.append(cot)
        
        row = 0
        col = 0
        for i, cot in enumerate(cot_sap_xep):
            var = tk.BooleanVar()
            self.checkbox_vars[cot] = var
            
            # Highlight các cột ưu tiên
            if cot in cot_uu_tien:
                cb = ttk.Checkbutton(
                    self.frame_checkbox, 
                    text=f"⭐ {cot}",  # Thêm dấu sao cho cột ưu tiên
                    variable=var,
                    command=lambda c=cot: self.cap_nhat_cot_duoc_chon(c)
                )
            else:
                cb = ttk.Checkbutton(
                    self.frame_checkbox, 
                    text=cot, 
                    variable=var,
                    command=lambda c=cot: self.cap_nhat_cot_duoc_chon(c)
                )
            
            cb.grid(row=row, column=col, sticky=tk.W, padx=(0, 20), pady=2)
            
            col += 1
            if col > 3:  # 4 cột mỗi hàng
                col = 0
                row += 1
                
        # Cập nhật panel xem trước sau khi tạo checkbox
        self.cap_nhat_panel_xem_truoc()
        
        # Cập nhật danh sách cột chế độ thi nếu đang ở chế độ hybrid
        if self.che_do_thi == "hybrid":
            self.cap_nhat_danh_sach_cot_che_do()
        
    def cap_nhat_cot_duoc_chon(self, cot_duoc_click=None):
        if cot_duoc_click is not None:
            # Nếu cột được check (True)
            if self.checkbox_vars[cot_duoc_click].get():
                # Thêm vào thứ tự nếu chưa có
                if cot_duoc_click not in self.thu_tu_cot_duoc_chon:
                    self.thu_tu_cot_duoc_chon.append(cot_duoc_click)
            else:
                # Nếu cột được uncheck (False), xóa khỏi thứ tự
                if cot_duoc_click in self.thu_tu_cot_duoc_chon:
                    self.thu_tu_cot_duoc_chon.remove(cot_duoc_click)
        
        # Cập nhật danh sách cột được chọn theo thứ tự click
        self.cac_cot_duoc_chon = [cot for cot in self.thu_tu_cot_duoc_chon if self.checkbox_vars[cot].get()]
        
        # Cập nhật panel xem trước
        self.cap_nhat_panel_xem_truoc()
        
    def cap_nhat_panel_xem_truoc(self):
        """Cập nhật panel xem trước bên phải"""
        if not hasattr(self, 'text_xem_truoc'):
            return
            
        # Xóa nội dung cũ
        self.text_xem_truoc.delete(1.0, tk.END)
        
        if not self.cac_cot_duoc_chon:
            self.text_xem_truoc.insert(tk.END, "Chưa chọn cột nào")
            return
            
        # Hiển thị thứ tự cột
        self.text_xem_truoc.insert(tk.END, "Thứ tự cột:\n")
        for i, cot in enumerate(self.cac_cot_duoc_chon, 1):
            self.text_xem_truoc.insert(tk.END, f"({i}) {cot}\n")
        
        # Hiển thị các cột bổ sung theo chế độ thi
        self.text_xem_truoc.insert(tk.END, "\nCột bổ sung:\n")
        if self.che_do_thi == "offline":
            self.text_xem_truoc.insert(tk.END, "• STT\n• Chữ ký\n")
        elif self.che_do_thi == "online":
            self.text_xem_truoc.insert(tk.END, "• STT\n• Thời gian gọi trước ngày thi\n")
            self.text_xem_truoc.insert(tk.END, "• Thời gian gọi trong ngày thi\n")
            self.text_xem_truoc.insert(tk.END, "• Điểm danh\n• Thời gian nộp bài\n• Ghi chú\n")
        elif self.che_do_thi == "hybrid":
            self.text_xem_truoc.insert(tk.END, "• STT\n• Chế độ thi (Offline/Online)\n")
            self.text_xem_truoc.insert(tk.END, "• Ghi chú (chung)\n")
            self.text_xem_truoc.insert(tk.END, "• Thời gian gọi trước (Online)\n")
            self.text_xem_truoc.insert(tk.END, "• Thời gian gọi trong ngày (Online)\n")
            self.text_xem_truoc.insert(tk.END, "• Điểm danh (Online)\n• Thời gian nộp bài (Online)\n• Chữ ký (Offline)\n")
            
            # Hiển thị thông tin cột chế độ thi được chọn
            if self.cot_che_do_thi:
                self.text_xem_truoc.insert(tk.END, f"\n📋 Cột chế độ thi: {self.cot_che_do_thi}\n")
                self.text_xem_truoc.insert(tk.END, "💡 Tiêu đề cột thông minh:\n")
                self.text_xem_truoc.insert(tk.END, "   - Cột có ghi chú (Online) chỉ dành cho thí sinh Online\n")
                self.text_xem_truoc.insert(tk.END, "   - Cột có ghi chú (Offline) chỉ dành cho thí sinh Offline\n")
                self.text_xem_truoc.insert(tk.END, "   - Cột (chung) dành cho cả hai loại\n")
                self.text_xem_truoc.insert(tk.END, "   - Giám thị dễ nhận biết cột nào cần điền\n")
            else:
                self.text_xem_truoc.insert(tk.END, "\n⚠️ Chưa chọn cột chế độ thi!\n")
        
        # Hiển thị dữ liệu mẫu nếu có file Excel
        if self.df_goc is not None:
            self.text_xem_truoc.insert(tk.END, "\nDữ liệu mẫu:\n")
            df_truoc = self.df_goc[self.cac_cot_duoc_chon]
            self.text_xem_truoc.insert(tk.END, df_truoc.head(5).to_string(index=False))
        
    def thuc_hien_chia_phong(self):
        if self.df_goc is None:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn file Excel trước!")
            return
            
        if not self.cac_cot_duoc_chon:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn ít nhất một cột dữ liệu!")
            return
        
        # Kiểm tra cột chế độ thi cho chế độ hybrid
        if self.che_do_thi == "hybrid":
            if not self.cot_che_do_thi:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn cột chứa thông tin chế độ thi (Offline/Online)!")
                return
            if self.cot_che_do_thi not in self.df_goc.columns:
                messagebox.showwarning("Cảnh báo", f"Cột '{self.cot_che_do_thi}' không tồn tại trong dữ liệu!")
                return
            
        try:
            # Lấy thông tin bổ sung - đảm bảo không None
            self.ban_to_chuc = self.combo_ban_to_chuc.get() or ""
            self.diem_thi = self.combo_diem_thi.get() or ""
            self.mon_thi = self.combo_mon_thi.get() or "Toán"
            self.thoi_gian_co_mat = self.combo_thoi_gian.get() or ""
            
            # Lấy số thí sinh mỗi phòng với validation
            so_hs_text = self.entry_so_hs.get().strip()
            if not so_hs_text:
                messagebox.showwarning("Cảnh báo", "Vui lòng nhập số thí sinh mỗi phòng!")
                return
                
            try:
                self.so_thi_sinh_moi_phong = int(so_hs_text)
                if self.so_thi_sinh_moi_phong <= 0:
                    messagebox.showwarning("Cảnh báo", "Số thí sinh mỗi phòng phải lớn hơn 0!")
                    return
            except ValueError:
                messagebox.showerror("Lỗi", "Số thí sinh mỗi phòng phải là số nguyên!")
                return
                
            ten_phong_goc = self.entry_ten_phong.get().strip()
            if not ten_phong_goc:
                ten_phong_goc = "Phòng"  # Giá trị mặc định
            
            # Tạo bản sao của dataframe gốc
            self.df_da_chia = self.df_goc.copy()
            
            # Chia phòng với số lượng khác nhau cho từng phòng
            phong_thi = self.chia_phong_voi_so_luong_khac_nhau(len(self.df_da_chia), ten_phong_goc)
            self.df_da_chia['Phòng thi'] = phong_thi
            
            # Hiển thị kết quả
            self.hien_thi_ket_qua()
            
        except ValueError:
            messagebox.showerror("Lỗi", "Số thí sinh mỗi phòng phải là số nguyên!")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra: {str(e)}")
            
    def hien_thi_ket_qua(self):
        self.text_ket_qua.delete(1.0, tk.END)
        
        # Thống kê tổng quan
        so_hoc_sinh = len(self.df_da_chia)
        so_phong = len(self.df_da_chia['Phòng thi'].unique())
        
        self.text_ket_qua.insert(tk.END, "=== KẾT QUẢ CHIA PHÒNG THI ===\n\n")
        self.text_ket_qua.insert(tk.END, f"Tổng số thí sinh: {so_hoc_sinh}\n")
        self.text_ket_qua.insert(tk.END, f"Số phòng thi: {so_phong}\n")
        
        # Hiển thị thông tin về cấu hình phòng
        if self.cau_hinh_phong:
            self.text_ket_qua.insert(tk.END, f"📋 Cấu hình phòng tùy chỉnh: {len(self.cau_hinh_phong)} phòng\n")
            self.text_ket_qua.insert(tk.END, f"📋 Số thí sinh mặc định: {self.so_thi_sinh_moi_phong}\n\n")
        else:
            self.text_ket_qua.insert(tk.END, f"Số thí sinh mỗi phòng: {self.so_thi_sinh_moi_phong}\n\n")
        
        # Thống kê từng phòng
        self.text_ket_qua.insert(tk.END, "=== THỐNG KÊ TỪNG PHÒNG ===\n")
        thong_ke_phong = self.df_da_chia['Phòng thi'].value_counts().sort_index()
        
        for phong, so_hs in thong_ke_phong.items():
            # Kiểm tra xem phòng này có cấu hình tùy chỉnh không
            if phong in self.cau_hinh_phong:
                so_hs_cau_hinh = self.cau_hinh_phong[phong]
                if so_hs == so_hs_cau_hinh:
                    self.text_ket_qua.insert(tk.END, f"{phong}: {so_hs} thí sinh ✅ (đúng cấu hình)\n")
                else:
                    self.text_ket_qua.insert(tk.END, f"{phong}: {so_hs} thí sinh ⚠️ (cấu hình: {so_hs_cau_hinh})\n")
            else:
                self.text_ket_qua.insert(tk.END, f"{phong}: {so_hs} thí sinh (mặc định)\n")
        
        # Hiển thị thông tin cấu hình nếu có
        if self.cau_hinh_phong:
            self.text_ket_qua.insert(tk.END, "\n=== CẤU HÌNH PHÒNG TÙY CHỈNH ===\n")
            for phong, so_hs in self.cau_hinh_phong.items():
                self.text_ket_qua.insert(tk.END, f"{phong}: {so_hs} thí sinh\n")
        
        # Hiển thị thống kê chế độ thi cho hybrid
        if self.che_do_thi == "hybrid" and self.cot_che_do_thi:
            self.hien_thi_thong_ke_che_do_thi()
            
        self.text_ket_qua.insert(tk.END, "\n=== DANH SÁCH CHI TIẾT ===\n")
        
        # Hiển thị danh sách chi tiết (chỉ các cột được chọn + phòng thi)
        # Sử dụng thứ tự click chọn thay vì thứ tự trong file Excel
        cot_hien_thi = self.cac_cot_duoc_chon + ['Phòng thi']
        df_hien_thi = self.df_da_chia[cot_hien_thi]
        
        # Hiển thị 20 dòng đầu
        self.text_ket_qua.insert(tk.END, df_hien_thi.head(20).to_string(index=False))
        
        if len(df_hien_thi) > 20:
            self.text_ket_qua.insert(tk.END, f"\n... và {len(df_hien_thi) - 20} dòng khác")
    
    def hien_thi_thong_ke_che_do_thi(self):
        """Hiển thị thống kê chế độ thi cho chế độ hybrid"""
        try:
            if not self.cot_che_do_thi or self.cot_che_do_thi not in self.df_da_chia.columns:
                return
            
            self.text_ket_qua.insert(tk.END, "\n=== THỐNG KÊ CHẾ ĐỘ THI (HYBRID) ===\n")
            
            # Đếm số lượng theo chế độ thi
            thong_ke_che_do = {}
            for _, hs in self.df_da_chia.iterrows():
                che_do = str(hs[self.cot_che_do_thi]).strip().lower()
                
                # Chuẩn hóa giống như trong hàm dien_du_lieu_cac_cot_bo_sung
                if che_do in ['online', 'on', 'trực tuyến', 'online thi', 'thi online']:
                    che_do_chuan = "Online"
                elif che_do in ['offline', 'off', 'truyền thống', 'offline thi', 'thi offline']:
                    che_do_chuan = "Offline"
                else:
                    che_do_chuan = "Offline"  # Mặc định
                
                thong_ke_che_do[che_do_chuan] = thong_ke_che_do.get(che_do_chuan, 0) + 1
            
            # Hiển thị thống kê
            for che_do, so_luong in thong_ke_che_do.items():
                self.text_ket_qua.insert(tk.END, f"📊 {che_do}: {so_luong} thí sinh\n")
            
            # Hiển thị cảnh báo nếu có vấn đề
            if "Online" not in thong_ke_che_do:
                self.text_ket_qua.insert(tk.END, "⚠️ CẢNH BÁO: Không có thí sinh Online nào được nhận diện!\n")
                self.text_ket_qua.insert(tk.END, f"💡 Kiểm tra cột '{self.cot_che_do_thi}' có chứa giá trị 'Online' không?\n")
            elif "Offline" not in thong_ke_che_do:
                self.text_ket_qua.insert(tk.END, "⚠️ CẢNH BÁO: Không có thí sinh Offline nào được nhận diện!\n")
                self.text_ket_qua.insert(tk.END, f"💡 Kiểm tra cột '{self.cot_che_do_thi}' có chứa giá trị 'Offline' không?\n")
            
        except Exception as e:
            self.text_ket_qua.insert(tk.END, f"❌ Lỗi khi thống kê chế độ thi: {str(e)}\n")
            
    def xem_truoc_du_lieu(self):
        """Hiển thị xem trước thứ tự cột và dữ liệu mẫu"""
        if self.df_goc is None:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn file Excel trước!")
            return
            
        if not self.cac_cot_duoc_chon:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn ít nhất một cột dữ liệu!")
            return
        
        # Tạo cửa sổ xem trước
        window_truoc = tk.Toplevel(self.root)
        window_truoc.title("Xem trước dữ liệu")
        window_truoc.geometry("800x600")
        window_truoc.grab_set()  # Modal window
        
        # Frame chính
        main_frame_truoc = ttk.Frame(window_truoc, padding="10")
        main_frame_truoc.pack(fill=tk.BOTH, expand=True)
        
        # Tiêu đề
        ttk.Label(main_frame_truoc, text="XEM TRƯỚC DỮ LIỆU", font=("Arial", 14, "bold")).pack(pady=(0, 10))
        
        # Hiển thị thứ tự cột được chọn
        ttk.Label(main_frame_truoc, text="Thứ tự cột theo click chọn:", font=("Arial", 12, "bold")).pack(anchor=tk.W, pady=(0, 5))
        
        frame_thu_tu = ttk.Frame(main_frame_truoc)
        frame_thu_tu.pack(fill=tk.X, pady=(0, 10))
        
        for i, cot in enumerate(self.cac_cot_duoc_chon, 1):
            ttk.Label(frame_thu_tu, text=f"({i}) {cot}", font=("Arial", 10)).pack(anchor=tk.W)
        
        # Hiển thị dữ liệu mẫu
        ttk.Label(main_frame_truoc, text="Dữ liệu mẫu (10 dòng đầu):", font=("Arial", 12, "bold")).pack(anchor=tk.W, pady=(10, 5))
        
        # Tạo dataframe với thứ tự cột đã chọn
        df_truoc = self.df_goc[self.cac_cot_duoc_chon]
        
        # Text widget để hiển thị dữ liệu
        text_truoc = tk.Text(main_frame_truoc, height=15, width=80)
        scrollbar_truoc = ttk.Scrollbar(main_frame_truoc, orient="vertical", command=text_truoc.yview)
        text_truoc.configure(yscrollcommand=scrollbar_truoc.set)
        
        # Hiển thị dữ liệu
        text_truoc.insert(tk.END, df_truoc.head(10).to_string(index=False))
        
        # Pack widgets
        text_truoc.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_truoc.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Nút đóng
        ttk.Button(main_frame_truoc, text="Đóng", command=window_truoc.destroy).pack(pady=(10, 0))
    
    def preview_form_phong_thi(self):
        """Preview danh sách để in"""
        if self.df_goc is None:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn file Excel trước!")
            return
            
        if not self.cac_cot_duoc_chon:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn ít nhất một cột dữ liệu!")
            return
        
        # Cập nhật thông tin từ entry
        self.ban_to_chuc = self.combo_ban_to_chuc.get() or ""
        self.diem_thi = self.combo_diem_thi.get() or ""
        self.mon_thi = self.combo_mon_thi.get() or "Toán"
        self.thoi_gian_co_mat = self.combo_thoi_gian.get() or ""
        
        try:
            # Tạo file tạm để preview
            import tempfile
            temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            temp_file.close()
            
            # Tạo dữ liệu mẫu cho preview (chỉ 5 thí sinh)
            df_preview = self.df_goc.head(5).copy()
            df_preview['Phòng thi'] = 'Phòng 1'
            
            # Lưu dữ liệu gốc tạm thời
            df_goc_backup = self.df_goc
            df_da_chia_backup = self.df_da_chia
            cac_cot_backup = self.cac_cot_duoc_chon
            
            # Thiết lập dữ liệu preview
            self.df_goc = df_preview
            self.df_da_chia = df_preview
            self.cac_cot_duoc_chon = cac_cot_backup
            
            # Tạo form preview
            self.tao_form_phong_thi(temp_file.name)
            
            # Khôi phục dữ liệu gốc
            self.df_goc = df_goc_backup
            self.df_da_chia = df_da_chia_backup
            self.cac_cot_duoc_chon = cac_cot_backup
            
            # Mở file preview
            import subprocess
            import platform
            
            if platform.system() == 'Windows':
                os.startfile(temp_file.name)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.call(['open', temp_file.name])
            else:  # Linux
                subprocess.call(['xdg-open', temp_file.name])
                
            messagebox.showinfo("Thành công", f"Đã tạo danh sách để in preview: {os.path.basename(temp_file.name)}\nFile sẽ được mở tự động!")
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể tạo preview: {str(e)}")
            import traceback
            traceback.print_exc()
            
    def xuat_file_ket_qua(self):
        if self.df_da_chia is None:
            messagebox.showwarning("Cảnh báo", "Chưa có dữ liệu để xuất!")
            return
            
        # Tạo tên file mặc định theo thứ tự: Hình thức thi - Loại file - Điểm thi - Môn thi
        diem_thi = self.combo_diem_thi.get().strip()
        hinh_thuc_thi = self.var_che_do.get().upper()
        mon_thi = self.combo_mon_thi.get().strip()
        
        if diem_thi and mon_thi:
            ten_file_mac_dinh = f"{hinh_thuc_thi} - Kết quả chia phòng thi - {diem_thi} - {mon_thi}.xlsx"
        elif diem_thi:
            ten_file_mac_dinh = f"{hinh_thuc_thi} - Kết quả chia phòng thi - {diem_thi}.xlsx"
        else:
            ten_file_mac_dinh = f"{hinh_thuc_thi} - Kết quả chia phòng thi.xlsx"
            
        file_path = filedialog.asksaveasfilename(
            title="Lưu file kết quả",
            defaultextension=".xlsx",
            initialfile=ten_file_mac_dinh,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                # Xuất file với định dạng đẹp
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    # Sheet 1: Danh sách đầy đủ
                    self.df_da_chia.to_excel(writer, sheet_name='Danh sách đầy đủ', index=False)
                    
                    # Format tất cả các cột dạng text (để giữ số 0 đầu, số điện thoại...)
                    ws_full = writer.sheets['Danh sách đầy đủ']
                    for col_idx, col in enumerate(self.df_da_chia.columns, 1):
                        for row_idx in range(2, len(self.df_da_chia) + 2):  # Bỏ qua header
                            cell = ws_full.cell(row=row_idx, column=col_idx)
                            cell.number_format = '@'  # Format dạng text
                    
                    # Sheet 2: Chỉ các cột được chọn (theo thứ tự click)
                    cot_hien_thi = self.cac_cot_duoc_chon + ['Phòng thi']
                    df_hien_thi = self.df_da_chia[cot_hien_thi]
                    df_hien_thi.to_excel(writer, sheet_name='Danh sách rút gọn', index=False)
                    
                    # Format cột text cho sheet rút gọn
                    ws_rutgon = writer.sheets['Danh sách rút gọn']
                    for col_idx in range(1, len(cot_hien_thi) + 1):
                        for row_idx in range(2, len(df_hien_thi) + 2):
                            cell = ws_rutgon.cell(row=row_idx, column=col_idx)
                            cell.number_format = '@'  # Format dạng text
                    
                    # Sheet 3: Thống kê theo phòng với ghi chú chi tiết (sắp xếp theo số phòng)
                    thong_ke = self.tao_thong_ke_chi_tiet_theo_khoi()
                    if thong_ke is not None:
                        thong_ke.to_excel(writer, sheet_name='Thống kê phòng', index=False)
                        
                        # Format sheet Thống kê phòng
                        ws_thong_ke = writer.sheets['Thống kê phòng']
                        
                        # Thiết lập độ rộng cột
                        ws_thong_ke.column_dimensions['A'].width = 15  # Phòng thi
                        ws_thong_ke.column_dimensions['B'].width = 15  # Số thí sinh
                        ws_thong_ke.column_dimensions['C'].width = 80  # GHI CHÚ
                        
                        # Format header
                        for cell in ws_thong_ke[1]:
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Format dữ liệu
                        for row in ws_thong_ke.iter_rows(min_row=2, max_row=ws_thong_ke.max_row):
                            row[0].alignment = Alignment(horizontal='center', vertical='center')  # Phòng thi
                            row[1].alignment = Alignment(horizontal='center', vertical='center')  # Số thí sinh
                            row[2].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)  # GHI CHÚ
                    else:
                        # Fallback nếu không tạo được thống kê chi tiết
                        thong_ke = self.df_da_chia.groupby('Phòng thi').size().reset_index(name='Số thí sinh')
                        thong_ke = self.sap_xep_thong_ke_phong(thong_ke)
                        thong_ke.to_excel(writer, sheet_name='Thống kê phòng', index=False)
                
                messagebox.showinfo("Thành công", f"Đã xuất file thành công: {os.path.basename(file_path)}")
                
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể xuất file: {str(e)}")
                import traceback
                traceback.print_exc()
                
    def xuat_danh_sach_chia_phong(self):
        if self.df_da_chia is None:
            messagebox.showwarning("Cảnh báo", "Chưa có dữ liệu để xuất!")
            return
            
        # Tạo tên file mặc định theo thứ tự: Hình thức thi - Loại file - Điểm thi - Môn thi
        diem_thi = self.combo_diem_thi.get().strip()
        hinh_thuc_thi = self.var_che_do.get().upper()
        mon_thi = self.combo_mon_thi.get().strip()
        
        if diem_thi and mon_thi:
            ten_file_mac_dinh = f"{hinh_thuc_thi} - Danh sách chia phòng - {diem_thi} - {mon_thi}.xlsx"
        elif diem_thi:
            ten_file_mac_dinh = f"{hinh_thuc_thi} - Danh sách chia phòng - {diem_thi}.xlsx"
        else:
            ten_file_mac_dinh = f"{hinh_thuc_thi} - Danh sách chia phòng.xlsx"
            
        file_path = filedialog.asksaveasfilename(
            title="Lưu danh sách chia phòng",
            defaultextension=".xlsx",
            initialfile=ten_file_mac_dinh,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                # Tạo workbook mới
                wb = openpyxl.Workbook()
                wb.remove(wb.active)  # Xóa sheet mặc định
                
                # Lấy danh sách các phòng và sắp xếp theo thứ tự số
                cac_phong = self.sap_xep_phong_theo_so()
                
                for phong in cac_phong:
                    # Tạo sheet cho mỗi phòng
                    ws = wb.create_sheet(title=phong)
                    
                    # Lấy dữ liệu của phòng này
                    df_phong = self.df_da_chia[self.df_da_chia['Phòng thi'] == phong]
                    
                    # Chỉ lấy các cột được chọn (theo thứ tự click)
                    cot_hien_thi = self.cac_cot_duoc_chon + ['Phòng thi']
                    df_hien_thi = df_phong[cot_hien_thi]
                    
                    # Ghi dữ liệu vào sheet
                    for r in dataframe_to_rows(df_hien_thi, index=False, header=True):
                        ws.append(r)
                    
                    # Định dạng header
                    font_header = Font(name='Times New Roman', bold=True, size=12)
                    for cell in ws[1]:
                        cell.font = font_header
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Tự động điều chỉnh độ rộng cột
                    for col_idx, column in enumerate(ws.columns, 1):
                        max_length = 0
                        column_letter = get_column_letter(col_idx)
                        
                        for cell in column:
                            try:
                                # Kiểm tra nếu cell không phải là MergedCell
                                if hasattr(cell, 'value') and cell.value is not None:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                
                # Lưu file
                wb.save(file_path)
                messagebox.showinfo("Thành công", f"Đã xuất danh sách chia phòng thành công: {os.path.basename(file_path)}")
                
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể xuất danh sách chia phòng: {str(e)}")
                
    def xuat_form_phong_thi(self):
        if self.df_da_chia is None:
            messagebox.showwarning("Cảnh báo", "Chưa có dữ liệu để xuất!")
            return
            
        # Tạo tên file mặc định theo thứ tự: Hình thức thi - Loại file - Điểm thi - Môn thi
        diem_thi = (self.combo_diem_thi.get() or "").strip()
        hinh_thuc_thi = self.var_che_do.get().upper()
        mon_thi = (self.combo_mon_thi.get() or "").strip()
        
        if diem_thi and mon_thi:
            ten_file_mac_dinh = f"{hinh_thuc_thi} - Danh sách để in - {diem_thi} - {mon_thi}.xlsx"
        elif diem_thi:
            ten_file_mac_dinh = f"{hinh_thuc_thi} - Danh sách để in - {diem_thi}.xlsx"
        else:
            ten_file_mac_dinh = f"{hinh_thuc_thi} - Danh sách để in.xlsx"
            
        file_path = filedialog.asksaveasfilename(
            title="Lưu danh sách để in",
            defaultextension=".xlsx",
            initialfile=ten_file_mac_dinh,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                self.tao_form_phong_thi(file_path)
                messagebox.showinfo("Thành công", f"Đã xuất danh sách để in thành công: {os.path.basename(file_path)}")
                
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể xuất danh sách để in: {str(e)}")
                
    def xuat_danh_sach_cho_truong(self):
        """Xuất danh sách đơn giản cho trường (chỉ tiêu đề và danh sách, không có chữ ký và phần dưới)"""
        if self.df_da_chia is None:
            messagebox.showwarning("Cảnh báo", "Chưa có dữ liệu để xuất!")
            return
            
        # Kiểm tra chế độ thi
        if self.che_do_thi != "offline":
            messagebox.showwarning("Cảnh báo", "Tính năng này chỉ dành cho chế độ thi Offline!")
            return
            
        # Tạo tên file mặc định theo thứ tự: Hình thức thi - Loại file - Điểm thi - Môn thi
        diem_thi = (self.combo_diem_thi.get() or "").strip()
        hinh_thuc_thi = self.var_che_do.get().upper()
        mon_thi = (self.combo_mon_thi.get() or "").strip()
        
        if diem_thi and mon_thi:
            ten_file_mac_dinh = f"{hinh_thuc_thi} - Danh sách cho trường - {diem_thi} - {mon_thi}.xlsx"
        elif diem_thi:
            ten_file_mac_dinh = f"{hinh_thuc_thi} - Danh sách cho trường - {diem_thi}.xlsx"
        else:
            ten_file_mac_dinh = f"{hinh_thuc_thi} - Danh sách cho trường.xlsx"
            
        file_path = filedialog.asksaveasfilename(
            title="Lưu danh sách cho trường",
            defaultextension=".xlsx",
            initialfile=ten_file_mac_dinh,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                self.tao_danh_sach_cho_truong(file_path)
                messagebox.showinfo("Thành công", f"Đã xuất danh sách cho trường thành công: {os.path.basename(file_path)}")
                
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể xuất danh sách cho trường: {str(e)}")
                
    def tao_danh_sach_cho_truong(self, file_path):
        """Tạo danh sách đơn giản cho trường (chỉ tiêu đề và danh sách, không có chữ ký và phần dưới)"""
        print("🔄 Bắt đầu tạo danh sách cho trường...")
        
        # Kiểm tra dữ liệu cần thiết
        if self.df_da_chia is None or self.df_da_chia.empty:
            raise Exception("Không có dữ liệu để tạo danh sách cho trường!")
        
        if not self.cac_cot_duoc_chon:
            raise Exception("Chưa chọn cột dữ liệu nào!")
        
        try:
            # Tạo workbook mới
            wb = openpyxl.Workbook()
            
            # Xóa sheet mặc định
            wb.remove(wb.active)
            
            # Lấy danh sách các phòng và sắp xếp theo thứ tự số
            cac_phong = self.sap_xep_phong_theo_so()
            
            for phong in cac_phong:
                # Tạo sheet cho mỗi phòng
                ws = wb.create_sheet(title=phong)
                
                # Thiết lập lề in cho mỗi sheet - tất cả lề = 0
                ws.page_margins = PageMargins(
                    left=0,        # Lề trái: 0 inch
                    right=0,       # Lề phải: 0 inch
                    top=0,          # Lề trên: 0 inch
                    bottom=0,       # Lề dưới: 0 inch
                    header=0,       # Lề header: 0 inch
                    footer=0        # Lề footer: 0 inch
                )
            
                # Thiết lập khổ giấy A4 nằm ngang
                ws.page_setup.orientation = 'landscape'  # Nằm ngang
                ws.page_setup.paperSize = 9  # A4 (9 = A4)
                ws.page_setup.fitToPage = True  # Fit to page
                ws.page_setup.fitToWidth = 1  # Fit to 1 page wide
                ws.page_setup.fitToHeight = 1  # Fit to 1 page tall
                
            
                # Lấy dữ liệu của phòng này
                df_phong = self.df_da_chia[self.df_da_chia['Phòng thi'] == phong]
                
                # Tính số thí sinh dự kiến cho phòng này
                so_thi_sinh_du_kien_goc = self.tinh_so_thi_sinh_du_kien_cho_phong(phong)
                so_thi_sinh_du_kien = self.so_thi_sinh_moi_phong  # Luôn tạo bảng theo số mặc định
                so_thi_sinh_thuc_te = len(df_phong)
                
                print(f"📊 Phòng {phong} (danh sách trường): Cấu hình {so_thi_sinh_du_kien_goc} thí sinh, thực tế {so_thi_sinh_thuc_te} thí sinh, tạo bảng {so_thi_sinh_du_kien} dòng")
                
                # Tạo tiêu đề đơn giản (5 dòng)
                self.tao_tieu_de_don_gian(ws, phong, so_thi_sinh_du_kien)
                
                # Chèn ảnh vào góc trái và phải (cho danh sách trường)
                self.chen_anh_vao_trang(ws, is_for_school=True)
                
                # Tạo header cho bảng dữ liệu
                hang_tieu_de = 7  # Dòng 7 vì có 5 dòng tiêu đề + 1 dòng trống
                cot_bat_dau = 1
                
                # Tạo cột STT (số thứ tự)
                font_header = Font(name='Times New Roman', bold=True, size=12)
                cell_stt = ws.cell(row=hang_tieu_de, column=cot_bat_dau, value="STT")
                cell_stt.font = font_header
                cell_stt.alignment = Alignment(horizontal='center', vertical='center')
                
                # Header cho các cột được chọn
                for i, cot in enumerate(self.cac_cot_duoc_chon):
                    cell = ws.cell(row=hang_tieu_de, column=cot_bat_dau + 1 + i, value=cot)
                    cell.font = font_header
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Thêm cột G với tiêu đề "Ghi chú" cho danh sách trường
                cot_ghi_chu = cot_bat_dau + 1 + len(self.cac_cot_duoc_chon)
                cell_ghi_chu = ws.cell(row=hang_tieu_de, column=cot_ghi_chu, value="Ghi chú")
                cell_ghi_chu.font = font_header
                cell_ghi_chu.alignment = Alignment(horizontal='center', vertical='center')
                
                # Điền dữ liệu thí sinh - tạo đủ số dòng theo dự kiến
                font_du_lieu = Font(name='Times New Roman', size=12)
                
                for i in range(so_thi_sinh_du_kien):
                    hang = hang_tieu_de + 1 + i
                    
                    # Cột STT
                    cell_stt = ws.cell(row=hang, column=cot_bat_dau, value=i + 1)
                    cell_stt.font = font_du_lieu
                    cell_stt.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Các cột dữ liệu - chỉ điền nếu có thí sinh thực tế
                    if i < so_thi_sinh_thuc_te:
                        # Có thí sinh thực tế - điền dữ liệu
                        hs = df_phong.iloc[i]
                        for j, cot in enumerate(self.cac_cot_duoc_chon):
                            cell = ws.cell(row=hang, column=cot_bat_dau + 1 + j, value=hs[cot])
                            cell.font = font_du_lieu
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                    else:
                        # Không có thí sinh thực tế - để trống các ô
                        for j, cot in enumerate(self.cac_cot_duoc_chon):
                            cell = ws.cell(row=hang, column=cot_bat_dau + 1 + j, value="")
                            cell.font = font_du_lieu
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Thêm cột "Ghi chú" (luôn để trống)
                    cell_ghi_chu = ws.cell(row=hang, column=cot_ghi_chu, value="")
                    cell_ghi_chu.font = font_du_lieu
                    cell_ghi_chu.alignment = Alignment(horizontal='left', vertical='center')
            
                # Tạo border cho bảng - sử dụng số thí sinh dự kiến
                so_cot_tong = len(self.cac_cot_duoc_chon) + 2  # +1 cho STT, +1 cho cột Ghi chú
                self.tao_border_cho_bang(ws, hang_tieu_de, so_thi_sinh_du_kien + 1, so_cot_tong)  # +1 để bao gồm header
                
                # Tính toán độ rộng cột từ tất cả dữ liệu (cho danh sách trường)
                do_rong_cot = self.tinh_do_rong_cot_cho_truong()
                
                # Áp dụng độ rộng cột đã tính toán
                self.ap_dung_do_rong_cot(ws, do_rong_cot)
                
                # Thiết lập print area cho danh sách trường (chỉ đến cột G) - sau khi có dữ liệu
                so_cot_tong = len(self.cac_cot_duoc_chon) + 2  # +1 cho STT, +1 cho cột Ghi chú
                cot_cuoi_print = get_column_letter(so_cot_tong)  # Cột cuối là G
                ws.print_area = f'A1:{cot_cuoi_print}{ws.max_row}'  # Print area từ A1 đến cột G
                print(f"📄 Print area cho {phong}: A1:{cot_cuoi_print}{ws.max_row}")
            
            # Lưu file
            print(f"💾 Đang lưu file: {file_path}")
            wb.save(file_path)
            print(f"✅ Đã lưu file thành công!")
            
        except Exception as e:
            print(f"❌ Lỗi trong tao_danh_sach_cho_truong: {str(e)}")
            import traceback
            traceback.print_exc()
            raise e
    
    def tao_tieu_de_don_gian(self, ws, ten_phong, so_hoc_sinh):
        """Tạo tiêu đề đơn giản cho danh sách trường (5 dòng giống như danh sách để in)"""
        # Định nghĩa font Times New Roman - đồng nhất kích thước cho tất cả dòng
        font_tieu_de = Font(name='Times New Roman', bold=True, size=12)
        
        # Tính toán cột cuối cùng dựa trên số cột thực tế
        # Cột A: STT, Cột B-F: các cột được chọn, Cột G: Ghi chú
        so_cot_thuc_te = len(self.cac_cot_duoc_chon) + 2  # +1 cho STT, +1 cho cột Ghi chú
        cot_cuoi = get_column_letter(so_cot_thuc_te)  # Cột cuối là G (cột thứ 7)
        
        # Tiêu đề luôn merge từ B đến F cho tất cả chế độ thi
        merge_range = 'B1:F1'
        
        # Dòng 1: Ban tổ chức kỳ thi
        ws.merge_cells(merge_range)
        cell = ws.cell(row=1, column=2, value=f"BAN TỔ CHỨC KỲ THI {(self.ban_to_chuc or '').upper()}")
        cell.font = font_tieu_de
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Đảm bảo dòng 1 có chiều cao đồng nhất với các dòng khác
        ws.row_dimensions[1].height = 15.6  # Đồng nhất với chiều cao mặc định
        
        # Dòng 2: Điểm thi
        ws.merge_cells('B2:F2')
        cell = ws.cell(row=2, column=2, value=f"ĐIỂM THI: {(self.diem_thi or '').upper()}")
        cell.font = font_tieu_de
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Dòng 3: Môn thi
        ws.merge_cells('B3:F3')
        cell = ws.cell(row=3, column=2, value=f"MÔN THI: {(self.mon_thi or '').upper()}")
        cell.font = font_tieu_de
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Dòng 4: Phòng thi - chỉ hiển thị số phòng
        ws.merge_cells('B4:F4')
        # Tách số phòng từ tên phòng (ví dụ: "PT-LC 1" -> "1")
        so_phong = ten_phong.split()[-1] if ' ' in ten_phong else ten_phong
        cell = ws.cell(row=4, column=2, value=f"PHÒNG THI: {so_phong}")
        cell.font = font_tieu_de
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Dòng 5: Thời gian có mặt
        ws.merge_cells('B5:F5')
        cell = ws.cell(row=5, column=2, value=f"THÍ SINH CÓ MẶT TẠI PHÒNG THI LÚC {(self.thoi_gian_co_mat or '').upper()}")
        cell.font = font_tieu_de
        cell.alignment = Alignment(horizontal='center', vertical='center')
                
    def tao_form_phong_thi(self, file_path):
        print("🔄 Bắt đầu tạo danh sách để in...")
        
        # Kiểm tra dữ liệu cần thiết
        if self.df_da_chia is None or self.df_da_chia.empty:
            raise Exception("Không có dữ liệu để tạo form phòng thi!")
        
        if not self.cac_cot_duoc_chon:
            raise Exception("Chưa chọn cột dữ liệu nào!")
        
        try:
            # Tạo workbook mới
            wb = openpyxl.Workbook()
            
            # Xóa sheet mặc định
            wb.remove(wb.active)
            
            # Lấy danh sách các phòng và sắp xếp theo thứ tự số
            cac_phong = self.sap_xep_phong_theo_so()
            
            # Tính toán độ rộng cột từ tất cả dữ liệu
            do_rong_cot = self.tinh_do_rong_cot_toi_uu()
            
            for phong in cac_phong:
                # Tạo sheet cho mỗi phòng
                ws = wb.create_sheet(title=phong)
                
                
                # Thiết lập lề in cho mỗi sheet - tất cả lề = 0
                ws.page_margins = PageMargins(
                    left=0,        # Lề trái: 0 inch
                    right=0,       # Lề phải: 0 inch
                    top=0,          # Lề trên: 0 inch
                    bottom=0,       # Lề dưới: 0 inch
                    header=0,       # Lề header: 0 inch
                    footer=0        # Lề footer: 0 inch
                )
            
                # Thiết lập khổ giấy A4 nằm ngang
                ws.page_setup.orientation = 'landscape'  # Nằm ngang
                ws.page_setup.paperSize = 9  # A4 (9 = A4)
                ws.page_setup.fitToPage = True  # Fit to page
                ws.page_setup.fitToWidth = 1  # Fit to 1 page wide
                ws.page_setup.fitToHeight = 1  # Fit to 1 page tall
            
                # Thiết lập print area để đảm bảo in đúng - căn giữa nội dung
                # Tính toán số cột thực tế dựa trên dữ liệu
                so_cot_thuc_te = len(self.cac_cot_duoc_chon) + 2  # +2 cho STT và Chữ ký
                cot_cuoi = get_column_letter(so_cot_thuc_te)
                ws.print_area = f'A1:{cot_cuoi}{ws.max_row}'  # In từ A1 đến cột cuối cùng
                
                # Thiết lập print titles (lặp lại header)
                ws.print_title_rows = '7:7'  # Lặp lại dòng 7 (header) trên mỗi trang
            
                # Debug: In thông tin page setup
                print(f"📄 Page setup cho {phong}:")
                print(f"   - Orientation: {ws.page_setup.orientation}")
                print(f"   - Paper size: {ws.page_setup.paperSize}")
                print(f"   - Fit to page: {ws.page_setup.fitToPage}")
                print(f"   - Print area: {ws.print_area}")
            
                # Lấy dữ liệu của phòng này
                df_phong = self.df_da_chia[self.df_da_chia['Phòng thi'] == phong]
                
                # Tính số thí sinh dự kiến cho phòng này
                so_thi_sinh_du_kien_goc = self.tinh_so_thi_sinh_du_kien_cho_phong(phong)
                so_thi_sinh_du_kien = self.so_thi_sinh_moi_phong  # Luôn tạo bảng theo số mặc định
                so_thi_sinh_thuc_te = len(df_phong)
                
                print(f"📊 Phòng {phong}: Cấu hình {so_thi_sinh_du_kien_goc} thí sinh, thực tế {so_thi_sinh_thuc_te} thí sinh, tạo bảng {so_thi_sinh_du_kien} dòng")
                
                # Tạo tiêu đề (5 dòng)
                self.tao_tieu_de_phong_moi(ws, phong, so_thi_sinh_du_kien)
                
                # Chèn ảnh vào góc trái và phải (cho danh sách để in)
                self.chen_anh_vao_trang(ws, is_for_school=False)
            
                # Tạo header cho bảng dữ liệu
                hang_tieu_de = 7  # Dòng 7 vì có 5 dòng tiêu đề + 1 dòng trống
                cot_bat_dau = 1
                
                # Tạo cột STT (số thứ tự)
                font_header = Font(name='Times New Roman', bold=True, size=12)
                cell_stt = ws.cell(row=hang_tieu_de, column=cot_bat_dau, value="STT")
                cell_stt.font = font_header
                cell_stt.alignment = Alignment(horizontal='center', vertical='center')
                
                # Header cho các cột được chọn
                for i, cot in enumerate(self.cac_cot_duoc_chon):
                    cell = ws.cell(row=hang_tieu_de, column=cot_bat_dau + 1 + i, value=cot)
                    cell.font = font_header
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Thêm các cột bổ sung theo chế độ thi
                self.tao_cac_cot_bo_sung_theo_che_do(ws, hang_tieu_de, cot_bat_dau + 1 + len(self.cac_cot_duoc_chon), font_header)
                
                # Điền dữ liệu thí sinh - tạo đủ số dòng theo dự kiến
                font_du_lieu = Font(name='Times New Roman', size=12)
                
                for i in range(so_thi_sinh_du_kien):
                    hang = hang_tieu_de + 1 + i
                    
                    # Cột STT
                    cell_stt = ws.cell(row=hang, column=cot_bat_dau, value=i + 1)
                    cell_stt.font = font_du_lieu
                    cell_stt.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Các cột dữ liệu - chỉ điền nếu có thí sinh thực tế
                    if i < so_thi_sinh_thuc_te:
                        # Có thí sinh thực tế - điền dữ liệu
                        hs = df_phong.iloc[i]
                        for j, cot in enumerate(self.cac_cot_duoc_chon):
                            cell = ws.cell(row=hang, column=cot_bat_dau + 1 + j, value=hs[cot])
                            cell.font = font_du_lieu
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        
                        # Điền dữ liệu cho các cột bổ sung theo chế độ thi
                        self.dien_du_lieu_cac_cot_bo_sung(ws, hang, cot_bat_dau + 1 + len(self.cac_cot_duoc_chon), font_du_lieu, hs)
                    else:
                        # Không có thí sinh thực tế - để trống các ô
                        for j, cot in enumerate(self.cac_cot_duoc_chon):
                            cell = ws.cell(row=hang, column=cot_bat_dau + 1 + j, value="")
                            cell.font = font_du_lieu
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        
                        # Để trống các cột bổ sung
                        self.dien_du_lieu_cac_cot_bo_sung_trong(ws, hang, cot_bat_dau + 1 + len(self.cac_cot_duoc_chon), font_du_lieu)
            
                # Tạo border cho bảng - sử dụng số thí sinh dự kiến
                so_cot_bo_sung = self.tinh_so_cot_bo_sung()
                so_cot_tong = len(self.cac_cot_duoc_chon) + 1 + so_cot_bo_sung  # +1 cho STT
                self.tao_border_cho_bang(ws, hang_tieu_de, so_thi_sinh_du_kien + 1, so_cot_tong)  # +1 để bao gồm header
                
                # Thêm phần cuối danh sách
                hang_cuoi = hang_tieu_de + so_thi_sinh_du_kien + 2  # 2 dòng cách
                self.tao_phan_cuoi_danh_sach_moi(ws, hang_cuoi)
                        
                # Áp dụng độ rộng cột đã tính toán
                self.ap_dung_do_rong_cot(ws, do_rong_cot)
                
                # Thiết lập page setup cuối cùng sau khi có dữ liệu
                self.thiet_lap_page_setup_cuoi_cung(ws)
            
            # Lưu file
            print(f"💾 Đang lưu file: {file_path}")
            wb.save(file_path)
            print(f"✅ Đã lưu file thành công!")
            
        except Exception as e:
            print(f"❌ Lỗi trong tao_form_phong_thi: {str(e)}")
            import traceback
            traceback.print_exc()
            raise e
        
    def chen_anh_vao_trang(self, ws, is_for_school=False):
        """Chèn ảnh vào góc trái và phải của trang"""
        try:
            # Chọn ảnh phù hợp dựa trên loại danh sách
            if is_for_school:
                # Sử dụng ảnh cho danh sách trường
                anh_trai = self.duong_dan_anh_trai_truong
                anh_phai = self.duong_dan_anh_phai_truong
                loai_ds = "danh sách trường"
            else:
                # Sử dụng ảnh cho danh sách để in
                anh_trai = self.duong_dan_anh_trai
                anh_phai = self.duong_dan_anh_phai
                loai_ds = "danh sách để in"
            
            # Chèn ảnh góc trái (cột A, dòng 1-5)
            if anh_trai and os.path.exists(anh_trai):
                img_trai = Image(anh_trai)
                # Điều chỉnh kích thước ảnh: chiều cao 2cm, chiều rộng theo tỷ lệ
                self.dieu_chinh_kich_thuoc_anh(img_trai)
                # Chèn vào góc trái (A1)
                ws.add_image(img_trai, 'A1')
                print(f"✅ Đã chèn ảnh trái cho {loai_ds}: {os.path.basename(anh_trai)}")
            
            # Chèn ảnh góc phải (cột cuối, dòng 1-5)
            if anh_phai and os.path.exists(anh_phai):
                img_phai = Image(anh_phai)
                # Điều chỉnh kích thước ảnh: chiều cao 2cm, chiều rộng theo tỷ lệ
                self.dieu_chinh_kich_thuoc_anh(img_phai)
                # Ảnh phải luôn ở cột G cho đồng bộ
                ws.add_image(img_phai, 'G1')
                print(f"✅ Đã chèn ảnh phải cho {loai_ds} vào cột G: {os.path.basename(anh_phai)}")
                
        except Exception as e:
            print(f"❌ Lỗi khi chèn ảnh: {str(e)}")
            # Không dừng chương trình nếu lỗi chèn ảnh
            pass
    
    def tao_cac_cot_bo_sung_theo_che_do(self, ws, hang, cot_bat_dau, font_header):
        """Tạo các cột bổ sung theo chế độ thi"""
        if self.che_do_thi == "offline":
            # Chỉ có cột Chữ ký
            cell_chu_ky = ws.cell(row=hang, column=cot_bat_dau, value="Chữ ký")
            cell_chu_ky.font = font_header
            cell_chu_ky.alignment = Alignment(horizontal='center', vertical='center')
            
        elif self.che_do_thi == "online":
            # Các cột cho thi online
            cot_hien_tai = cot_bat_dau
            cot_ten = ["Thời gian gọi trước", "Thời gian gọi trong ngày", "Điểm danh", "Thời gian nộp bài", "Ghi chú"]
            for ten_cot in cot_ten:
                cell = ws.cell(row=hang, column=cot_hien_tai, value=ten_cot)
                cell.font = font_header
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cot_hien_tai += 1
                
        elif self.che_do_thi == "hybrid":
            # Các cột cho thi hỗn hợp - bố trí thông minh với ghi chú
            cot_hien_tai = cot_bat_dau
            # Cột chung cho cả offline và online
            cot_ten = ["Chế độ thi", "Ghi chú (chung)"]
            for ten_cot in cot_ten:
                cell = ws.cell(row=hang, column=cot_hien_tai, value=ten_cot)
                cell.font = font_header
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cot_hien_tai += 1
            
            # Cột riêng cho Online - thêm ghi chú
            cell_online = ws.cell(row=hang, column=cot_hien_tai, value="Thời gian gọi trước (Online)")
            cell_online.font = font_header
            cell_online.alignment = Alignment(horizontal='center', vertical='center')
            cot_hien_tai += 1
            
            cell_online = ws.cell(row=hang, column=cot_hien_tai, value="Thời gian gọi trong ngày (Online)")
            cell_online.font = font_header
            cell_online.alignment = Alignment(horizontal='center', vertical='center')
            cot_hien_tai += 1
            
            cell_online = ws.cell(row=hang, column=cot_hien_tai, value="Điểm danh (Online)")
            cell_online.font = font_header
            cell_online.alignment = Alignment(horizontal='center', vertical='center')
            cot_hien_tai += 1
            
            cell_online = ws.cell(row=hang, column=cot_hien_tai, value="Thời gian nộp bài (Online)")
            cell_online.font = font_header
            cell_online.alignment = Alignment(horizontal='center', vertical='center')
            cot_hien_tai += 1
            
            # Cột riêng cho Offline - thêm ghi chú
            cell_offline = ws.cell(row=hang, column=cot_hien_tai, value="Chữ ký (Offline)")
            cell_offline.font = font_header
            cell_offline.alignment = Alignment(horizontal='center', vertical='center')
    
    def dien_du_lieu_cac_cot_bo_sung(self, ws, hang, cot_bat_dau, font_du_lieu, hs):
        """Điền dữ liệu cho các cột bổ sung theo chế độ thi"""
        if self.che_do_thi == "offline":
            # Chỉ có cột Chữ ký (để trống)
            cell_chu_ky = ws.cell(row=hang, column=cot_bat_dau, value="")
            cell_chu_ky.font = font_du_lieu
            
        elif self.che_do_thi == "online":
            # Điền dữ liệu cho thi online
            cot_hien_tai = cot_bat_dau
            # Thời gian gọi trước (để trống)
            ws.cell(row=hang, column=cot_hien_tai, value="").font = font_du_lieu
            cot_hien_tai += 1
            # Thời gian gọi trong ngày (để trống)
            ws.cell(row=hang, column=cot_hien_tai, value="").font = font_du_lieu
            cot_hien_tai += 1
            # Điểm danh (để trống)
            ws.cell(row=hang, column=cot_hien_tai, value="").font = font_du_lieu
            cot_hien_tai += 1
            # Thời gian nộp bài (để trống)
            ws.cell(row=hang, column=cot_hien_tai, value="").font = font_du_lieu
            cot_hien_tai += 1
            # Ghi chú (để trống)
            ws.cell(row=hang, column=cot_hien_tai, value="").font = font_du_lieu
            
        elif self.che_do_thi == "hybrid":
            # Điền dữ liệu cho thi hỗn hợp - bố trí thông minh
            cot_hien_tai = cot_bat_dau
            
            # Xác định chế độ thi của thí sinh này
            if self.cot_che_do_thi and self.cot_che_do_thi in hs:
                che_do_thi_value = str(hs[self.cot_che_do_thi]).strip()
                print(f"🔍 Debug: Thí sinh {hs.get('Họ tên', 'N/A')} - Giá trị gốc: '{che_do_thi_value}'")
                
                # Chuẩn hóa giá trị - ưu tiên Online trước
                if che_do_thi_value.lower() in ['online', 'on', 'trực tuyến', 'online thi', 'thi online']:
                    che_do_thi_value = "Online"
                    print(f"✅ Nhận diện Online: {che_do_thi_value}")
                elif che_do_thi_value.lower() in ['offline', 'off', 'truyền thống', 'offline thi', 'thi offline']:
                    che_do_thi_value = "Offline"
                    print(f"✅ Nhận diện Offline: {che_do_thi_value}")
                else:
                    che_do_thi_value = "Offline"  # Mặc định
                    print(f"⚠️ Không nhận diện được, mặc định Offline: {che_do_thi_value}")
            else:
                che_do_thi_value = "Offline"  # Mặc định
                print(f"⚠️ Không có cột chế độ thi, mặc định Offline: {che_do_thi_value}")
            
            # Cột 1: Chế độ thi (chung cho cả offline và online)
            ws.cell(row=hang, column=cot_hien_tai, value=che_do_thi_value).font = font_du_lieu
            cot_hien_tai += 1
            
            # Cột 2: Ghi chú (chung cho cả offline và online)
            ws.cell(row=hang, column=cot_hien_tai, value="").font = font_du_lieu
            cot_hien_tai += 1
            
            # Điền dữ liệu theo chế độ thi thực tế
            if che_do_thi_value == "Online":
                # Thí sinh Online: điền các cột online, để trống cột offline
                # Thời gian gọi trước (để trống)
                ws.cell(row=hang, column=cot_hien_tai, value="").font = font_du_lieu
                cot_hien_tai += 1
                # Thời gian gọi trong ngày (để trống)
                ws.cell(row=hang, column=cot_hien_tai, value="").font = font_du_lieu
                cot_hien_tai += 1
                # Điểm danh (để trống)
                ws.cell(row=hang, column=cot_hien_tai, value="").font = font_du_lieu
                cot_hien_tai += 1
                # Thời gian nộp bài (để trống)
                ws.cell(row=hang, column=cot_hien_tai, value="").font = font_du_lieu
                cot_hien_tai += 1
                # Chữ ký (không cần cho online - để trống)
                ws.cell(row=hang, column=cot_hien_tai, value="").font = font_du_lieu
            else:
                # Thí sinh Offline: để trống các cột online, chỉ cần chữ ký
                # Thời gian gọi trước (không cần cho offline - để trống)
                ws.cell(row=hang, column=cot_hien_tai, value="").font = font_du_lieu
                cot_hien_tai += 1
                # Thời gian gọi trong ngày (không cần cho offline - để trống)
                ws.cell(row=hang, column=cot_hien_tai, value="").font = font_du_lieu
                cot_hien_tai += 1
                # Điểm danh (không cần cho offline - để trống)
                ws.cell(row=hang, column=cot_hien_tai, value="").font = font_du_lieu
                cot_hien_tai += 1
                # Thời gian nộp bài (không cần cho offline - để trống)
                ws.cell(row=hang, column=cot_hien_tai, value="").font = font_du_lieu
                cot_hien_tai += 1
                # Chữ ký (cần cho offline - để trống để ký)
                ws.cell(row=hang, column=cot_hien_tai, value="").font = font_du_lieu
    
    def dien_du_lieu_cac_cot_bo_sung_trong(self, ws, hang, cot_bat_dau, font_du_lieu):
        """Điền dữ liệu trống cho các cột bổ sung (dành cho các dòng trống)"""
        if self.che_do_thi == "offline":
            # Chỉ có cột Chữ ký (để trống)
            cell_chu_ky = ws.cell(row=hang, column=cot_bat_dau, value="")
            cell_chu_ky.font = font_du_lieu
            
        elif self.che_do_thi == "online":
            # Điền dữ liệu trống cho thi online
            cot_hien_tai = cot_bat_dau
            # Thời gian gọi trước (để trống)
            ws.cell(row=hang, column=cot_hien_tai, value="").font = font_du_lieu
            cot_hien_tai += 1
            # Thời gian gọi trong ngày (để trống)
            ws.cell(row=hang, column=cot_hien_tai, value="").font = font_du_lieu
            cot_hien_tai += 1
            # Điểm danh (để trống)
            ws.cell(row=hang, column=cot_hien_tai, value="").font = font_du_lieu
            cot_hien_tai += 1
            # Thời gian nộp bài (để trống)
            ws.cell(row=hang, column=cot_hien_tai, value="").font = font_du_lieu
            cot_hien_tai += 1
            # Ghi chú (để trống)
            ws.cell(row=hang, column=cot_hien_tai, value="").font = font_du_lieu
            
        elif self.che_do_thi == "hybrid":
            # Điền dữ liệu trống cho thi hỗn hợp
            cot_hien_tai = cot_bat_dau
            
            # Cột 1: Chế độ thi (để trống)
            ws.cell(row=hang, column=cot_hien_tai, value="").font = font_du_lieu
            cot_hien_tai += 1
            
            # Cột 2: Ghi chú (để trống)
            ws.cell(row=hang, column=cot_hien_tai, value="").font = font_du_lieu
            cot_hien_tai += 1
            
            # Các cột online (để trống)
            for _ in range(4):  # 4 cột online
                ws.cell(row=hang, column=cot_hien_tai, value="").font = font_du_lieu
                cot_hien_tai += 1
            
            # Cột chữ ký offline (để trống)
            ws.cell(row=hang, column=cot_hien_tai, value="").font = font_du_lieu
    
    def tinh_so_thi_sinh_du_kien_cho_phong(self, ten_phong):
        """Tính số thí sinh dự kiến cho một phòng cụ thể"""
        # Kiểm tra xem phòng này có cấu hình tùy chỉnh không
        if ten_phong in self.cau_hinh_phong:
            return self.cau_hinh_phong[ten_phong]
        else:
            # Sử dụng số thí sinh mặc định
            return self.so_thi_sinh_moi_phong
    
    def tinh_so_cot_bo_sung(self):
        if self.che_do_thi == "offline":
            return 1  # Chỉ có cột Chữ ký
        elif self.che_do_thi == "online":
            return 5  # Thời gian gọi trước, thời gian gọi trong ngày, điểm danh, thời gian nộp bài, ghi chú
        elif self.che_do_thi == "hybrid":
            return 7  # Chế độ thi, ghi chú, thời gian gọi trước, thời gian gọi trong ngày, điểm danh, thời gian nộp bài, chữ ký
        return 1  # Mặc định
    
    def dieu_chinh_kich_thuoc_anh(self, img):
        """Điều chỉnh kích thước ảnh: chiều cao tùy chỉnh, chiều rộng theo tỷ lệ"""
        try:
            # Sử dụng chiều cao tùy chỉnh từ biến self.chieu_cao_anh (đơn vị: cm)
            # Chuyển đổi từ cm sang points (1cm = 28.35 points) + thêm 20% để đảm bảo hiển thị đúng
            chieu_cao_muc_tieu = self.chieu_cao_anh * 28.35 * 1.2
            
            # Lấy kích thước gốc của ảnh
            chieu_rong_goc = img.width
            chieu_cao_goc = img.height
            
            # Tính tỷ lệ để đạt chiều cao mục tiêu
            ty_le = chieu_cao_muc_tieu / chieu_cao_goc
            
            # Tính chiều rộng mới theo tỷ lệ
            chieu_rong_moi = chieu_rong_goc * ty_le
            
            # Áp dụng kích thước mới
            img.height = chieu_cao_muc_tieu
            img.width = chieu_rong_moi
            
            print(f"📏 Kích thước ảnh: {chieu_rong_moi:.1f} x {chieu_cao_muc_tieu:.1f} points (chiều cao: {self.chieu_cao_anh}cm)")
            
        except Exception as e:
            print(f"❌ Lỗi khi điều chỉnh kích thước ảnh: {str(e)}")
            # Nếu lỗi, dùng kích thước mặc định
            chieu_cao_mac_dinh = self.chieu_cao_anh * 28.35 * 1.2
            img.height = chieu_cao_mac_dinh
            img.width = chieu_cao_mac_dinh

    def tao_tieu_de_phong_moi(self, ws, ten_phong, so_hoc_sinh):
        """Tạo tiêu đề danh sách để in theo mẫu mới"""
        # Định nghĩa font Times New Roman - đồng nhất kích thước cho tất cả dòng
        font_tieu_de = Font(name='Times New Roman', bold=True, size=12)
        
        # Tính toán cột cuối cùng dựa trên số cột thực tế
        so_cot_bo_sung = self.tinh_so_cot_bo_sung()
        so_cot_thuc_te = len(self.cac_cot_duoc_chon) + 1 + so_cot_bo_sung  # +1 cho STT + số cột bổ sung
        cot_cuoi = get_column_letter(so_cot_thuc_te)
        
        # Tiêu đề luôn merge từ B đến F cho tất cả chế độ thi
        merge_range = 'B1:F1'
        
        # Dòng 1: Ban tổ chức kỳ thi
        ws.merge_cells(merge_range)
        cell = ws.cell(row=1, column=2, value=f"BAN TỔ CHỨC KỲ THI {(self.ban_to_chuc or '').upper()}")
        cell.font = font_tieu_de
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Đảm bảo dòng 1 có chiều cao đồng nhất với các dòng khác
        ws.row_dimensions[1].height = 15.6  # Đồng nhất với chiều cao mặc định
        
        # Dòng 2: Điểm thi
        ws.merge_cells('B2:F2')
        cell = ws.cell(row=2, column=2, value=f"ĐIỂM THI: {(self.diem_thi or '').upper()}")
        cell.font = font_tieu_de
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Dòng 3: Môn thi
        ws.merge_cells('B3:F3')
        cell = ws.cell(row=3, column=2, value=f"MÔN THI: {(self.mon_thi or '').upper()}")
        cell.font = font_tieu_de
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Dòng 4: Phòng thi - chỉ hiển thị số phòng
        ws.merge_cells('B4:F4')
        # Tách số phòng từ tên phòng (ví dụ: "PT-LC 1" -> "1")
        so_phong = ten_phong.split()[-1] if ' ' in ten_phong else ten_phong
        cell = ws.cell(row=4, column=2, value=f"PHÒNG THI: {so_phong}")
        cell.font = font_tieu_de
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Dòng 5: Thời gian có mặt
        ws.merge_cells('B5:F5')
        cell = ws.cell(row=5, column=2, value=f"THÍ SINH CÓ MẶT TẠI PHÒNG THI LÚC {(self.thoi_gian_co_mat or '').upper()}")
        cell.font = font_tieu_de
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def sap_xep_phong_theo_so(self):
        """Sắp xếp danh sách phòng theo thứ tự số"""
        cac_phong = list(self.df_da_chia['Phòng thi'].unique())
        
        def extract_number(phong_name):
            """Trích xuất số từ tên phòng để sắp xếp"""
            # Tách phần cuối của tên phòng (số)
            parts = phong_name.split()
            if len(parts) > 0:
                try:
                    # Lấy phần cuối và chuyển thành số
                    return int(parts[-1])
                except ValueError:
                    # Nếu không phải số, trả về 0
                    return 0
            return 0
        
        # Sắp xếp theo số
        cac_phong.sort(key=extract_number)
        return cac_phong
    
    def sap_xep_thong_ke_phong(self, thong_ke_df):
        """Sắp xếp thống kê phòng theo thứ tự số"""
        def extract_number(phong_name):
            """Trích xuất số từ tên phòng để sắp xếp"""
            # Tách phần cuối của tên phòng (số)
            parts = phong_name.split()
            if len(parts) > 0:
                try:
                    # Lấy phần cuối và chuyển thành số
                    return int(parts[-1])
                except ValueError:
                    # Nếu không phải số, trả về 0
                    return 0
            return 0
        
        # Sắp xếp theo số phòng
        thong_ke_df = thong_ke_df.sort_values('Phòng thi', key=lambda x: x.apply(extract_number))
        return thong_ke_df.reset_index(drop=True)
    
    def tao_thong_ke_chi_tiet_theo_khoi(self):
        """Tạo thống kê chi tiết số lượng bài thi theo từng khối cho mỗi phòng"""
        if self.df_da_chia is None:
            return None
        
        thong_ke_list = []
        
        # Tìm cột chứa thông tin về môn học và khối
        # Thường các cột này có tên chứa từ "môn", "khối", "bài thi", etc.
        cot_mon = None
        cot_khoi = None
        
        # Duyệt qua các cột để tìm cột môn và khối
        for cot in self.df_da_chia.columns:
            cot_lower = cot.lower()
            if any(keyword in cot_lower for keyword in ['môn', 'mon', 'subject', 'bài thi', 'bai thi']):
                cot_mon = cot
            if any(keyword in cot_lower for keyword in ['khối', 'khoi', 'grade', 'lớp', 'lop']):
                cot_khoi = cot
        
        # Lấy danh sách phòng và sắp xếp
        cac_phong = self.df_da_chia['Phòng thi'].unique()
        cac_phong = sorted(cac_phong, key=lambda x: self.extract_number_from_room(x))
        
        for phong in cac_phong:
            df_phong = self.df_da_chia[self.df_da_chia['Phòng thi'] == phong]
            so_thi_sinh = len(df_phong)
            
            # Thống kê theo khối
            thong_ke_khoi = {}
            
            if cot_mon and cot_khoi:
                # Nếu có cả cột môn và khối
                for _, hs in df_phong.iterrows():
                    mon = str(hs[cot_mon]).strip() if pd.notna(hs[cot_mon]) else ""
                    khoi = str(hs[cot_khoi]).strip() if pd.notna(hs[cot_khoi]) else ""
                    
                    # Tạo key dạng "TOÁN KHỐI 1"
                    if mon and khoi:
                        key = f"{mon.upper()} KHỐI {khoi}"
                    elif mon:
                        key = mon.upper()
                    elif khoi:
                        key = f"KHỐI {khoi}"
                    else:
                        key = "KHÁC"
                    
                    thong_ke_khoi[key] = thong_ke_khoi.get(key, 0) + 1
                    
            elif cot_khoi:
                # Chỉ có cột khối
                mon_hien_tai = self.mon_thi or "TOÁN"
                for _, hs in df_phong.iterrows():
                    khoi = str(hs[cot_khoi]).strip() if pd.notna(hs[cot_khoi]) else ""
                    
                    if khoi:
                        key = f"{mon_hien_tai.upper()} KHỐI {khoi}"
                    else:
                        key = mon_hien_tai.upper()
                    
                    thong_ke_khoi[key] = thong_ke_khoi.get(key, 0) + 1
                    
            elif cot_mon:
                # Chỉ có cột môn
                for _, hs in df_phong.iterrows():
                    mon = str(hs[cot_mon]).strip() if pd.notna(hs[cot_mon]) else ""
                    
                    if mon:
                        key = mon.upper()
                    else:
                        key = "KHÁC"
                    
                    thong_ke_khoi[key] = thong_ke_khoi.get(key, 0) + 1
            else:
                # Không tìm thấy cột môn hoặc khối - chỉ đếm tổng số
                mon_hien_tai = self.mon_thi or "TOÁN"
                key = f"{so_thi_sinh} {mon_hien_tai.upper()}"
                thong_ke_khoi[key] = so_thi_sinh
            
            # Tạo chuỗi ghi chú
            ghi_chu_parts = []
            
            # Kiểm tra nếu tất cả thí sinh cùng loại
            if len(thong_ke_khoi) == 1:
                key = list(thong_ke_khoi.keys())[0]
                count = thong_ke_khoi[key]
                ghi_chu = f"{count} {key}"
            else:
                # Sắp xếp theo khối (nếu có số khối trong key)
                sorted_items = sorted(thong_ke_khoi.items(), key=lambda x: self.extract_grade_from_key(x[0]))
                
                for key, count in sorted_items:
                    ghi_chu_parts.append(f"{count} {key}")
                
                ghi_chu = ", ".join(ghi_chu_parts)
            
            thong_ke_list.append({
                'Phòng thi': phong,
                'Số thí sinh': so_thi_sinh,
                'GHI CHÚ': ghi_chu
            })
        
        return pd.DataFrame(thong_ke_list)
    
    def extract_number_from_room(self, phong_name):
        """Trích xuất số từ tên phòng để sắp xếp"""
        parts = phong_name.split()
        if len(parts) > 0:
            try:
                return int(parts[-1])
            except ValueError:
                return 0
        return 0
    
    def extract_grade_from_key(self, key):
        """Trích xuất số khối từ key để sắp xếp"""
        import re
        match = re.search(r'KHỐI\s+(\d+)', key)
        if match:
            return int(match.group(1))
        return 999  # Đặt ở cuối nếu không tìm thấy số khối
    
    def tinh_do_rong_cot_cho_truong(self):
        """Tính toán độ rộng cột tối ưu cho danh sách trường"""
        do_rong_cot = {}
        
        # Cột STT cố định
        do_rong_cot['A'] = 8  # STT
        
        # Tính toán độ rộng cho các cột dữ liệu
        for i, cot in enumerate(self.cac_cot_duoc_chon, 1):
            cot_letter = get_column_letter(i + 1)  # Bắt đầu từ cột B
            
            # Tìm độ dài tối đa trong cột này
            max_length = 0
            
            # Kiểm tra tên cột (header)
            if len(cot) > max_length:
                max_length = len(cot)
            
            # Kiểm tra tất cả dữ liệu trong cột
            if self.df_da_chia is not None:
                for _, row in self.df_da_chia.iterrows():
                    if cot in row and pd.notna(row[cot]):
                        cell_value = str(row[cot])
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
            
            # Điều chỉnh độ rộng - tăng padding để đảm bảo không bị mất chữ
            if 'tên' in cot.lower() or 'name' in cot.lower():
                # Cột tên cần rộng hơn
                adjusted_width = min(max_length + 5, 60)
            elif 'trường' in cot.lower() or 'school' in cot.lower():
                # Cột trường cần rộng hơn
                adjusted_width = min(max_length + 5, 70)
            else:
                # Các cột khác
                adjusted_width = min(max_length + 3, 50)
            
            do_rong_cot[cot_letter] = adjusted_width
        
        # Cột Ghi chú cho danh sách trường
        cot_ghi_chu = get_column_letter(len(self.cac_cot_duoc_chon) + 2)  # Cột cuối
        do_rong_cot[cot_ghi_chu] = 20  # Ghi chú
        
        return do_rong_cot
    
    def tinh_do_rong_cot_toi_uu(self):
        """Tính toán độ rộng cột tối ưu từ tất cả dữ liệu - tối ưu để không bị mất chữ"""
        do_rong_cot = {}
        
        # Cột STT cố định - nhỏ hơn
        do_rong_cot['A'] = 8  # STT
        
        # Tính toán độ rộng cho các cột dữ liệu
        for i, cot in enumerate(self.cac_cot_duoc_chon, 1):
            cot_letter = get_column_letter(i + 1)  # Bắt đầu từ cột B
            
            # Tìm độ dài tối đa trong cột này
            max_length = 0
            
            # Kiểm tra tên cột (header)
            if len(cot) > max_length:
                max_length = len(cot)
            
            # Kiểm tra tất cả dữ liệu trong cột
            if self.df_da_chia is not None:
                for _, row in self.df_da_chia.iterrows():
                    if cot in row and pd.notna(row[cot]):
                        cell_value = str(row[cot])
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
            
            # Điều chỉnh độ rộng - tăng padding để đảm bảo không bị mất chữ
            # Đặc biệt tối ưu cho cột tên và trường
            if 'tên' in cot.lower() or 'name' in cot.lower():
                # Cột tên cần rộng hơn
                adjusted_width = min(max_length + 5, 60)
            elif 'trường' in cot.lower() or 'school' in cot.lower():
                # Cột trường cần rộng hơn
                adjusted_width = min(max_length + 5, 70)
            else:
                # Các cột khác
                adjusted_width = min(max_length + 3, 50)
            
            do_rong_cot[cot_letter] = adjusted_width
        
        # Các cột bổ sung theo chế độ thi
        so_cot_bo_sung = self.tinh_so_cot_bo_sung()
        for i in range(so_cot_bo_sung):
            cot_letter = get_column_letter(len(self.cac_cot_duoc_chon) + 2 + i)  # +2 cho STT và bắt đầu từ cột tiếp theo
            if self.che_do_thi == "offline":
                do_rong_cot[cot_letter] = 15  # Chữ ký
            elif self.che_do_thi == "online":
                if i == 0:  # Thời gian gọi trước
                    do_rong_cot[cot_letter] = 25
                elif i == 1:  # Thời gian gọi trong ngày
                    do_rong_cot[cot_letter] = 25
                elif i == 2:  # Điểm danh
                    do_rong_cot[cot_letter] = 15
                elif i == 3:  # Thời gian nộp bài
                    do_rong_cot[cot_letter] = 25
                elif i == 4:  # Ghi chú
                    do_rong_cot[cot_letter] = 20
            elif self.che_do_thi == "hybrid":
                if i == 0:  # Chế độ thi
                    do_rong_cot[cot_letter] = 15
                elif i == 1:  # Thời gian gọi trước
                    do_rong_cot[cot_letter] = 25
                elif i == 2:  # Thời gian gọi trong ngày
                    do_rong_cot[cot_letter] = 25
                elif i == 3:  # Điểm danh
                    do_rong_cot[cot_letter] = 15
                elif i == 4:  # Thời gian nộp bài
                    do_rong_cot[cot_letter] = 25
                elif i == 5:  # Chữ ký
                    do_rong_cot[cot_letter] = 15
                elif i == 6:  # Ghi chú
                    do_rong_cot[cot_letter] = 20
        
        return do_rong_cot
    
    def ap_dung_do_rong_cot(self, ws, do_rong_cot):
        """Áp dụng độ rộng cột đã tính toán"""
        for cot_letter, width in do_rong_cot.items():
            ws.column_dimensions[cot_letter].width = width
    
    def thiet_lap_page_setup_cuoi_cung(self, ws):
        """Thiết lập page setup cuối cùng sau khi có đầy đủ dữ liệu"""
        try:
            # Đảm bảo orientation là landscape
            ws.page_setup.orientation = 'landscape'
            
            # Đảm bảo paper size là A4
            ws.page_setup.paperSize = 9  # 9 = A4
            
            # Thiết lập fit to page
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1
            
            # Thiết lập print area dựa trên dữ liệu thực tế - căn giữa nội dung
            if ws.max_row > 0:
                so_cot_bo_sung = self.tinh_so_cot_bo_sung()
                so_cot_thuc_te = len(self.cac_cot_duoc_chon) + 1 + so_cot_bo_sung  # +1 cho STT + số cột bổ sung
                cot_cuoi = get_column_letter(so_cot_thuc_te)
                # Print area bao gồm tất cả các cột dữ liệu thực tế (cần in)
                ws.print_area = f'A1:{cot_cuoi}{ws.max_row}'
            
            # Thiết lập print titles
            ws.print_title_rows = '7:7'  # Lặp lại header
            
            # Thiết lập margins - tất cả lề = 0
            ws.page_margins = PageMargins(
                left=0,        # Lề trái: 0 inch
                right=0,       # Lề phải: 0 inch
                top=0,          # Lề trên: 0 inch
                bottom=0,       # Lề dưới: 0 inch
                header=0,       # Lề header: 0 inch
                footer=0        # Lề footer: 0 inch
            )
            
            # Thiết lập căn giữa nội dung trên trang
            ws.page_setup.horizontalCentered = True  # Căn giữa theo chiều ngang
            ws.page_setup.verticalCentered = False    # Không căn giữa theo chiều dọc
            
            print(f"✅ Page setup cuối cùng đã được thiết lập:")
            print(f"   - Orientation: {ws.page_setup.orientation}")
            print(f"   - Paper size: {ws.page_setup.paperSize}")
            print(f"   - Print area: {ws.print_area}")
            print(f"   - Print titles: {ws.print_title_rows}")
            
        except Exception as e:
            print(f"❌ Lỗi khi thiết lập page setup: {str(e)}")
            import traceback
            traceback.print_exc()
        
    def tao_border_cho_bang(self, ws, hang_bat_dau, so_hang, so_cot):
        """Tạo border cho bảng dữ liệu"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(hang_bat_dau, hang_bat_dau + so_hang):
            for col in range(1, so_cot + 1):
                ws.cell(row=row, column=col).border = thin_border
                
    def tao_phan_cuoi_danh_sach_moi(self, ws, hang_bat_dau):
        """Tạo phần cuối danh sách theo mẫu mới"""
        font_bold = Font(name='Times New Roman', bold=True, size=12)
        
        # Dòng 1: Số lượng thí sinh trên danh sách
        ws.cell(row=hang_bat_dau, column=1, value="Số lượng thí sinh trên danh sách:")
        ws.cell(row=hang_bat_dau, column=1).font = font_bold
        
        # Dòng 2: Số lượng thí sinh vắng mặt
        ws.cell(row=hang_bat_dau + 1, column=1, value="Số lượng thí sinh vắng mặt:")
        ws.cell(row=hang_bat_dau + 1, column=1).font = font_bold
        
        # Dòng 3: Trống
        ws.cell(row=hang_bat_dau + 2, column=1, value="")
        
        # Dòng 4: Chữ ký giám thị
        # Chữ ký giám thị số 1
        ws.cell(row=hang_bat_dau + 3, column=2, value="Chữ ký giám thị số 1")
        ws.cell(row=hang_bat_dau + 3, column=2).font = font_bold
        ws.cell(row=hang_bat_dau + 3, column=2).alignment = Alignment(horizontal='center', vertical='center')
        
        # Chữ ký giám thị số 2
        ws.cell(row=hang_bat_dau + 3, column=6, value="Chữ ký giám thị số 2")
        ws.cell(row=hang_bat_dau + 3, column=6).font = font_bold
        ws.cell(row=hang_bat_dau + 3, column=6).alignment = Alignment(horizontal='center', vertical='center')
    
    def bo_sung_thi_sinh(self):
        """Bổ sung thí sinh vào danh sách đã chia phòng"""
        # Kiểm tra đã chia phòng chưa
        if self.df_da_chia is None:
            messagebox.showwarning("Cảnh báo", "Vui lòng thực hiện chia phòng trước khi bổ sung thí sinh!")
            return
        
        # Tạo cửa sổ mới
        window = tk.Toplevel(self.root)
        window.title("Bổ sung thí sinh vào danh sách đã chia")
        window.geometry("800x600")
        
        # Frame chính
        main_frame = ttk.Frame(window, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Thông tin hiện tại
        phong_cuoi = self.df_da_chia['Phòng thi'].iloc[-1]
        so_hs_phong_cuoi = len(self.df_da_chia[self.df_da_chia['Phòng thi'] == phong_cuoi])
        so_phong_hien_tai = len(self.df_da_chia['Phòng thi'].unique())
        
        # Lấy số phòng cuối
        try:
            # Tách số phòng từ tên phòng (ví dụ: "Phòng thi số 14" -> 14)
            so_phong_cuoi_text = phong_cuoi.split()[-1]
            so_phong_cuoi = int(so_phong_cuoi_text)
        except:
            so_phong_cuoi = so_phong_hien_tai
        
        # Hiển thị thông tin
        thong_tin = f"""📊 THÔNG TIN HIỆN TẠI:

• Tổng số thí sinh đã chia: {len(self.df_da_chia)}
• Số phòng hiện tại: {so_phong_hien_tai}
• Phòng cuối cùng: {phong_cuoi}
• Số thí sinh ở phòng cuối: {so_hs_phong_cuoi}/{self.so_thi_sinh_moi_phong}
• Số chỗ trống ở phòng cuối: {self.so_thi_sinh_moi_phong - so_hs_phong_cuoi}

💡 CÁCH THỨC BỔ SUNG:
1. Chọn file Excel chứa danh sách thí sinh mới
2. Hệ thống sẽ tự động:
   - Điền đầy phòng cuối ({phong_cuoi})
   - Tạo phòng mới nếu cần (bắt đầu từ phòng {so_phong_cuoi + 1})
   - Giữ nguyên cấu hình số lượng thí sinh/phòng"""
        
        ttk.Label(main_frame, text=thong_tin, font=("Courier New", 10), justify=tk.LEFT, foreground="blue").pack(anchor=tk.W, pady=(0, 15))
        
        # Frame chọn file
        frame_file = ttk.LabelFrame(main_frame, text="Chọn file danh sách thí sinh bổ sung", padding="10")
        frame_file.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(frame_file, text="📂 Chọn file Excel", 
                  command=lambda: self.chon_file_bo_sung(lbl_file)).pack(side=tk.LEFT, padx=(0, 10))
        
        lbl_file = ttk.Label(frame_file, text="Chưa chọn file", foreground="red")
        lbl_file.pack(side=tk.LEFT)
        
        # Lưu đường dẫn file được chọn
        self.file_bo_sung = None
        
        # Frame xem trước
        frame_preview = ttk.LabelFrame(main_frame, text="Xem trước dữ liệu bổ sung", padding="10")
        frame_preview.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        text_preview = tk.Text(frame_preview, height=15, width=80, wrap=tk.WORD)
        scrollbar_preview = ttk.Scrollbar(frame_preview, orient="vertical", command=text_preview.yview)
        text_preview.configure(yscrollcommand=scrollbar_preview.set)
        
        text_preview.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_preview.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.text_preview_bo_sung = text_preview
        
        # Frame nút điều khiển
        frame_nut = ttk.Frame(main_frame)
        frame_nut.pack(fill=tk.X)
        
        ttk.Button(frame_nut, text="✅ Thực hiện bổ sung", 
                  command=lambda: self.xu_ly_bo_sung_thi_sinh(window)).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(frame_nut, text="❌ Hủy", 
                  command=window.destroy).pack(side=tk.RIGHT)
    
    def chon_file_bo_sung(self, lbl_file):
        """Chọn file Excel chứa danh sách thí sinh bổ sung"""
        file_path = filedialog.askopenfilename(
            title="Chọn file danh sách thí sinh bổ sung",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                # Đọc tất cả sheet names
                xl_file = pd.ExcelFile(file_path, engine='openpyxl')
                sheet_names = xl_file.sheet_names
                
                # Nếu có nhiều hơn 1 sheet, cho chọn
                if len(sheet_names) > 1:
                    # Tạo dialog chọn sheet
                    sheet_window = tk.Toplevel(self.root)
                    sheet_window.title("Chọn trang tính")
                    sheet_window.geometry("400x300")
                    
                    ttk.Label(sheet_window, text=f"File có {len(sheet_names)} trang tính. Chọn trang cần lấy:", 
                             font=("Arial", 10, "bold")).pack(pady=10)
                    
                    # Listbox để chọn sheet
                    listbox = tk.Listbox(sheet_window, height=10)
                    listbox.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
                    
                    for sheet_name in sheet_names:
                        listbox.insert(tk.END, sheet_name)
                    listbox.select_set(0)  # Chọn sheet đầu tiên mặc định
                    
                    selected_sheet = [None]  # Dùng list để lưu giá trị trong closure
                    
                    def chon_sheet():
                        selection = listbox.curselection()
                        if selection:
                            selected_sheet[0] = sheet_names[selection[0]]
                            sheet_window.destroy()
                    
                    ttk.Button(sheet_window, text="✅ Chọn", command=chon_sheet).pack(pady=5)
                    
                    sheet_window.transient(self.root)
                    sheet_window.grab_set()
                    self.root.wait_window(sheet_window)
                    
                    if selected_sheet[0] is None:
                        return  # Người dùng đã hủy
                    
                    sheet_name = selected_sheet[0]
                else:
                    sheet_name = sheet_names[0]
                
                # Đọc file Excel từ sheet đã chọn
                # QUAN TRỌNG: dtype=str để giữ nguyên định dạng (số điện thoại, mã số, số 0 đầu...)
                df_bo_sung = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl', dtype=str)
                
                # Lưu thông tin
                self.file_bo_sung = file_path
                self.df_bo_sung = df_bo_sung
                
                # Cập nhật label
                lbl_file.config(text=f"✅ {os.path.basename(file_path)} [{sheet_name}] ({len(df_bo_sung)} thí sinh)", foreground="green")
                
                # Hiển thị preview
                self.text_preview_bo_sung.delete(1.0, tk.END)
                self.text_preview_bo_sung.insert(tk.END, f"📋 File: {os.path.basename(file_path)}\n")
                self.text_preview_bo_sung.insert(tk.END, f"📄 Sheet: {sheet_name}\n")
                self.text_preview_bo_sung.insert(tk.END, f"📊 Số lượng: {len(df_bo_sung)} thí sinh\n\n")
                self.text_preview_bo_sung.insert(tk.END, "=== XEM TRƯỚC 10 DÒNG ĐẦU ===\n")
                
                # Lấy các cột giống với danh sách gốc (trừ cột 'Phòng thi')
                cot_hien_thi = [col for col in self.cac_cot_duoc_chon if col in df_bo_sung.columns]
                if cot_hien_thi:
                    self.text_preview_bo_sung.insert(tk.END, df_bo_sung[cot_hien_thi].head(10).to_string(index=False))
                else:
                    self.text_preview_bo_sung.insert(tk.END, df_bo_sung.head(10).to_string(index=False))
                
                if len(df_bo_sung) > 10:
                    self.text_preview_bo_sung.insert(tk.END, f"\n... và {len(df_bo_sung) - 10} dòng khác")
                
            except Exception as e:
                import traceback
                messagebox.showerror("Lỗi", f"Không thể đọc file: {str(e)}")
    
    def xu_ly_bo_sung_thi_sinh(self, window):
        """Xử lý việc bổ sung thí sinh vào danh sách đã chia"""
        try:
            # Kiểm tra đã chọn file chưa
            if not hasattr(self, 'file_bo_sung') or self.file_bo_sung is None:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn file danh sách thí sinh bổ sung!")
                return
            
            # Lấy thông tin phòng cuối
            phong_cuoi = self.df_da_chia['Phòng thi'].iloc[-1]
            so_hs_phong_cuoi = len(self.df_da_chia[self.df_da_chia['Phòng thi'] == phong_cuoi])
            
            # Lấy tên phòng gốc và số phòng cuối
            ten_phong_goc = self.entry_ten_phong.get().strip() or "Phòng"
            try:
                so_phong_cuoi_text = phong_cuoi.split()[-1]
                so_phong_cuoi = int(so_phong_cuoi_text)
            except:
                so_phong_cuoi = len(self.df_da_chia['Phòng thi'].unique())
            
            # Tính số chỗ trống ở phòng cuối
            so_cho_trong_phong_cuoi = self.so_thi_sinh_moi_phong - so_hs_phong_cuoi
            
            # Kiểm tra cấu hình phòng cuối
            if phong_cuoi in self.cau_hinh_phong:
                so_cho_trong_phong_cuoi = self.cau_hinh_phong[phong_cuoi] - so_hs_phong_cuoi
            
            # Tạo danh sách phòng thi cho thí sinh bổ sung
            so_hs_bo_sung = len(self.df_bo_sung)
            phong_thi_bo_sung = []
            
            # Bước 1: Điền đầy phòng cuối
            so_hs_dien_phong_cuoi = min(so_cho_trong_phong_cuoi, so_hs_bo_sung)
            for i in range(so_hs_dien_phong_cuoi):
                phong_thi_bo_sung.append(phong_cuoi)
            
            # Bước 2: Tạo phòng mới nếu còn thí sinh
            so_hs_con_lai = so_hs_bo_sung - so_hs_dien_phong_cuoi
            if so_hs_con_lai > 0:
                so_phong_bat_dau = so_phong_cuoi + 1
                vi_tri_hien_tai = 0
                
                while vi_tri_hien_tai < so_hs_con_lai:
                    ten_phong_moi = f"{ten_phong_goc} {so_phong_bat_dau}"
                    
                    # Lấy số lượng thí sinh cho phòng mới
                    if ten_phong_moi in self.cau_hinh_phong:
                        so_hs_phong_moi = self.cau_hinh_phong[ten_phong_moi]
                    else:
                        so_hs_phong_moi = self.so_thi_sinh_moi_phong
                    
                    # Tính số thí sinh thực tế cho phòng này
                    so_hs_thuc_te = min(so_hs_phong_moi, so_hs_con_lai - vi_tri_hien_tai)
                    
                    # Gán phòng
                    for i in range(so_hs_thuc_te):
                        phong_thi_bo_sung.append(ten_phong_moi)
                    
                    vi_tri_hien_tai += so_hs_thuc_te
                    so_phong_bat_dau += 1
            
            # Gán phòng thi cho df bổ sung
            self.df_bo_sung['Phòng thi'] = phong_thi_bo_sung
            
            # Ghép vào danh sách chính
            self.df_da_chia = pd.concat([self.df_da_chia, self.df_bo_sung], ignore_index=True)
            
            # Hiển thị kết quả
            messagebox.showinfo("Thành công", 
                               f"✅ Đã bổ sung {so_hs_bo_sung} thí sinh!\n\n"
                               f"• Điền vào {phong_cuoi}: {so_hs_dien_phong_cuoi} thí sinh\n"
                               f"• Tạo phòng mới: {so_hs_con_lai} thí sinh\n\n"
                               f"Tổng số thí sinh hiện tại: {len(self.df_da_chia)}")
            
            # Cập nhật hiển thị kết quả
            self.hien_thi_ket_qua()
            
            # Đóng cửa sổ
            window.destroy()
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể bổ sung thí sinh: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def chia_lai_tu_phong_x(self):
        """Giữ nguyên X phòng đầu, chia lại từ phòng X+1"""
        # Kiểm tra đã chia phòng chưa
        if self.df_da_chia is None:
            messagebox.showwarning("Cảnh báo", "Vui lòng thực hiện chia phòng trước!")
            return
        
        # Tạo cửa sổ mới
        window = tk.Toplevel(self.root)
        window.title("Chia lại từ phòng X")
        window.geometry("700x500")
        
        # Frame chính
        main_frame = ttk.Frame(window, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Thông tin hiện tại
        so_phong_hien_tai = len(self.df_da_chia['Phòng thi'].unique())
        danh_sach_phong = sorted(self.df_da_chia['Phòng thi'].unique())
        
        # Tiêu đề
        ttk.Label(main_frame, text="🔄 CHIA LẠI TỪ PHÒNG X", 
                 font=("Arial", 14, "bold"), foreground="blue").pack(pady=(0, 20))
        
        # Thông tin
        info = f"""📊 THÔNG TIN HIỆN TẠI:

• Tổng số thí sinh: {len(self.df_da_chia)}
• Số phòng hiện tại: {so_phong_hien_tai}
• Danh sách phòng: {', '.join(danh_sach_phong[:5])}{'...' if len(danh_sach_phong) > 5 else ''}

💡 CÁCH THỨC:
Giữ nguyên các phòng từ đầu đến phòng X-1, chia lại từ phòng X trở đi theo cấu hình mới."""
        
        ttk.Label(main_frame, text=info, font=("Courier New", 10), justify=tk.LEFT).pack(anchor=tk.W, pady=(0, 15))
        
        # Label ví dụ động
        self.lbl_vi_du = ttk.Label(main_frame, text="", font=("Courier New", 10), 
                                    justify=tk.LEFT, foreground="blue")
        self.lbl_vi_du.pack(anchor=tk.W, pady=(0, 20))
        
        # Frame chọn phòng bắt đầu
        frame_chon = ttk.LabelFrame(main_frame, text="Chọn phòng bắt đầu chia lại", padding="10")
        frame_chon.pack(fill=tk.X, pady=(0, 20))
        
        ttk.Label(frame_chon, text="Bắt đầu từ phòng:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        
        # Dropdown chọn phòng
        combo_phong = ttk.Combobox(frame_chon, values=danh_sach_phong, state="readonly", width=30)
        combo_phong.grid(row=0, column=1, sticky=(tk.W, tk.E))
        combo_phong.current(0)
        
        # Hiển thị thống kê khi chọn phòng
        lbl_thong_ke = ttk.Label(frame_chon, text="", font=("Arial", 9), foreground="green")
        lbl_thong_ke.grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=(10, 0))
        
        def cap_nhat_thong_ke(event=None):
            phong_chon = combo_phong.get()
            # Tìm index của phòng được chọn
            idx = danh_sach_phong.index(phong_chon)
            so_phong_giu = idx
            so_phong_chia_lai = so_phong_hien_tai - idx
            
            # Đếm số thí sinh
            so_hs_giu = len(self.df_da_chia[self.df_da_chia['Phòng thi'].isin(danh_sach_phong[:idx])])
            so_hs_chia_lai = len(self.df_da_chia[self.df_da_chia['Phòng thi'].isin(danh_sach_phong[idx:])])
            
            thong_ke = f"→ Giữ nguyên: {so_phong_giu} phòng ({so_hs_giu} thí sinh)\n"
            thong_ke += f"→ Chia lại: {so_phong_chia_lai} phòng ({so_hs_chia_lai} thí sinh)"
            lbl_thong_ke.config(text=thong_ke)
            
            # Cập nhật ví dụ động
            if idx == 0:
                vi_du = f"\n⚠️ LƯU Ý: Bạn đang chọn phòng đầu tiên!\n"
                vi_du += f"→ Sẽ CHIA LẠI TOÀN BỘ danh sách từ đầu"
            else:
                # Lấy danh sách phòng giữ nguyên
                phong_giu_str = ', '.join(danh_sach_phong[:min(idx, 3)])
                if idx > 3:
                    phong_giu_str += f', ... {danh_sach_phong[idx-1]}'
                
                vi_du = f"\nVÍ DỤ: Chọn '{phong_chon}'\n"
                vi_du += f"→ Giữ nguyên: {phong_giu_str}\n"
                vi_du += f"→ Chia lại: Từ {phong_chon} trở đi"
            
            self.lbl_vi_du.config(text=vi_du)
        
        combo_phong.bind('<<ComboboxSelected>>', cap_nhat_thong_ke)
        cap_nhat_thong_ke()  # Hiển thị ban đầu
        
        # Frame nút
        frame_nut = ttk.Frame(main_frame)
        frame_nut.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Button(frame_nut, text="✅ Thực hiện chia lại", 
                  command=lambda: self.xu_ly_chia_lai_tu_phong(combo_phong.get(), danh_sach_phong, window)).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(frame_nut, text="❌ Hủy", 
                  command=window.destroy).pack(side=tk.RIGHT)
    
    def xu_ly_chia_lai_tu_phong(self, phong_bat_dau, danh_sach_phong, window):
        """Xử lý việc chia lại từ phòng X"""
        try:
            # Tìm index của phòng bắt đầu
            idx = danh_sach_phong.index(phong_bat_dau)
            
            # CẢNH BÁO nếu chọn phòng đầu tiên (chia lại toàn bộ)
            if idx == 0:
                xac_nhan = messagebox.askyesno(
                    "Xác nhận", 
                    f"⚠️ Bạn đang chọn chia lại từ {phong_bat_dau} (phòng đầu tiên)!\n\n"
                    f"Điều này sẽ CHIA LẠI TOÀN BỘ danh sách từ đầu.\n\n"
                    f"Bạn có chắc chắn muốn tiếp tục?",
                    icon='warning'
                )
                if not xac_nhan:
                    return
            
            # Lấy các phòng giữ nguyên và phòng cần chia lại
            phong_giu_nguyen = danh_sach_phong[:idx]
            phong_chia_lai = danh_sach_phong[idx:]
            
            # Tách dữ liệu - GIỮ NGUYÊN THỨ TỰ BAN ĐẦU
            df_giu_nguyen = self.df_da_chia[self.df_da_chia['Phòng thi'].isin(phong_giu_nguyen)].copy()
            df_can_chia_lai = self.df_da_chia[self.df_da_chia['Phòng thi'].isin(phong_chia_lai)].copy()
            
            # LƯU THỨ TỰ GốC để khôi phục sau khi chia
            df_can_chia_lai = df_can_chia_lai.reset_index(drop=False).rename(columns={'index': 'thu_tu_goc'})
            
            # Lấy tên phòng gốc
            ten_phong_goc = self.entry_ten_phong.get().strip() or "Phòng"
            
            # Tính số phòng bắt đầu chia lại
            import re
            match = re.search(r'(\d+)', phong_bat_dau)
            if match:
                so_phong_bat_dau = int(match.group(1))
            else:
                so_phong_bat_dau = idx + 1
            
            # Chia lại thí sinh từ phòng X
            so_hs_can_chia = len(df_can_chia_lai)
            phong_thi_moi = []
            vi_tri_hien_tai = 0
            so_phong = so_phong_bat_dau
            
            # Tự động phát hiện format từ cấu hình
            co_so_0_dau = False
            if self.cau_hinh_phong:
                ten_phong_mau = list(self.cau_hinh_phong.keys())[0]
                if "01" in ten_phong_mau or "02" in ten_phong_mau:
                    co_so_0_dau = True
                match_prefix = re.match(r'(.+?)\s*(\d+)$', ten_phong_mau)
                if match_prefix:
                    ten_phong_goc = match_prefix.group(1).strip()
            
            while vi_tri_hien_tai < so_hs_can_chia:
                # Tạo tên phòng
                if co_so_0_dau:
                    ten_phong = f"{ten_phong_goc} {so_phong:02d}"
                else:
                    ten_phong = f"{ten_phong_goc} {so_phong}"
                
                # Lấy số lượng từ cấu hình
                if ten_phong in self.cau_hinh_phong:
                    so_hs_phong = self.cau_hinh_phong[ten_phong]
                else:
                    so_hs_phong = self.so_thi_sinh_moi_phong
                
                # Tính số thí sinh thực tế
                so_hs_con_lai = so_hs_can_chia - vi_tri_hien_tai
                so_hs_thuc_te = min(so_hs_phong, so_hs_con_lai)
                
                # Gán phòng
                for i in range(so_hs_thuc_te):
                    phong_thi_moi.append(ten_phong)
                
                vi_tri_hien_tai += so_hs_thuc_te
                so_phong += 1
            
            # Gán phòng mới cho df
            df_can_chia_lai['Phòng thi'] = phong_thi_moi
            
            # KHÔI PHỤC THỨ TỰ GỐC sau khi gán phòng mới
            df_can_chia_lai = df_can_chia_lai.sort_values('thu_tu_goc').drop(columns=['thu_tu_goc'])
            df_can_chia_lai = df_can_chia_lai.reset_index(drop=True)
            
            # Ghép lại - GIỮ NGUYÊN THỨ TỰ
            self.df_da_chia = pd.concat([df_giu_nguyen, df_can_chia_lai], ignore_index=True)
            
            # Thông báo
            messagebox.showinfo("Thành công", 
                               f"✅ Đã chia lại thành công!\n\n"
                               f"• Giữ nguyên: {len(df_giu_nguyen)} thí sinh ({len(phong_giu_nguyen)} phòng)\n"
                               f"• Chia lại: {len(df_can_chia_lai)} thí sinh\n"
                               f"• Số phòng mới: {so_phong - so_phong_bat_dau}\n\n"
                               f"Tổng số thí sinh: {len(self.df_da_chia)}")
            
            # Cập nhật hiển thị
            self.hien_thi_ket_qua()
            
            # Đóng cửa sổ
            window.destroy()
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể chia lại: {str(e)}")
            import traceback
            traceback.print_exc()
        
    def chay(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = ChiaPhongThi()
    app.chay()