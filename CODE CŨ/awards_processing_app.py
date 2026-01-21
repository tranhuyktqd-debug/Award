# -*- coding: utf-8 -*-
"""
ỨNG DỤNG XỬ LÝ MÃ CERT VÀ TẠO BÁO CÁO
Tích hợp toàn bộ quy trình trong 1 app
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import pandas as pd
import numpy as np
import os
import sys
import threading
from datetime import datetime
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
try:
    import qrcode
    from PIL import Image, ImageTk
    HAS_QR_PIL = True
except ImportError:
    HAS_QR_PIL = False
    print("Warning: qrcode hoặc PIL không được cài đặt. Tính năng QR/Photo sẽ bị giới hạn.")

class AwardsProcessingApp:
    def __init__(self, root):
        self.root = root
        self.root.title("🎓 HỆ THỐNG XỬ LÝ MÃ CERT ASMO")
        
        # Cấu hình cửa sổ và đặt ở giữa màn hình
        window_width = 900
        window_height = 700
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
        self.root.resizable(False, False)
        
        # Biến lưu trữ đường dẫn file
        self.file_full = tk.StringVar()
        self.file_trao_giai = tk.StringVar()
        self.output_dir = tk.StringVar(value=os.getcwd())
        
        # Biến trạng thái
        self.is_processing = False
        
        # === THÊM CHO TAB 2: CHIA DANH SÁCH ===
        self.df_nguon = None  # DataFrame nguồn từ Awards_Comparison_WITH_CERT.xlsx
        self.df_da_chia = None  # DataFrame đã chia theo STT túi
        self.checkboxes_cot = {}  # Dictionary lưu các checkbox cột
        self.cac_cot_mac_dinh = {}  # Dictionary lưu cấu hình các cột
        self.excel_sheets = []  # Danh sách sheet trong file Excel
        self.current_sheet = None  # Sheet hiện tại được chọn
        
        # === THÊM CHO TAB 3: TRA CỨU ===
        self.df_tracuu = None  # DataFrame dữ liệu tra cứu
        self.current_results = []  # Kết quả tìm kiếm hiện tại
        self.file_tracuu_var = tk.StringVar()  # Đường dẫn file tra cứu
        self.tracuu_sheets = []  # Danh sách sheet trong file
        self.sheet_checkboxes = {}  # Dictionary lưu checkbox các sheet
        
        # Tạo giao diện
        self.create_widgets()
        
    def create_widgets(self):
        """Tạo các widget cho giao diện với Notebook (Tabs)"""
        
        # ========== HEADER ==========
        header_frame = tk.Frame(self.root, bg="#2c3e50", height=80)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        title_label = tk.Label(
            header_frame,
            text="🎓 HỆ THỐNG XỬ LÝ MÃ CERT ASMO",
            font=("Arial", 18, "bold"),
            bg="#2c3e50",
            fg="white"
        )
        title_label.pack(pady=20)
        
        # ========== NOTEBOOK (TABS) ==========
        # Configure tab style for better visibility
        style = ttk.Style()
        style.theme_use('default')
        
        # Configure colors for tabs
        style.configure('TNotebook', background='#ecf0f1', borderwidth=0)
        style.configure('TNotebook.Tab', 
                       background='#bdc3c7',
                       foreground='#2c3e50',
                       padding=[20, 10],
                       font=('Arial', 10, 'bold'))
        
        # Selected tab (active)
        style.map('TNotebook.Tab',
                 background=[('selected', '#3498db')],
                 foreground=[('selected', 'white')],
                 expand=[('selected', [1, 1, 1, 0])])
        
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Tab 1: Xử lý Mã Cert
        self.tab1 = tk.Frame(self.notebook, bg="#ecf0f1")
        self.notebook.add(self.tab1, text="📋 Xử lý Mã Cert")
        self.create_tab1_content(self.tab1)
        
        # Tab 2: Chia danh sách
        self.tab2 = tk.Frame(self.notebook, bg="#ecf0f1")
        self.notebook.add(self.tab2, text="📦 Chia danh sách")
        self.create_tab2_content(self.tab2)
        
        # Tab 3: Tra cứu
        self.tab3 = tk.Frame(self.notebook, bg="#ecf0f1")
        self.notebook.add(self.tab3, text="🔍 Tra cứu")
        self.create_tab3_content(self.tab3)
    
    def create_tab1_content(self, parent):
        """Tạo nội dung Tab 1: Xử lý Mã Cert (code cũ)"""
        main_frame = tk.Frame(parent, bg="#ecf0f1", padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # --- File Input Section ---
        input_frame = tk.LabelFrame(
            main_frame,
            text="📂 CHỌN FILE ĐẦU VÀO",
            font=("Arial", 12, "bold"),
            bg="#ecf0f1",
            fg="#2c3e50",
            padx=15,
            pady=10
        )
        input_frame.pack(fill=tk.X, pady=(0, 15))
        
        # File 1: Awards_Template_Full.xlsx
        tk.Label(
            input_frame,
            text="File đầy đủ (Awards_Template_Full.xlsx):",
            font=("Arial", 10),
            bg="#ecf0f1"
        ).grid(row=0, column=0, sticky=tk.W, pady=5)
        
        file1_entry = tk.Entry(
            input_frame,
            textvariable=self.file_full,
            width=50,
            font=("Arial", 9)
        )
        file1_entry.grid(row=0, column=1, padx=10, pady=5)
        
        tk.Button(
            input_frame,
            text="Chọn file",
            command=lambda: self.browse_file(self.file_full),
            bg="#3498db",
            fg="white",
            font=("Arial", 9, "bold"),
            cursor="hand2"
        ).grid(row=0, column=2, pady=5)
        
        # File 2: Awards_TRAO GIAI.xlsx
        tk.Label(
            input_frame,
            text="File trao giải (Awards_TRAO GIAI.xlsx):",
            font=("Arial", 10),
            bg="#ecf0f1"
        ).grid(row=1, column=0, sticky=tk.W, pady=5)
        
        file2_entry = tk.Entry(
            input_frame,
            textvariable=self.file_trao_giai,
            width=50,
            font=("Arial", 9)
        )
        file2_entry.grid(row=1, column=1, padx=10, pady=5)
        
        tk.Button(
            input_frame,
            text="Chọn file",
            command=lambda: self.browse_file(self.file_trao_giai),
            bg="#3498db",
            fg="white",
            font=("Arial", 9, "bold"),
            cursor="hand2"
        ).grid(row=1, column=2, pady=5)
        
        # Output Directory
        tk.Label(
            input_frame,
            text="Thư mục lưu kết quả:",
            font=("Arial", 10),
            bg="#ecf0f1"
        ).grid(row=2, column=0, sticky=tk.W, pady=5)
        
        output_entry = tk.Entry(
            input_frame,
            textvariable=self.output_dir,
            width=50,
            font=("Arial", 9)
        )
        output_entry.grid(row=2, column=1, padx=10, pady=5)
        
        tk.Button(
            input_frame,
            text="Chọn thư mục",
            command=self.browse_directory,
            bg="#3498db",
            fg="white",
            font=("Arial", 9, "bold"),
            cursor="hand2"
        ).grid(row=2, column=2, pady=5)
        
        # --- Progress Section ---
        progress_frame = tk.LabelFrame(
            main_frame,
            text="📊 TIẾN TRÌNH XỬ LÝ",
            font=("Arial", 12, "bold"),
            bg="#ecf0f1",
            fg="#2c3e50",
            padx=15,
            pady=10
        )
        progress_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        # Progress bar
        self.progress = ttk.Progressbar(
            progress_frame,
            mode='determinate',
            length=800
        )
        self.progress.pack(fill=tk.X, pady=5)
        
        # Log text area
        self.log_text = scrolledtext.ScrolledText(
            progress_frame,
            height=10,
            width=95,
            font=("Consolas", 9),
            bg="#2c3e50",
            fg="#2ecc71",
            insertbackground="white"
        )
        self.log_text.pack(fill=tk.BOTH, expand=True, pady=(5, 0))
        
        # --- Action Buttons ---
        button_frame = tk.Frame(main_frame, bg="#ecf0f1")
        button_frame.pack(fill=tk.X, pady=10)
        
        # Left side buttons
        left_buttons = tk.Frame(button_frame, bg="#ecf0f1")
        left_buttons.pack(side=tk.LEFT)
        
        self.process_btn = tk.Button(
            left_buttons,
            text="▶ BẮT ĐẦU XỬ LÝ",
            command=self.start_processing,
            bg="#27ae60",
            fg="white",
            font=("Arial", 11, "bold"),
            height=1,
            width=18,
            cursor="hand2",
            relief=tk.RAISED,
            bd=3
        )
        self.process_btn.pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            left_buttons,
            text="🗑 XÓA LOG",
            command=self.clear_log,
            bg="#e67e22",
            fg="white",
            font=("Arial", 11, "bold"),
            height=1,
            width=12,
            cursor="hand2",
            relief=tk.RAISED,
            bd=3
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            left_buttons,
            text="📁 MỞ THƯ MỤC",
            command=self.open_output_folder,
            bg="#3498db",
            fg="white",
            font=("Arial", 11, "bold"),
            height=1,
            width=14,
            cursor="hand2",
            relief=tk.RAISED,
            bd=3
        ).pack(side=tk.LEFT, padx=5)
    
    def create_tab2_content(self, parent):
        """Tạo Tab 2: Chia danh sách theo STT túi"""
        
        main_frame = tk.Frame(parent, bg="#ecf0f1", padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # --- 1. Chọn file nguồn ---
        frame_file = tk.LabelFrame(
            main_frame,
            text="📂 CHỌN FILE NGUỒN",
            font=("Arial", 12, "bold"),
            bg="#ecf0f1",
            fg="#2c3e50",
            padx=15,
            pady=10
        )
        frame_file.pack(fill=tk.X, pady=(0, 15))
        
        self.file_nguon_var = tk.StringVar()
        default_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Awards_Comparison_WITH_CERT.xlsx")
        self.file_nguon_var.set(default_path)
        
        tk.Label(frame_file, text="File nguồn:", font=("Arial", 10), bg="#ecf0f1").grid(row=0, column=0, sticky=tk.W, pady=5)
        entry_file = tk.Entry(frame_file, textvariable=self.file_nguon_var, width=70, font=("Arial", 9))
        entry_file.grid(row=0, column=1, padx=10, pady=5)
        
        tk.Button(
            frame_file,
            text="📂 Chọn file",
            command=lambda: self.browse_file_tab2(self.file_nguon_var),
            bg="#3498db",
            fg="white",
            font=("Arial", 9, "bold"),
            cursor="hand2"
        ).grid(row=0, column=2, pady=5)
        
        tk.Button(
            frame_file,
            text="📖 Đọc file",
            command=self.load_file_nguon,
            bg="#27ae60",
            fg="white",
            font=("Arial", 9, "bold"),
            cursor="hand2"
        ).grid(row=0, column=3, padx=5, pady=5)
        
        # Chọn sheet
        tk.Label(frame_file, text="Chọn sheet:", font=("Arial", 10), bg="#ecf0f1").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.combo_sheet = ttk.Combobox(frame_file, state="readonly", width=68, font=("Arial", 9))
        self.combo_sheet.grid(row=1, column=1, padx=10, pady=5)
        self.combo_sheet.bind("<<ComboboxSelected>>", self.on_sheet_selected)
        
        tk.Button(
            frame_file,
            text="🔄 Load sheet",
            command=self.load_selected_sheet,
            bg="#e67e22",
            fg="white",
            font=("Arial", 9, "bold"),
            cursor="hand2"
        ).grid(row=1, column=2, pady=5, columnspan=2, sticky=tk.W)
        
        # --- 2. Chọn các cột ---
        frame_cot = tk.LabelFrame(
            main_frame,
            text="📋 CHỌN CÁC CỘT CẦN XUẤT",
            font=("Arial", 12, "bold"),
            bg="#ecf0f1",
            fg="#2c3e50",
            padx=15,
            pady=10
        )
        frame_cot.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        # Frame chứa checkbox với scrollbar
        canvas_cot = tk.Canvas(frame_cot, bg="#ecf0f1", height=120)
        scrollbar_cot = ttk.Scrollbar(frame_cot, orient="vertical", command=canvas_cot.yview)
        frame_checkbox = tk.Frame(canvas_cot, bg="#ecf0f1")
        
        frame_checkbox.bind(
            "<Configure>",
            lambda e: canvas_cot.configure(scrollregion=canvas_cot.bbox("all"))
        )
        
        canvas_cot.create_window((0, 0), window=frame_checkbox, anchor="nw")
        canvas_cot.configure(yscrollcommand=scrollbar_cot.set)
        
        # Danh sách các cột mặc định với mapping
        self.cac_cot_mac_dinh = {
            'SBD': {'var': tk.BooleanVar(value=True), 'hien_thi': 'SBD'},
            'FULL NAME': {'var': tk.BooleanVar(value=True), 'hien_thi': 'FULL NAME'},
            'Ngày sinh': {'var': tk.BooleanVar(value=True), 'hien_thi': 'Ngày sinh / D.O.B'},
            'KHỐI': {'var': tk.BooleanVar(value=True), 'hien_thi': 'KHỐI'},
            'TRƯỜNG': {'var': tk.BooleanVar(value=True), 'hien_thi': 'TRƯỜNG'},
            'KQ VQG TOÁN': {'var': tk.BooleanVar(value=True), 'hien_thi': 'TOÁN (Kết quả 1)'},
            'KQ VQG KHOA HỌC': {'var': tk.BooleanVar(value=True), 'hien_thi': 'Khoa học (Kết quả 2)'},
            'KQ VQG TIẾNG ANH': {'var': tk.BooleanVar(value=True), 'hien_thi': 'TA (Kết quả 3)'},
            'MÃ CERT ĐẦY ĐỦ': {'var': tk.BooleanVar(value=False), 'hien_thi': 'Mã Cert (đầy đủ)'},
            'MÃ CERT': {'var': tk.BooleanVar(value=True), 'hien_thi': 'Mã Cert (rút gọn)'},
            'SL GCN': {'var': tk.BooleanVar(value=True), 'hien_thi': 'SL GCN'},
            'STT TÚI': {'var': tk.BooleanVar(value=True), 'hien_thi': 'STT TÚI'}
        }
        
        # Tạo checkbox
        self.checkboxes_cot = {}
        col = 0
        row = 0
        for cot, config in self.cac_cot_mac_dinh.items():
            cb = tk.Checkbutton(
                frame_checkbox,
                text=config['hien_thi'],
                variable=config['var'],
                bg="#ecf0f1",
                font=("Arial", 9),
                anchor=tk.W
            )
            cb.grid(row=row, column=col, sticky=tk.W, padx=10, pady=5)
            self.checkboxes_cot[cot] = cb
            
            col += 1
            if col >= 4:
                col = 0
                row += 1
        
        canvas_cot.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_cot.pack(side=tk.RIGHT, fill=tk.Y)
        
        # --- 3. Nút xử lý ---
        frame_button = tk.Frame(main_frame, bg="#ecf0f1")
        frame_button.pack(fill=tk.X, pady=(0, 15))
        
        tk.Button(
            frame_button,
            text="▶ CHIA DANH SÁCH THEO STT TÚI",
            command=self.chia_danh_sach,
            bg="#27ae60",
            fg="white",
            font=("Arial", 12, "bold"),
            height=2,
            width=35,
            cursor="hand2"
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            frame_button,
            text="💾 XUẤT KẾT QUẢ",
            command=self.xuat_ket_qua_chia,
            bg="#3498db",
            fg="white",
            font=("Arial", 12, "bold"),
            height=2,
            width=20,
            cursor="hand2"
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            frame_button,
            text="📁 MỞ THƯ MỤC",
            command=self.open_output_folder,
            bg="#95a5a6",
            fg="white",
            font=("Arial", 12, "bold"),
            height=2,
            width=15,
            cursor="hand2"
        ).pack(side=tk.LEFT, padx=5)
        
        # --- 4. Hiển thị kết quả ---
        frame_ket_qua = tk.LabelFrame(
            main_frame,
            text="📊 KẾT QUẢ",
            font=("Arial", 12, "bold"),
            bg="#ecf0f1",
            fg="#2c3e50",
            padx=15,
            pady=10
        )
        frame_ket_qua.pack(fill=tk.BOTH, expand=True)
        
        self.text_ket_qua_chia = scrolledtext.ScrolledText(
            frame_ket_qua,
            height=12,
            width=95,
            font=("Consolas", 9),
            bg="#2c3e50",
            fg="#2ecc71",
            insertbackground="white"
        )
        self.text_ket_qua_chia.pack(fill=tk.BOTH, expand=True)
        
        # Hiển thị hướng dẫn ban đầu
        self.text_ket_qua_chia.insert(tk.END, "📖 HƯỚNG DẪN:\n")
        self.text_ket_qua_chia.insert(tk.END, "1. Chọn file Awards_Comparison_WITH_CERT.xlsx\n")
        self.text_ket_qua_chia.insert(tk.END, "2. Nhấn '📖 Đọc file' để tải dữ liệu\n")
        self.text_ket_qua_chia.insert(tk.END, "3. Chọn các cột cần xuất\n")
        self.text_ket_qua_chia.insert(tk.END, "4. Nhấn '▶ CHIA DANH SÁCH THEO STT TÚI'\n")
        self.text_ket_qua_chia.insert(tk.END, "5. Nhấn '💾 XUẤT KẾT QUẢ' để lưu file Excel\n\n")
    
    # ========== CÁC HÀM XỬ LÝ CHO TAB 2 ==========
    
    def browse_file_tab2(self, var):
        """Chọn file cho Tab 2"""
        filename = filedialog.askopenfilename(
            title="Chọn file Excel nguồn",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            var.set(filename)
            self.text_ket_qua_chia.insert(tk.END, f"✅ Đã chọn file: {os.path.basename(filename)}\n")
            self.text_ket_qua_chia.see(tk.END)
    
    def load_file_nguon(self):
        """Đọc file nguồn Awards_Comparison_WITH_CERT.xlsx và hiển thị danh sách sheet"""
        try:
            file_path = self.file_nguon_var.get()
            if not file_path or not os.path.exists(file_path):
                messagebox.showerror("Lỗi", "File không tồn tại!")
                return
            
            self.text_ket_qua_chia.delete(1.0, tk.END)
            self.text_ket_qua_chia.insert(tk.END, "⏳ Đang đọc file...\n")
            self.text_ket_qua_chia.update()
            
            # Đọc file Excel (có thể có nhiều sheet)
            xls = pd.ExcelFile(file_path)
            self.excel_sheets = xls.sheet_names
            
            # Cập nhật combobox với danh sách sheet
            self.combo_sheet['values'] = self.excel_sheets
            
            # Ưu tiên sheet 'TRAO GIẢI', nếu không có thì lấy sheet đầu tiên
            if 'TRAO GIẢI' in self.excel_sheets:
                sheet_name = 'TRAO GIẢI'
            else:
                sheet_name = self.excel_sheets[0]
            
            self.combo_sheet.set(sheet_name)
            self.current_sheet = sheet_name
            
            # Load dữ liệu từ sheet được chọn
            self.df_nguon = pd.read_excel(file_path, sheet_name=sheet_name, dtype={'SBD': str})
            
            # Kiểm tra có cột STT TÚI không
            if 'STT TÚI' not in self.df_nguon.columns:
                messagebox.showwarning(
                    "Cảnh báo", 
                    "File không có cột 'STT TÚI'!\n\n"
                    "Cần chạy Tab 'Xử lý Mã Cert' để tạo STT TÚI trước."
                )
                self.text_ket_qua_chia.insert(tk.END, "❌ File thiếu cột 'STT TÚI'!\n")
                self.text_ket_qua_chia.insert(tk.END, "   Vui lòng chạy Tab 'Xử lý Mã Cert' trước.\n")
                return
            
            # Cập nhật trạng thái checkbox theo cột thực tế
            self.update_checkbox_status()
            
            # Thông báo thành công
            self.text_ket_qua_chia.insert(tk.END, f"✅ Đã đọc file thành công!\n\n")
            self.text_ket_qua_chia.insert(tk.END, f"📄 File: {os.path.basename(file_path)}\n")
            self.text_ket_qua_chia.insert(tk.END, f"📑 Tổng số sheet: {len(self.excel_sheets)} ({', '.join(self.excel_sheets)})\n")
            self.text_ket_qua_chia.insert(tk.END, f"📋 Sheet đang xem: {sheet_name}\n")
            self.text_ket_qua_chia.insert(tk.END, f"📊 Số dòng: {len(self.df_nguon):,}\n")
            self.text_ket_qua_chia.insert(tk.END, f"📊 Số cột: {len(self.df_nguon.columns)}\n\n")
            
            # Hiển thị các túi
            if 'STT TÚI' in self.df_nguon.columns:
                thong_ke_tui = self.df_nguon['STT TÚI'].value_counts().sort_index()
                tui_co_du_lieu = thong_ke_tui[thong_ke_tui.index > 0]
                if len(tui_co_du_lieu) > 0:
                    self.text_ket_qua_chia.insert(tk.END, f"📦 Số túi: {len(tui_co_du_lieu)}\n")
                    self.text_ket_qua_chia.insert(tk.END, f"📦 Túi từ {int(tui_co_du_lieu.index.min())} đến {int(tui_co_du_lieu.index.max())}\n\n")
            
            messagebox.showinfo(
                "Thành công", 
                f"Đã đọc file thành công!\n\n"
                f"Tổng số sheet: {len(self.excel_sheets)}\n"
                f"Sheet đang xem: {sheet_name}\n"
                f"Số dòng: {len(self.df_nguon):,}\n"
                f"Số cột: {len(self.df_nguon.columns)}\n\n"
                f"💡 Có thể chọn sheet khác từ dropdown!"
            )
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể đọc file:\n{str(e)}")
            self.text_ket_qua_chia.insert(tk.END, f"\n❌ LỖI: {str(e)}\n")
            import traceback
            self.text_ket_qua_chia.insert(tk.END, traceback.format_exc())
    
    def update_checkbox_status(self):
        """Cập nhật trạng thái checkbox theo các cột thực tế trong file"""
        if self.df_nguon is None:
            return
        
        all_columns = list(self.df_nguon.columns)
        
        # Cập nhật trạng thái các checkbox
        for cot, config in self.cac_cot_mac_dinh.items():
            if cot in all_columns:
                self.checkboxes_cot[cot].config(state=tk.NORMAL)
            else:
                # Tắt checkbox nếu cột không tồn tại
                config['var'].set(False)
                self.checkboxes_cot[cot].config(state=tk.DISABLED, fg="gray")
    
    def on_sheet_selected(self, event=None):
        """Xử lý khi chọn sheet từ combobox"""
        selected_sheet = self.combo_sheet.get()
        if selected_sheet != self.current_sheet:
            self.text_ket_qua_chia.insert(tk.END, f"\n💡 Đã chọn sheet: {selected_sheet}\n")
            self.text_ket_qua_chia.insert(tk.END, f"   Nhấn '🔄 Load sheet' để tải dữ liệu từ sheet này.\n")
            self.text_ket_qua_chia.see(tk.END)
    
    def load_selected_sheet(self):
        """Load dữ liệu từ sheet được chọn"""
        try:
            selected_sheet = self.combo_sheet.get()
            if not selected_sheet:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn sheet!")
                return
            
            file_path = self.file_nguon_var.get()
            if not file_path or not os.path.exists(file_path):
                messagebox.showerror("Lỗi", "File không tồn tại!")
                return
            
            self.text_ket_qua_chia.delete(1.0, tk.END)
            self.text_ket_qua_chia.insert(tk.END, f"⏳ Đang load sheet '{selected_sheet}'...\n")
            self.text_ket_qua_chia.update()
            
            # Load dữ liệu từ sheet được chọn
            self.df_nguon = pd.read_excel(file_path, sheet_name=selected_sheet, dtype={'SBD': str})
            self.current_sheet = selected_sheet
            
            # Kiểm tra có cột STT TÚI không
            if 'STT TÚI' not in self.df_nguon.columns:
                messagebox.showwarning(
                    "Cảnh báo", 
                    f"Sheet '{selected_sheet}' không có cột 'STT TÚI'!\n\n"
                    "Cần chạy Tab 'Xử lý Mã Cert' để tạo STT TÚI trước."
                )
                self.text_ket_qua_chia.insert(tk.END, f"❌ Sheet '{selected_sheet}' thiếu cột 'STT TÚI'!\n")
                self.text_ket_qua_chia.insert(tk.END, "   Vui lòng chọn sheet khác hoặc chạy Tab 'Xử lý Mã Cert' trước.\n")
                return
            
            # Cập nhật trạng thái checkbox theo cột thực tế
            self.update_checkbox_status()
            
            # Thông báo thành công
            self.text_ket_qua_chia.insert(tk.END, f"✅ Đã load sheet '{selected_sheet}' thành công!\n\n")
            self.text_ket_qua_chia.insert(tk.END, f"📋 Sheet: {selected_sheet}\n")
            self.text_ket_qua_chia.insert(tk.END, f"📊 Số dòng: {len(self.df_nguon):,}\n")
            self.text_ket_qua_chia.insert(tk.END, f"📊 Số cột: {len(self.df_nguon.columns)}\n\n")
            
            # Hiển thị các túi
            if 'STT TÚI' in self.df_nguon.columns:
                thong_ke_tui = self.df_nguon['STT TÚI'].value_counts().sort_index()
                tui_co_du_lieu = thong_ke_tui[thong_ke_tui.index > 0]
                if len(tui_co_du_lieu) > 0:
                    self.text_ket_qua_chia.insert(tk.END, f"📦 Số túi: {len(tui_co_du_lieu)}\n")
                    self.text_ket_qua_chia.insert(tk.END, f"📦 Túi từ {int(tui_co_du_lieu.index.min())} đến {int(tui_co_du_lieu.index.max())}\n\n")
            
            messagebox.showinfo(
                "Thành công", 
                f"Đã load sheet '{selected_sheet}' thành công!\n\n"
                f"Số dòng: {len(self.df_nguon):,}\n"
                f"Số cột: {len(self.df_nguon.columns)}"
            )
            
            # Reset dữ liệu đã chia
            self.df_da_chia = None
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể load sheet:\n{str(e)}")
            self.text_ket_qua_chia.insert(tk.END, f"\n❌ LỖI: {str(e)}\n")
            import traceback
            self.text_ket_qua_chia.insert(tk.END, traceback.format_exc())
    
    def chia_danh_sach(self):
        """Chia danh sách theo STT túi"""
        try:
            if self.df_nguon is None:
                messagebox.showwarning("Cảnh báo", "Vui lòng đọc file nguồn trước!")
                return
            
            # Kiểm tra có cột STT TÚI
            if 'STT TÚI' not in self.df_nguon.columns:
                messagebox.showerror("Lỗi", "File không có cột 'STT TÚI'!")
                return
            
            # Lấy các cột được chọn
            cac_cot_chon = []
            for cot, config in self.cac_cot_mac_dinh.items():
                if config['var'].get() and cot in self.df_nguon.columns:
                    cac_cot_chon.append(cot)
            
            if not cac_cot_chon:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn ít nhất một cột!")
                return
            
            # Đảm bảo có cột STT TÚI (bắt buộc)
            if 'STT TÚI' not in cac_cot_chon:
                cac_cot_chon.append('STT TÚI')
            
            self.text_ket_qua_chia.delete(1.0, tk.END)
            self.text_ket_qua_chia.insert(tk.END, "⏳ Đang chia danh sách...\n")
            self.text_ket_qua_chia.update()
            
            # Tạo DataFrame đã chia
            self.df_da_chia = self.df_nguon[cac_cot_chon].copy()
            
            # Làm sạch dữ liệu: Bỏ từ "HUY CHƯƠNG" khỏi các cột kết quả
            ket_qua_cols = ['KQ VQG TOÁN', 'KQ VQG KHOA HỌC', 'KQ VQG TIẾNG ANH']
            for col in ket_qua_cols:
                if col in self.df_da_chia.columns:
                    self.df_da_chia[col] = self.df_da_chia[col].apply(
                        lambda x: str(x).replace('HUY CHƯƠNG ', '').replace('HUY CHUONG ', '').strip() if pd.notna(x) else x
                    )
            
            # Giữ nguyên thứ tự từ file nguồn (không sắp xếp lại)
            # Reset index để đảm bảo index liên tục
            self.df_da_chia = self.df_da_chia.reset_index(drop=True)
            
            # Hiển thị kết quả
            self.text_ket_qua_chia.delete(1.0, tk.END)
            self.text_ket_qua_chia.insert(tk.END, "=== KẾT QUẢ CHIA DANH SÁCH ===\n\n")
            
            # Thống kê theo túi
            if 'STT TÚI' in self.df_da_chia.columns:
                thong_ke = self.df_da_chia['STT TÚI'].value_counts().sort_index()
                tui_co_du_lieu = thong_ke[thong_ke.index > 0]  # Bỏ qua túi 0
                
                self.text_ket_qua_chia.insert(tk.END, f"✅ Tổng số thí sinh: {len(self.df_da_chia):,}\n")
                self.text_ket_qua_chia.insert(tk.END, f"📦 Số túi có dữ liệu: {len(tui_co_du_lieu)}\n")
                self.text_ket_qua_chia.insert(tk.END, f"📋 Số cột được chọn: {len(cac_cot_chon)}\n\n")
                
                self.text_ket_qua_chia.insert(tk.END, "=== THỐNG KÊ THEO TÚI ===\n")
                
                for stt_tui, so_luong in tui_co_du_lieu.items():
                    if pd.notna(stt_tui) and stt_tui > 0:
                        self.text_ket_qua_chia.insert(tk.END, f"📦 Túi {int(stt_tui):3d}: {so_luong:4d} thí sinh\n")
                
                # Thống kê SL GCN nếu có
                if 'SL GCN' in self.df_da_chia.columns:
                    tong_gcn = int(self.df_da_chia['SL GCN'].sum())
                    self.text_ket_qua_chia.insert(tk.END, f"\n📊 Tổng số GCN: {tong_gcn:,}\n")
            
            # Hiển thị 20 dòng đầu
            self.text_ket_qua_chia.insert(tk.END, "\n=== DỮ LIỆU MẪU (20 dòng đầu) ===\n")
            preview_df = self.df_da_chia.head(20).copy()
            self.text_ket_qua_chia.insert(tk.END, preview_df.to_string(index=False))
            
            if len(self.df_da_chia) > 20:
                self.text_ket_qua_chia.insert(tk.END, f"\n... và {len(self.df_da_chia) - 20:,} dòng khác")
            
            self.text_ket_qua_chia.see(tk.END)
            
            messagebox.showinfo("Thành công", f"Đã chia danh sách thành công!\n\nSố túi: {len(tui_co_du_lieu)}\nTổng số thí sinh: {len(self.df_da_chia):,}")
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra:\n{str(e)}")
            self.text_ket_qua_chia.insert(tk.END, f"\n❌ LỖI: {str(e)}\n")
            import traceback
            self.text_ket_qua_chia.insert(tk.END, traceback.format_exc())
    
    def xuat_ket_qua_chia(self):
        """Xuất kết quả chia danh sách ra Excel"""
        try:
            if self.df_da_chia is None:
                messagebox.showwarning("Cảnh báo", "Chưa có dữ liệu để xuất!\nVui lòng chia danh sách trước.")
                return
            
            # Đường dẫn mặc định với tên sheet
            sheet_name = self.current_sheet if self.current_sheet else "Unknown"
            # Làm sạch tên sheet để dùng làm tên file (bỏ ký tự đặc biệt)
            safe_sheet_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '_', '-')).strip()
            default_filename = f"Danh sách chia túi_{safe_sheet_name}.xlsx"
            
            file_path = filedialog.asksaveasfilename(
                title="Lưu danh sách đã chia",
                defaultextension=".xlsx",
                initialdir=self.output_dir.get(),
                initialfile=default_filename,
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if not file_path:
                return
            
            self.text_ket_qua_chia.insert(tk.END, f"\n⏳ Đang xuất file Excel...\n")
            self.text_ket_qua_chia.update()
            
            # Tạo file Excel với nhiều sheet
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Sheet tổng hợp
                df_tong_hop = self.df_da_chia.copy()
                # Đổi tên cột: Bỏ "KQ VQG " khỏi tiêu đề
                doi_ten_cot = {
                    'KQ VQG TOÁN': 'TOÁN',
                    'KQ VQG KHOA HỌC': 'KHOA HỌC',
                    'KQ VQG TIẾNG ANH': 'TIẾNG ANH'
                }
                df_tong_hop.rename(columns=doi_ten_cot, inplace=True)
                df_tong_hop.insert(0, 'STT', range(1, len(df_tong_hop) + 1))
                df_tong_hop.to_excel(writer, sheet_name='Tổng hợp', index=False)
                
                # Các sheet riêng theo từng túi
                if 'STT TÚI' in self.df_da_chia.columns:
                    df_tong_hop_tui = []
                    
                    for stt_tui in sorted(self.df_da_chia['STT TÚI'].dropna().unique()):
                        if stt_tui > 0:
                            df_tui = self.df_da_chia[self.df_da_chia['STT TÚI'] == stt_tui].copy()
                            # Đổi tên cột: Bỏ "KQ VQG " khỏi tiêu đề
                            df_tui.rename(columns=doi_ten_cot, inplace=True)
                            # Thêm cột STT vào đầu
                            df_tui.insert(0, 'STT', range(1, len(df_tui) + 1))
                            sheet_name = f'Túi {int(stt_tui)}'
                            df_tui.to_excel(writer, sheet_name=sheet_name, index=False)
                            df_tong_hop_tui.append({
                                'STT TÚI': int(stt_tui),
                                'Số thí sinh': len(df_tui),
                                'SL GCN': int(df_tui['SL GCN'].sum()) if 'SL GCN' in df_tui.columns else 0
                            })
                    
                    # Sheet thống kê túi
                    if df_tong_hop_tui:
                        df_thong_ke_tui = pd.DataFrame(df_tong_hop_tui)
                        df_thong_ke_tui.loc[len(df_thong_ke_tui)] = {
                            'STT TÚI': 'TỔNG CỘNG',
                            'Số thí sinh': df_thong_ke_tui['Số thí sinh'].sum(),
                            'SL GCN': df_thong_ke_tui['SL GCN'].sum()
                        }
                        # Thêm cột STT vào đầu (trừ dòng tổng cộng)
                        df_thong_ke_tui.insert(0, 'STT', [''] * len(df_thong_ke_tui))
                        # Đánh số từ 1 đến n-1 (n-1 vì dòng cuối là tổng cộng)
                        for i in range(len(df_thong_ke_tui) - 1):
                            df_thong_ke_tui.at[i, 'STT'] = i + 1
                        df_thong_ke_tui.to_excel(writer, sheet_name='Thống kê túi', index=False)
                
                # Áp dụng format cho tất cả các sheet
                for sheet_name in writer.sheets:
                    self.format_sheet_for_print(writer.sheets[sheet_name])
            
            messagebox.showinfo("Thành công", f"Đã xuất danh sách thành công!\n\nFile: {os.path.basename(file_path)}")
            self.text_ket_qua_chia.insert(tk.END, f"✅ Đã xuất file: {os.path.basename(file_path)}\n")
            self.text_ket_qua_chia.insert(tk.END, f"📁 Thư mục: {os.path.dirname(file_path)}\n")
            self.text_ket_qua_chia.see(tk.END)
                
        except PermissionError as e:
            messagebox.showerror("Lỗi quyền truy cập", 
                "Không thể ghi file!\n\n"
                "⚠️ Có thể file đang được mở trong Excel hoặc chương trình khác.\n\n"
                "Giải pháp:\n"
                "1. Đóng file Excel nếu đang mở\n"
                "2. Hoặc đổi tên file khác khi lưu")
            self.text_ket_qua_chia.insert(tk.END, f"\n❌ LỖI: File đang được mở bởi chương trình khác!\n")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể xuất file:\n{str(e)}")
            self.text_ket_qua_chia.insert(tk.END, f"\n❌ LỖI: {str(e)}\n")
            import traceback
            self.text_ket_qua_chia.insert(tk.END, traceback.format_exc())
    
    def format_sheet_for_print(self, ws):
        """Format sheet với border và page setup cho A4"""
        try:
            # Định nghĩa border
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Định nghĩa alignment
            center_alignment = Alignment(horizontal='center', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center')
            
            # Định nghĩa font cho header
            header_font = Font(bold=True, size=11)
            
            # Lấy kích thước dữ liệu
            max_row = ws.max_row
            max_col = ws.max_column
            
            # Format header (dòng 1)
            for col in range(1, max_col + 1):
                cell = ws.cell(row=1, column=col)
                cell.border = thin_border
                cell.alignment = center_alignment
                cell.font = header_font
            
            # Format các dòng dữ liệu
            for row in range(2, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    # STT, SBD căn giữa, các cột khác căn trái
                    if col == 1:  # Cột STT
                        cell.alignment = center_alignment
                    elif ws.cell(row=1, column=col).value in ['SBD', 'KHỐI', 'SL GCN', 'STT TÚI']:
                        cell.alignment = center_alignment
                    else:
                        cell.alignment = left_alignment
            
            # Auto-fit column width
            for col in range(1, max_col + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                for row in range(1, min(max_row + 1, 100)):  # Kiểm tra 100 dòng đầu
                    cell = ws.cell(row=row, column=col)
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                
                # Thiết lập width (tối thiểu 8, tối đa 50)
                adjusted_width = min(max(max_length + 2, 8), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Thiết lập page setup cho A4
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE  # Ngang
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0  # Không giới hạn chiều cao
            
            # Thiết lập margins (đơn vị inch)
            ws.page_margins.left = 0.5
            ws.page_margins.right = 0.5
            ws.page_margins.top = 0.75
            ws.page_margins.bottom = 0.75
            ws.page_margins.header = 0.3
            ws.page_margins.footer = 0.3
            
            # Thiết lập print options
            ws.print_options.horizontalCentered = True
            ws.print_options.verticalCentered = False
            
            # In lặp lại header trên mỗi trang
            ws.print_title_rows = '1:1'
            
        except Exception as e:
            print(f"Lỗi format sheet: {str(e)}")
        
    def browse_file(self, var):
        """Chọn file"""
        filename = filedialog.askopenfilename(
            title="Chọn file Excel",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            var.set(filename)
            self.log(f"✅ Đã chọn file: {os.path.basename(filename)}")
    
    def browse_directory(self):
        """Chọn thư mục"""
        directory = filedialog.askdirectory(title="Chọn thư mục lưu kết quả")
        if directory:
            self.output_dir.set(directory)
            self.log(f"✅ Đã chọn thư mục: {directory}")
    
    def log(self, message):
        """Ghi log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
    
    def clear_log(self):
        """Xóa log"""
        self.log_text.delete(1.0, tk.END)
    
    def open_output_folder(self):
        """Mở thư mục output"""
        output_dir = self.output_dir.get()
        if os.path.exists(output_dir):
            os.startfile(output_dir)
        else:
            messagebox.showerror("Lỗi", "Thư mục không tồn tại!")
    
    def update_progress(self, value):
        """Cập nhật progress bar"""
        self.progress['value'] = value
        self.root.update_idletasks()
    
    def start_processing(self):
        """Bắt đầu xử lý"""
        if self.is_processing:
            messagebox.showwarning("Cảnh báo", "Đang xử lý, vui lòng đợi!")
            return
        
        # Kiểm tra file input
        if not self.file_full.get() or not self.file_trao_giai.get():
            messagebox.showerror("Lỗi", "Vui lòng chọn đầy đủ 2 file đầu vào!")
            return
        
        if not os.path.exists(self.file_full.get()):
            messagebox.showerror("Lỗi", "File Awards_Template_Full.xlsx không tồn tại!")
            return
        
        if not os.path.exists(self.file_trao_giai.get()):
            messagebox.showerror("Lỗi", "File Awards_TRAO GIAI.xlsx không tồn tại!")
            return
        
        # Chạy xử lý trong thread riêng
        self.is_processing = True
        self.process_btn.config(state=tk.DISABLED, text="⏳ ĐANG XỬ LÝ...")
        
        thread = threading.Thread(target=self.process_all)
        thread.daemon = True
        thread.start()
    
    def process_all(self):
        """Xử lý toàn bộ quy trình"""
        try:
            self.clear_log()
            self.log("="*80)
            self.log("🎓 BẮT ĐẦU QUY TRÌNH XỬ LÝ MÃ CERT")
            self.log("="*80)
            
            output_dir = self.output_dir.get()
            
            # Định nghĩa đường dẫn file output
            file_step1 = os.path.join(output_dir, "Awards_Comparison_Result.xlsx")
            file_step2 = os.path.join(output_dir, "Awards_Comparison_WITH_RANK.xlsx")
            file_step3 = os.path.join(output_dir, "Awards_Comparison_WITH_CERT.xlsx")
            file_step4 = os.path.join(output_dir, "Awards_Comparison_WITH_REPORT.xlsx")
            
            # BƯỚC 1: So sánh và tách file
            self.update_progress(10)
            self.log("\n📌 BƯỚC 1/4: Tạo file so sánh...")
            self.step1_compare_files(file_step1)
            self.update_progress(25)
            
            # BƯỚC 2: Thêm rank và sắp xếp
            self.update_progress(30)
            self.log("\n📌 BƯỚC 2/4: Thêm RANK NHẬN GIẢI và sắp xếp...")
            self.step2_add_rank(file_step1, file_step2)
            self.update_progress(50)
            
            # BƯỚC 3: Tạo mã CERT
            self.update_progress(55)
            self.log("\n📌 BƯỚC 3/4: Tạo mã CERT...")
            self.step3_generate_cert(file_step2, file_step3)
            self.update_progress(75)
            
            # BƯỚC 4: Tạo báo cáo thống kê
            self.update_progress(80)
            self.log("\n📌 BƯỚC 4/4: Tạo báo cáo thống kê...")
            self.step4_create_report(file_step3, file_step4)
            self.update_progress(100)
            
            # Hoàn thành
            self.log("\n" + "="*80)
            self.log("🎉 HOÀN THÀNH TOÀN BỘ QUY TRÌNH!")
            self.log("="*80)
            self.log(f"\n✅ Các file đã tạo:")
            self.log(f"   1. {os.path.basename(file_step1)}")
            self.log(f"   2. {os.path.basename(file_step2)}")
            self.log(f"   3. {os.path.basename(file_step3)}")
            self.log(f"   4. {os.path.basename(file_step4)}")
            
            messagebox.showinfo("Thành công", "Đã hoàn thành toàn bộ quy trình!\n\nCác file đã được lưu vào thư mục output.")
            
        except PermissionError as e:
            self.log(f"\n❌ LỖI: File đang được mở bởi chương trình khác!")
            messagebox.showerror("Lỗi quyền truy cập", 
                "Không thể ghi file!\n\n"
                "⚠️ Có thể file đang được mở trong Excel hoặc chương trình khác.\n\n"
                "Giải pháp:\n"
                "1. Đóng tất cả file Excel đang mở\n"
                "2. Chạy lại chương trình")
        except Exception as e:
            self.log(f"\n❌ LỖI: {str(e)}")
            import traceback
            self.log(traceback.format_exc())
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra:\n{str(e)}")
        
        finally:
            self.is_processing = False
            self.process_btn.config(state=tk.NORMAL, text="▶ BẮT ĐẦU XỬ LÝ")
            self.update_progress(0)
    
    # ========== CÁC HÀM XỬ LÝ ==========
    
    def step1_compare_files(self, output_file):
        """Bước 1: So sánh và tách file"""
        df_trao_giai = pd.read_excel(self.file_trao_giai.get())
        sbd_trao_giai = set(df_trao_giai['SBD'].dropna().astype(str))
        self.log(f"   ✓ Đọc file TRAO GIẢI: {len(df_trao_giai)} dòng")
        
        df_full = pd.read_excel(self.file_full.get())
        self.log(f"   ✓ Đọc file FULL: {len(df_full)} dòng")
        
        df_full['SBD_str'] = df_full['SBD'].astype(str)
        df_sheet1 = df_full[df_full['SBD_str'].isin(sbd_trao_giai)].drop('SBD_str', axis=1)
        df_sheet2 = df_full[~df_full['SBD_str'].isin(sbd_trao_giai)].drop('SBD_str', axis=1)
        
        self.log(f"   ✓ Sheet TRAO GIẢI: {len(df_sheet1)} học sinh")
        self.log(f"   ✓ Sheet KO ĐK: {len(df_sheet2)} học sinh")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_sheet1.to_excel(writer, sheet_name='TRAO GIẢI', index=False)
            df_sheet2.to_excel(writer, sheet_name='KO ĐK', index=False)
        
        self.log(f"   ✅ Đã lưu: {os.path.basename(output_file)}")
    
    def step2_add_rank(self, input_file, output_file):
        """Bước 2: Thêm rank và sắp xếp"""
        df_trao_giai = pd.read_excel(input_file, sheet_name='TRAO GIẢI')
        df_ko_dk = pd.read_excel(input_file, sheet_name='KO ĐK')
        
        # Tính RANK NHẬN GIẢI
        rank_cols = ['RANK T', 'RANK S', 'RANK E']
        df_trao_giai['RANK NHẬN GIẢI'] = df_trao_giai[rank_cols].min(axis=1, skipna=True)
        
        all_nan_mask = df_trao_giai[rank_cols].isna().all(axis=1)
        df_trao_giai.loc[all_nan_mask, 'RANK NHẬN GIẢI'] = np.nan
        
        # Sắp xếp
        df_trao_giai = df_trao_giai.sort_values(
            ['RANK NHẬN GIẢI', 'KHỐI', 'TRƯỜNG'],
            na_position='last'
        ).reset_index(drop=True)
        
        self.log(f"   ✓ Đã thêm RANK NHẬN GIẢI và sắp xếp")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_trao_giai.to_excel(writer, sheet_name='TRAO GIẢI', index=False)
            df_ko_dk.to_excel(writer, sheet_name='KO ĐK', index=False)
        
        self.log(f"   ✅ Đã lưu: {os.path.basename(output_file)}")
    
    def step3_generate_cert(self, input_file, output_file):
        """Bước 3: Tạo mã CERT"""
        df_trao_giai = pd.read_excel(input_file, sheet_name='TRAO GIẢI')
        df_ko_dk = pd.read_excel(input_file, sheet_name='KO ĐK')
        
        # Sắp xếp sheet KO ĐK theo MÃ TRƯỜNG (cột AC), sau đó đến KHỐI (cột E)
        if 'MÃ TRƯỜNG' in df_ko_dk.columns and 'KHỐI' in df_ko_dk.columns:
            df_ko_dk['MÃ TRƯỜNG'] = df_ko_dk['MÃ TRƯỜNG'].astype(str)
            df_ko_dk = df_ko_dk.sort_values(['MÃ TRƯỜNG', 'KHỐI'], na_position='last').reset_index(drop=True)
        
        # Xử lý sheet TRAO GIẢI
        df_trao_giai = self.process_sheet_cert(df_trao_giai, 1)
        bags1 = df_trao_giai['STT TÚI'].max() if 'STT TÚI' in df_trao_giai.columns else 0
        self.log(f"   ✓ Sheet TRAO GIẢI: {len(df_trao_giai)} HS, {int(bags1)} túi")
        
        # Xử lý sheet KO ĐK
        start_bag = int(bags1) + 1
        df_ko_dk = self.process_sheet_cert(df_ko_dk, start_bag)
        bags2 = df_ko_dk['STT TÚI'].max() - bags1 if 'STT TÚI' in df_ko_dk.columns else 0
        self.log(f"   ✓ Sheet KO ĐK: {len(df_ko_dk)} HS, {int(bags2)} túi")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_trao_giai.to_excel(writer, sheet_name='TRAO GIẢI', index=False)
            df_ko_dk.to_excel(writer, sheet_name='KO ĐK', index=False)
        
        self.log(f"   ✅ Đã lưu: {os.path.basename(output_file)}")
    
    def process_sheet_cert(self, df, start_bag_num):
        """Xử lý tạo mã CERT cho 1 sheet"""
        df['MÃ CERT ĐẦY ĐỦ'] = df.apply(self.generate_cert_code_full, axis=1)
        df['MÃ CERT'] = df.apply(self.generate_cert_code_short, axis=1)
        df['Rank nhận giải'] = df.apply(self.get_highest_rank, axis=1)
        df['SL GCN'] = df.apply(self.count_certificates, axis=1)
        
        # Phân túi
        bag_series = self.assign_bags(df, start_bag_num)
        df['STT TÚI'] = bag_series
        
        # Cập nhật mã CERT với STT túi
        df['MÃ CERT ĐẦY ĐỦ'] = df.apply(
            lambda row: f"{row['MÃ CERT ĐẦY ĐỦ']}*{int(row['STT TÚI'])}" if row['STT TÚI'] > 0 else row['MÃ CERT ĐẦY ĐỦ'],
            axis=1
        )
        df['MÃ CERT'] = df.apply(
            lambda row: f"{row['MÃ CERT']}*{int(row['STT TÚI'])}" if row['STT TÚI'] > 0 else row['MÃ CERT'],
            axis=1
        )
        
        return df
    
    def step4_create_report(self, input_file, output_file):
        """Bước 4: Tạo báo cáo thống kê"""
        df_trao_giai = pd.read_excel(input_file, sheet_name='TRAO GIẢI')
        df_ko_dk = pd.read_excel(input_file, sheet_name='KO ĐK')
        
        # Báo cáo TRAO GIẢI theo khối
        report1 = self.create_report_by_khoi(df_trao_giai)
        self.log(f"   ✓ Báo cáo TRAO GIẢI: {len(report1)-1} khối")
        
        # Báo cáo KO ĐK theo mã trường
        report2 = self.create_report_by_truong(df_ko_dk)
        self.log(f"   ✓ Báo cáo KO ĐK: {len(report2)-1} trường")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            report1.to_excel(writer, sheet_name='BÁO CÁO TRAO GIẢI', index=False)
            report2.to_excel(writer, sheet_name='BÁO CÁO KO ĐK', index=False)
        
        self.log(f"   ✅ Đã lưu: {os.path.basename(output_file)}")
    
    # ========== CÁC HÀM HỖ TRỢ ==========
    
    @staticmethod
    def map_result_to_code(result, subject):
        if pd.isna(result) or result == '':
            return f'NULL-{subject}'
        result = str(result).strip().upper()
        if 'VÀNG' in result or 'VANG' in result:
            return f'V-{subject}'
        elif 'BẠC' in result or 'BAC' in result:
            return f'B-{subject}'
        elif 'ĐỒNG' in result or 'DONG' in result:
            return f'D-{subject}'
        elif 'KHUYẾN KHÍCH' in result or 'KHUYEN KHICH' in result or 'KK' in result:
            return f'KK-{subject}'
        elif 'CHỨNG NHẬN' in result or 'CHUNG NHAN' in result or 'CN' in result:
            return f'CN-{subject}'
        return f'NULL-{subject}'
    
    def generate_cert_code_full(self, row):
        """Tạo mã CERT đầy đủ theo thứ tự: MATH → ENGLISH → SCIENCE"""
        khoi = str(int(row['KHỐI'])) if not pd.isna(row['KHỐI']) else 'X'
        math = self.map_result_to_code(row['KQ VQG TOÁN'], 'MATH')
        english = self.map_result_to_code(row['KQ VQG TIẾNG ANH'], 'ENGLISH')
        science = self.map_result_to_code(row['KQ VQG KHOA HỌC'], 'SCIENCE')
        return f"{khoi}*{math}*{english}*{science}"
    
    def generate_cert_code_short(self, row):
        """Tạo mã CERT rút gọn theo thứ tự: M → E → S"""
        khoi = str(int(row['KHỐI'])) if not pd.isna(row['KHỐI']) else 'X'
        math = self.map_result_to_code(row['KQ VQG TOÁN'], 'M')
        english = self.map_result_to_code(row['KQ VQG TIẾNG ANH'], 'E')
        science = self.map_result_to_code(row['KQ VQG KHOA HỌC'], 'S')
        
        parts = [khoi]
        # Thứ tự: MATH → ENGLISH → SCIENCE
        if not math.startswith('NULL'):
            parts.append(math)
        if not english.startswith('NULL'):
            parts.append(english)
        if not science.startswith('NULL'):
            parts.append(science)
        
        return '*'.join(parts)
    
    def get_highest_rank(self, row):
        math = self.map_result_to_code(row['KQ VQG TOÁN'], 'M')
        science = self.map_result_to_code(row['KQ VQG KHOA HỌC'], 'S')
        english = self.map_result_to_code(row['KQ VQG TIẾNG ANH'], 'E')
        
        rank_priority = {'V': 3, 'B': 2, 'D': 1}
        awards = []
        
        for code, priority in [(math, 3), (english, 2), (science, 1)]:
            if code.startswith(('V-', 'B-', 'D-')):
                rank_type = code.split('-')[0]
                awards.append((rank_priority.get(rank_type, 0), priority, code))
        
        if not awards:
            return ''
        
        awards.sort(key=lambda x: (x[0], x[1]), reverse=True)
        return awards[0][2]
    
    @staticmethod
    def count_certificates(row):
        count = 0
        for col in ['KQ VQG TOÁN', 'KQ VQG KHOA HỌC', 'KQ VQG TIẾNG ANH']:
            if not pd.isna(row[col]) and row[col] != '':
                count += 1
        return count
    
    @staticmethod
    def assign_bags(df, start_bag_number=1, max_gcn=30):
        current_bag = start_bag_number
        current_gcn = 0
        bag_assignments = {}
        
        for idx, row in df.iterrows():
            student_gcn = row['SL GCN']
            
            if student_gcn == 0:
                bag_assignments[idx] = 0
                continue
            
            if current_gcn + student_gcn <= max_gcn:
                current_gcn += student_gcn
                bag_assignments[idx] = current_bag
            else:
                current_bag += 1
                current_gcn = student_gcn
                bag_assignments[idx] = current_bag
        
        return pd.Series(bag_assignments, name='STT TÚI')
    
    def create_report_by_khoi(self, df):
        """Tạo báo cáo theo khối"""
        khoi_list = sorted(df['KHỐI'].dropna().unique())
        report = []
        
        for khoi in khoi_list:
            df_khoi = df[df['KHỐI'] == khoi]
            
            # Đếm số học sinh theo giải cao nhất
            hs_vang = self.count_students_by_highest_award(df_khoi, 'VÀNG')
            hs_bac = self.count_students_by_highest_award(df_khoi, 'BẠC')
            hs_dong = self.count_students_by_highest_award(df_khoi, 'ĐỒNG')
            
            # Đếm số GCN
            vang_gcn = self.count_gcn_for_award(df_khoi, 'VÀNG|VANG')
            bac_gcn = self.count_gcn_for_award(df_khoi, 'BẠC|BAC')
            dong_gcn = self.count_gcn_for_award(df_khoi, 'ĐỒNG|DONG')
            kk_gcn = self.count_gcn_for_award(df_khoi, 'KHUYẾN KHÍCH|KHUYEN KHICH')
            cn_gcn = self.count_gcn_for_award(df_khoi, 'CHỨNG NHẬN|CHUNG NHAN')
            
            report.append({
                'Khối': int(khoi),
                'Tổng HS': len(df_khoi),
                'Số HS VÀNG': hs_vang,
                'Số HS BẠC': hs_bac,
                'Số HS ĐỒNG': hs_dong,
                'GCN VÀNG': vang_gcn,
                'GCN BẠC': bac_gcn,
                'GCN ĐỒNG': dong_gcn,
                'GCN KHUYẾN KHÍCH': kk_gcn,
                'GCN CHỨNG NHẬN': cn_gcn,
                'TỔNG GCN': int(df_khoi['SL GCN'].sum())
            })
        
        # Tổng cộng
        report.append({
            'Khối': 'TỔNG CỘNG',
            'Tổng HS': sum([r['Tổng HS'] for r in report]),
            'Số HS VÀNG': sum([r['Số HS VÀNG'] for r in report]),
            'Số HS BẠC': sum([r['Số HS BẠC'] for r in report]),
            'Số HS ĐỒNG': sum([r['Số HS ĐỒNG'] for r in report]),
            'GCN VÀNG': sum([r['GCN VÀNG'] for r in report]),
            'GCN BẠC': sum([r['GCN BẠC'] for r in report]),
            'GCN ĐỒNG': sum([r['GCN ĐỒNG'] for r in report]),
            'GCN KHUYẾN KHÍCH': sum([r['GCN KHUYẾN KHÍCH'] for r in report]),
            'GCN CHỨNG NHẬN': sum([r['GCN CHỨNG NHẬN'] for r in report]),
            'TỔNG GCN': sum([r['TỔNG GCN'] for r in report])
        })
        
        return pd.DataFrame(report)
    
    def create_report_by_truong(self, df):
        """Tạo báo cáo theo trường"""
        # Sắp xếp theo Mã trường trước, sau đó đến Khối
        df_sorted = df.sort_values(['MÃ TRƯỜNG', 'KHỐI'], na_position='last')
        
        # Lấy danh sách kết hợp (Mã trường, Khối)
        group_keys = df_sorted.groupby(['MÃ TRƯỜNG', 'KHỐI'], dropna=False).size().index.tolist()
        
        report = []
        
        for ma_truong, khoi in group_keys:
            df_truong = df[(df['MÃ TRƯỜNG'] == ma_truong) & (df['KHỐI'] == khoi)]
            
            vang = self.count_gcn_for_award(df_truong, 'VÀNG|VANG')
            bac = self.count_gcn_for_award(df_truong, 'BẠC|BAC')
            dong = self.count_gcn_for_award(df_truong, 'ĐỒNG|DONG')
            
            ten_truong = df_truong['TRƯỜNG'].iloc[0] if 'TRƯỜNG' in df_truong.columns and len(df_truong) > 0 else ''
            khoi_display = int(khoi) if not pd.isna(khoi) else ''
            
            report.append({
                'MÃ TRƯỜNG': str(ma_truong),
                'TÊN TRƯỜNG': ten_truong,
                'Khối': khoi_display,
                'Tổng HS': len(df_truong),
                'GCN VÀNG': vang,
                'GCN BẠC': bac,
                'GCN ĐỒNG': dong,
                'TỔNG GCN': int(df_truong['SL GCN'].sum())
            })
        
        # Tổng cộng
        report.append({
            'MÃ TRƯỜNG': 'TỔNG CỘNG',
            'TÊN TRƯỜNG': '',
            'Khối': '',
            'Tổng HS': sum([r['Tổng HS'] for r in report]),
            'GCN VÀNG': sum([r['GCN VÀNG'] for r in report]),
            'GCN BẠC': sum([r['GCN BẠC'] for r in report]),
            'GCN ĐỒNG': sum([r['GCN ĐỒNG'] for r in report]),
            'TỔNG GCN': sum([r['TỔNG GCN'] for r in report])
        })
        
        return pd.DataFrame(report)
    
    @staticmethod
    def count_gcn_for_award(df, award_type):
        """Đếm số GCN cho loại giải"""
        count = 0
        for col in ['KQ VQG TOÁN', 'KQ VQG KHOA HỌC', 'KQ VQG TIẾNG ANH']:
            if col in df.columns:
                count += df[col].astype(str).str.upper().str.contains(award_type, na=False).sum()
        return count
    
    @staticmethod
    def count_students_by_highest_award(df, award_level):
        """
        Đếm số học sinh theo giải cao nhất
        - VÀNG: có ít nhất 1 giải VÀNG
        - BẠC: giải cao nhất là BẠC (không có VÀNG)
        - ĐỒNG: chỉ có huy chương ĐỒNG (không có VÀNG hoặc BẠC)
        """
        count = 0
        for idx, row in df.iterrows():
            has_vang = False
            has_bac = False
            has_dong = False
            
            for col in ['KQ VQG TOÁN', 'KQ VQG KHOA HỌC', 'KQ VQG TIẾNG ANH']:
                if col in df.columns:
                    val = str(row[col]).upper()
                    if 'VÀNG' in val or 'VANG' in val:
                        has_vang = True
                    elif 'BẠC' in val or 'BAC' in val:
                        has_bac = True
                    elif 'ĐỒNG' in val or 'DONG' in val:
                        has_dong = True
            
            if award_level == 'VÀNG' and has_vang:
                count += 1
            elif award_level == 'BẠC' and not has_vang and has_bac:
                count += 1
            elif award_level == 'ĐỒNG' and not has_vang and not has_bac and has_dong:
                count += 1
        
        return count
    
    # ========== TAB 3: TRA CỨU ==========
    
    def create_tab3_content(self, parent):
        """Tạo Tab 3: Tra cứu thông tin học sinh"""
        
        # Container chính với scroll
        main_container = tk.Frame(parent, bg="#ecf0f1")
        main_container.pack(fill=tk.BOTH, expand=True)
        
        # Top Frame: Upload file và tìm kiếm
        top_frame = tk.Frame(main_container, bg="#ecf0f1", padx=20, pady=10)
        top_frame.pack(fill=tk.X)
        
        # === 1. UPLOAD FILE ===
        upload_frame = tk.LabelFrame(
            top_frame,
            text="📂 CHỌN FILE DỮ LIỆU",
            font=("Arial", 11, "bold"),
            bg="#ecf0f1",
            fg="#2c3e50",
            padx=10,
            pady=10
        )
        upload_frame.pack(fill=tk.X, pady=(0, 10))
        
        file_frame = tk.Frame(upload_frame, bg="#ecf0f1")
        file_frame.pack(fill=tk.X)
        
        tk.Label(file_frame, text="File:", font=("Arial", 10), bg="#ecf0f1").pack(side=tk.LEFT, padx=5)
        tk.Entry(file_frame, textvariable=self.file_tracuu_var, width=50, font=("Arial", 9)).pack(side=tk.LEFT, padx=5)
        tk.Button(
            file_frame,
            text="📂 Chọn file",
            command=self.browse_file_tracuu,
            bg="#3498db",
            fg="white",
            font=("Arial", 9, "bold"),
            cursor="hand2",
            relief=tk.RAISED,
            bd=2
        ).pack(side=tk.LEFT, padx=5)
        tk.Button(
            file_frame,
            text="📖 Đọc file",
            command=self.load_file_tracuu,
            bg="#27ae60",
            fg="white",
            font=("Arial", 9, "bold"),
            cursor="hand2",
            relief=tk.RAISED,
            bd=2
        ).pack(side=tk.LEFT, padx=5)
        
        # Sheet selection frame
        self.sheet_selection_frame = tk.Frame(upload_frame, bg="#ecf0f1")
        self.sheet_selection_frame.pack(fill=tk.X, pady=(10, 0))
        
        tk.Label(
            self.sheet_selection_frame,
            text="📑 Chọn sheet:",
            font=("Arial", 10, "bold"),
            bg="#ecf0f1",
            fg="#2c3e50"
        ).pack(side=tk.LEFT, padx=5)
        
        # Container for checkboxes (will be populated after reading file)
        self.sheet_checkbox_container = tk.Frame(self.sheet_selection_frame, bg="#ecf0f1")
        self.sheet_checkbox_container.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Load selected sheets button
        self.load_sheets_btn = tk.Button(
            self.sheet_selection_frame,
            text="✅ Load dữ liệu từ sheet đã chọn",
            command=self.load_selected_sheets,
            bg="#e67e22",
            fg="white",
            font=("Arial", 9, "bold"),
            cursor="hand2",
            relief=tk.RAISED,
            bd=2,
            state=tk.DISABLED
        )
        self.load_sheets_btn.pack(side=tk.RIGHT, padx=5)
        
        # === 2. SEARCH FORM ===
        search_frame = tk.LabelFrame(
            top_frame,
            text="🔍 TÌM KIẾM",
            font=("Arial", 11, "bold"),
            bg="#ecf0f1",
            fg="#2c3e50",
            padx=10,
            pady=10
        )
        search_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Row 1: SBD và Họ tên
        row1 = tk.Frame(search_frame, bg="#ecf0f1")
        row1.pack(fill=tk.X, pady=5)
        
        tk.Label(row1, text="SBD:", font=("Arial", 10), bg="#ecf0f1", width=10, anchor='w').pack(side=tk.LEFT)
        self.sbd_entry = tk.Entry(row1, font=("Arial", 10), width=20)
        self.sbd_entry.pack(side=tk.LEFT, padx=5)
        
        tk.Label(row1, text="Họ tên:", font=("Arial", 10), bg="#ecf0f1", width=10, anchor='w').pack(side=tk.LEFT, padx=(20, 0))
        self.hoten_entry = tk.Entry(row1, font=("Arial", 10), width=30)
        self.hoten_entry.pack(side=tk.LEFT, padx=5)
        
        # Row 2: Ngày sinh
        row2 = tk.Frame(search_frame, bg="#ecf0f1")
        row2.pack(fill=tk.X, pady=5)
        
        tk.Label(row2, text="Ngày sinh:", font=("Arial", 10), bg="#ecf0f1", width=10, anchor='w').pack(side=tk.LEFT)
        
        self.day_var = tk.StringVar()
        self.month_var = tk.StringVar()
        self.year_var = tk.StringVar()
        
        tk.Label(row2, text="Ngày:", font=("Arial", 9), bg="#ecf0f1").pack(side=tk.LEFT, padx=(5, 2))
        day_combo = ttk.Combobox(row2, textvariable=self.day_var, width=5, state='readonly')
        day_combo['values'] = [''] + list(range(1, 32))
        day_combo.pack(side=tk.LEFT, padx=2)
        
        tk.Label(row2, text="Tháng:", font=("Arial", 9), bg="#ecf0f1").pack(side=tk.LEFT, padx=(10, 2))
        month_combo = ttk.Combobox(row2, textvariable=self.month_var, width=5, state='readonly')
        month_combo['values'] = [''] + list(range(1, 13))
        month_combo.pack(side=tk.LEFT, padx=2)
        
        tk.Label(row2, text="Năm:", font=("Arial", 9), bg="#ecf0f1").pack(side=tk.LEFT, padx=(10, 2))
        year_combo = ttk.Combobox(row2, textvariable=self.year_var, width=8, state='readonly')
        year_combo['values'] = [''] + list(range(2020, 1989, -1))
        year_combo.pack(side=tk.LEFT, padx=2)
        
        # Buttons
        btn_frame = tk.Frame(search_frame, bg="#ecf0f1")
        btn_frame.pack(fill=tk.X, pady=(10, 0))
        
        tk.Button(
            btn_frame,
            text="🔍 TÌM KIẾM",
            command=self.search_students,
            bg="#e74c3c",
            fg="white",
            font=("Arial", 10, "bold"),
            cursor="hand2",
            width=15
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            btn_frame,
            text="🔄 XÓA BỘ LỌC",
            command=self.clear_search_form,
            bg="#95a5a6",
            fg="white",
            font=("Arial", 10, "bold"),
            cursor="hand2",
            width=15
        ).pack(side=tk.LEFT, padx=5)
        
        self.result_count_label = tk.Label(
            btn_frame,
            text="Kết quả: 0",
            font=("Arial", 10, "bold"),
            bg="#ecf0f1",
            fg="#2c3e50"
        )
        self.result_count_label.pack(side=tk.RIGHT, padx=10)
        
        # === 3. RESULTS TABLE & DETAILS ===
        content_frame = tk.Frame(main_container, bg="#ecf0f1", padx=20)
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        # Left: Results Table
        left_frame = tk.Frame(content_frame, bg="white", relief=tk.SUNKEN, bd=2)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        tk.Label(
            left_frame,
            text="📋 KẾT QUẢ TÌM KIẾM",
            font=("Arial", 11, "bold"),
            bg="white",
            fg="#2c3e50"
        ).pack(pady=5)
        
        # Treeview for results
        tree_frame = tk.Frame(left_frame, bg="white")
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        tree_scroll_y = tk.Scrollbar(tree_frame)
        tree_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        
        tree_scroll_x = tk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)
        tree_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.results_tree = ttk.Treeview(
            tree_frame,
            columns=("SBD", "Họ tên", "Ngày sinh", "Khối", "Trường", "Cert", "Toán", "TA", "KH", "Sheet"),
            show="headings",
            yscrollcommand=tree_scroll_y.set,
            xscrollcommand=tree_scroll_x.set,
            height=15
        )
        
        tree_scroll_y.config(command=self.results_tree.yview)
        tree_scroll_x.config(command=self.results_tree.xview)
        
        # Define columns
        columns_config = [
            ("SBD", 100),
            ("Họ tên", 150),
            ("Ngày sinh", 100),
            ("Khối", 50),
            ("Trường", 200),
            ("Cert", 150),
            ("Toán", 120),
            ("TA", 120),
            ("KH", 120),
            ("Sheet", 120)
        ]
        
        for col, width in columns_config:
            self.results_tree.heading(col, text=col)
            self.results_tree.column(col, width=width, anchor='center')
        
        self.results_tree.pack(fill=tk.BOTH, expand=True)
        self.results_tree.bind('<<TreeviewSelect>>', self.on_student_select)
        
        # Right: Student Details
        right_frame = tk.Frame(content_frame, bg="white", relief=tk.SUNKEN, bd=2, width=350)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, padx=(10, 0))
        right_frame.pack_propagate(False)
        
        tk.Label(
            right_frame,
            text="👤 THÔNG TIN CHI TIẾT",
            font=("Arial", 11, "bold"),
            bg="white",
            fg="#2c3e50"
        ).pack(pady=10)
        
        # Details content
        details_scroll = tk.Scrollbar(right_frame)
        details_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.details_canvas = tk.Canvas(right_frame, bg="white", yscrollcommand=details_scroll.set, highlightthickness=0)
        self.details_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        details_scroll.config(command=self.details_canvas.yview)
        
        self.details_frame = tk.Frame(self.details_canvas, bg="white")
        self.details_canvas.create_window((0, 0), window=self.details_frame, anchor='nw')
        
        # Placeholder
        self.details_placeholder = tk.Label(
            self.details_frame,
            text="Chọn học sinh để xem chi tiết",
            font=("Arial", 10, "italic"),
            bg="white",
            fg="#7f8c8d"
        )
        self.details_placeholder.pack(pady=50)
        
        self.details_frame.bind('<Configure>', lambda e: self.details_canvas.configure(scrollregion=self.details_canvas.bbox("all")))
    
    def browse_file_tracuu(self):
        """Chọn file tra cứu"""
        file_path = filedialog.askopenfilename(
            title="Chọn file dữ liệu tra cứu",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.file_tracuu_var.set(file_path)
    
    def load_file_tracuu(self):
        """Đọc danh sách sheet từ file Excel"""
        file_path = self.file_tracuu_var.get()
        if not file_path or not os.path.exists(file_path):
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn file hợp lệ!")
            return
        
        try:
            # Đọc danh sách sheet
            xl_file = pd.ExcelFile(file_path)
            self.tracuu_sheets = xl_file.sheet_names
            
            # Clear previous checkboxes
            for widget in self.sheet_checkbox_container.winfo_children():
                widget.destroy()
            self.sheet_checkboxes.clear()
            
            # Create checkboxes for each sheet
            for i, sheet_name in enumerate(self.tracuu_sheets):
                var = tk.BooleanVar(value=False)
                # Auto-select 'TRAO GIẢI' or first sheet
                if sheet_name == 'TRAO GIẢI' or (i == 0 and 'TRAO GIẢI' not in self.tracuu_sheets):
                    var.set(True)
                
                cb = tk.Checkbutton(
                    self.sheet_checkbox_container,
                    text=sheet_name,
                    variable=var,
                    bg="#ecf0f1",
                    font=("Arial", 9),
                    activebackground="#ecf0f1"
                )
                cb.pack(side=tk.LEFT, padx=5)
                self.sheet_checkboxes[sheet_name] = var
            
            # Enable load button
            self.load_sheets_btn.config(state=tk.NORMAL)
            
            messagebox.showinfo("Thành công", f"Đã tìm thấy {len(self.tracuu_sheets)} sheet!\n\nVui lòng chọn sheet và click 'Load dữ liệu'.")
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể đọc file:\n{str(e)}")
    
    def load_selected_sheets(self):
        """Load dữ liệu từ các sheet đã chọn"""
        file_path = self.file_tracuu_var.get()
        if not file_path or not os.path.exists(file_path):
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn file hợp lệ!")
            return
        
        # Get selected sheets
        selected_sheets = [name for name, var in self.sheet_checkboxes.items() if var.get()]
        
        if not selected_sheets:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn ít nhất 1 sheet!")
            return
        
        try:
            # Read data from selected sheets
            all_data = []
            total_students = 0
            
            for sheet_name in selected_sheets:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                # Add sheet name column
                df['_SHEET_NAME'] = sheet_name
                all_data.append(df)
                total_students += len(df)
            
            # Merge all data
            self.df_tracuu = pd.concat(all_data, ignore_index=True)
            
            # Làm sạch dữ liệu: Bỏ từ "HUY CHƯƠNG" khỏi các cột kết quả
            ket_qua_cols = ['TOÁN', 'KQ VQG TOÁN', 'KHOA HỌC', 'KQ VQG KHOA HỌC', 'TIẾNG ANH', 'KQ VQG TIẾNG ANH']
            for col in ket_qua_cols:
                if col in self.df_tracuu.columns:
                    self.df_tracuu[col] = self.df_tracuu[col].apply(
                        lambda x: str(x).replace('HUY CHƯƠNG ', '').replace('HUY CHUONG ', '').replace('HUY CHƯƠNG', '').replace('HUY CHUONG', '') 
                        if pd.notna(x) and str(x).strip() != '' else x
                    )
            
            messagebox.showinfo(
                "Thành công", 
                f"Đã load {total_students} học sinh từ {len(selected_sheets)} sheet!\n\n" +
                "\n".join([f"• {name}" for name in selected_sheets])
            )
            
            # Hiển thị tất cả học sinh ban đầu
            self.current_results = self.df_tracuu.to_dict('records')
            self.display_search_results()
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể đọc dữ liệu:\n{str(e)}")
    
    def search_students(self):
        """Tìm kiếm học sinh"""
        if self.df_tracuu is None:
            messagebox.showwarning("Cảnh báo", "Vui lòng đọc file dữ liệu trước!")
            return
        
        # Lấy điều kiện tìm kiếm
        sbd = self.sbd_entry.get().strip().lower()
        hoten = self.hoten_entry.get().strip().lower()
        day = self.day_var.get()
        month = self.month_var.get()
        year = self.year_var.get()
        
        # Filter
        results = self.df_tracuu.copy()
        
        if sbd:
            results = results[results['SBD'].astype(str).str.lower().str.contains(sbd, na=False)]
        
        if hoten:
            results = results[results['FULL NAME'].astype(str).str.lower().str.contains(hoten, na=False)]
        
        # Filter theo ngày sinh
        if day or month or year:
            def match_dob(dob_str):
                if pd.isna(dob_str):
                    return False
                dob = str(dob_str)
                
                if day:
                    day_padded = str(day).zfill(2)
                    if not dob.startswith(day_padded):
                        return False
                
                if month:
                    month_padded = str(month).zfill(2)
                    if f'-{month_padded}-' not in dob and f'/{month_padded}/' not in dob:
                        return False
                
                if year:
                    if not dob.endswith(str(year)):
                        return False
                
                return True
            
            results = results[results['Ngày sinh'].apply(match_dob)]
        
        self.current_results = results.to_dict('records')
        self.display_search_results()
        
        # Tự động hiển thị chi tiết nếu chỉ có 1 kết quả
        if len(self.current_results) == 1:
            self.results_tree.selection_set(self.results_tree.get_children()[0])
            self.show_student_details(self.current_results[0])
    
    def display_search_results(self):
        """Hiển thị kết quả tìm kiếm trong bảng"""
        # Clear tree
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)
        
        # Update count
        count = len(self.current_results)
        self.result_count_label.config(text=f"Kết quả: {count}")
        
        # Populate tree
        for student in self.current_results:
            # Helper function to clean nan values
            def clean_value(val):
                if pd.isna(val) or str(val).lower() in ['nan', '<nan>', 'none']:
                    return ''
                return str(val) if val else ''
            
            values = (
                clean_value(student.get('SBD', '')),
                clean_value(student.get('FULL NAME', '')),
                clean_value(student.get('Ngày sinh', '')),
                clean_value(student.get('KHỐI', '')),
                clean_value(student.get('TRƯỜNG', '')),
                clean_value(student.get('MÃ CERT', student.get('MÃ CERT ĐẦY ĐỦ', ''))),
                clean_value(student.get('TOÁN', student.get('KQ VQG TOÁN', ''))),
                clean_value(student.get('TIẾNG ANH', student.get('KQ VQG TIẾNG ANH', ''))),
                clean_value(student.get('KHOA HỌC', student.get('KQ VQG KHOA HỌC', ''))),
                clean_value(student.get('_SHEET_NAME', ''))
            )
            self.results_tree.insert('', 'end', values=values)
    
    def clear_search_form(self):
        """Xóa form tìm kiếm"""
        self.sbd_entry.delete(0, tk.END)
        self.hoten_entry.delete(0, tk.END)
        self.day_var.set('')
        self.month_var.set('')
        self.year_var.set('')
        
        # Hiển thị lại tất cả
        if self.df_tracuu is not None:
            self.current_results = self.df_tracuu.to_dict('records')
            self.display_search_results()
    
    def on_student_select(self, event):
        """Xử lý khi chọn học sinh trong bảng"""
        selection = self.results_tree.selection()
        if not selection:
            return
        
        # Get selected index
        item = selection[0]
        index = self.results_tree.index(item)
        
        if 0 <= index < len(self.current_results):
            student = self.current_results[index]
            self.show_student_details(student)
    
    def show_student_details(self, student):
        """Hiển thị chi tiết học sinh"""
        # Clear previous details
        for widget in self.details_frame.winfo_children():
            widget.destroy()
        
        # Helper function to get medal class color
        def get_medal_color(score):
            if not score or pd.isna(score):
                return "#bdc3c7"
            score_str = str(score).upper()
            if 'VÀNG' in score_str or 'VANG' in score_str:
                return "#f39c12"  # Gold
            if 'BẠC' in score_str or 'BAC' in score_str:
                return "#95a5a6"  # Silver
            if 'ĐỒNG' in score_str or 'DONG' in score_str:
                return "#cd7f32"  # Bronze
            if 'KHUYẾN KHÍCH' in score_str or 'KK' in score_str:
                return "#3498db"  # Blue
            if 'CHỨNG NHẬN' in score_str or 'CN' in score_str:
                return "#27ae60"  # Green
            return "#bdc3c7"
        
        # Info boxes
        info_data = [
            ("SBD", student.get('SBD', '')),
            ("Họ tên", student.get('FULL NAME', '')),
            ("Ngày sinh", student.get('Ngày sinh', '')),
            ("Khối - Trường", f"{student.get('KHỐI', '')} - {student.get('TRƯỜNG', '')}"),
        ]
        
        # Add sheet name if available
        if student.get('_SHEET_NAME'):
            info_data.append(("Nguồn (Sheet)", student.get('_SHEET_NAME', '')))
        
        for label, value in info_data:
            frame = tk.Frame(self.details_frame, bg="white")
            frame.pack(fill=tk.X, padx=10, pady=3)
            
            # Clean value
            display_value = ''
            if value and not pd.isna(value) and str(value).lower() not in ['nan', '<nan>', 'none']:
                display_value = str(value)
            
            tk.Label(frame, text=label + ":", font=("Arial", 9, "bold"), bg="white", fg="#34495e", width=15, anchor='w').pack(side=tk.LEFT)
            tk.Label(frame, text=display_value, font=("Arial", 9), bg="white", fg="#2c3e50", anchor='w').pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Scores section
        tk.Label(
            self.details_frame,
            text="📊 KẾT QUẢ",
            font=("Arial", 10, "bold"),
            bg="white",
            fg="#2c3e50"
        ).pack(pady=(15, 5))
        
        scores = [
            ("Toán", student.get('TOÁN', student.get('KQ VQG TOÁN', ''))),
            ("Tiếng Anh", student.get('TIẾNG ANH', student.get('KQ VQG TIẾNG ANH', ''))),
            ("Khoa học", student.get('KHOA HỌC', student.get('KQ VQG KHOA HỌC', '')))
        ]
        
        for subject, score in scores:
            score_frame = tk.Frame(self.details_frame, bg="white")
            score_frame.pack(fill=tk.X, padx=10, pady=3)
            
            tk.Label(score_frame, text=subject + ":", font=("Arial", 9), bg="white", width=12, anchor='w').pack(side=tk.LEFT)
            
            # Clean score value
            score_display = ''
            if score and not pd.isna(score) and str(score).lower() not in ['nan', '<nan>', 'none']:
                score_display = str(score)
            
            score_label = tk.Label(
                score_frame,
                text=score_display,
                font=("Arial", 9, "bold"),
                bg=get_medal_color(score),
                fg="white",
                padx=10,
                pady=3,
                relief=tk.RAISED
            )
            score_label.pack(side=tk.LEFT, padx=5)
        
        # Cert code
        tk.Label(
            self.details_frame,
            text="🎖️ MÃ CERT",
            font=("Arial", 10, "bold"),
            bg="white",
            fg="#2c3e50"
        ).pack(pady=(15, 5))
        
        cert_code = student.get('MÃ CERT', student.get('MÃ CERT ĐẦY ĐỦ', ''))
        # Clean cert code
        cert_display = ''
        if cert_code and not pd.isna(cert_code) and str(cert_code).lower() not in ['nan', '<nan>', 'none']:
            cert_display = str(cert_code)
        
        tk.Label(
            self.details_frame,
            text=cert_display,
            font=("Arial", 10),
            bg="white",
            fg="#e74c3c",
            wraplength=300
        ).pack(pady=5)
        
        # Photo section
        if HAS_QR_PIL:
            tk.Label(
                self.details_frame,
                text="📷 ẢNH",
                font=("Arial", 10, "bold"),
                bg="white",
                fg="#2c3e50"
            ).pack(pady=(15, 5))
            
            photo_frame = tk.Frame(self.details_frame, bg="#ecf0f1", relief=tk.SUNKEN, bd=2, width=200, height=200)
            photo_frame.pack(pady=5)
            photo_frame.pack_propagate(False)
            
            # Try to load photo
            photo_loaded = False
            sbd = student.get('SBD', '')
            if sbd:
                photo_path = os.path.join('photos', f"{sbd}.jpg")
                if os.path.exists(photo_path):
                    try:
                        img = Image.open(photo_path)
                        img.thumbnail((190, 190))
                        photo = ImageTk.PhotoImage(img)
                        photo_label = tk.Label(photo_frame, image=photo, bg="#ecf0f1")
                        photo_label.image = photo  # Keep reference
                        photo_label.pack(expand=True)
                        photo_loaded = True
                    except:
                        pass
            
            if not photo_loaded:
                tk.Label(photo_frame, text="No Photo", font=("Arial", 10, "italic"), bg="#ecf0f1", fg="#7f8c8d").pack(expand=True)
            
            # QR Code section
            tk.Label(
                self.details_frame,
                text="🔲 QR CODE",
                font=("Arial", 10, "bold"),
                bg="white",
                fg="#2c3e50"
            ).pack(pady=(15, 5))
            
            qr_frame = tk.Frame(self.details_frame, bg="#ecf0f1", relief=tk.SUNKEN, bd=2, width=200, height=200)
            qr_frame.pack(pady=5)
            qr_frame.pack_propagate(False)
            
            # Generate QR
            try:
                qr_data = f"""STUDENT INFORMATION
Candidate: {student.get('SBD', '')}
Name: {student.get('FULL NAME', '')}
Date of Birth: {student.get('Ngày sinh', '')}
Grade {student.get('KHỐI', '')} - {student.get('TRƯỜNG', '')}

RESULTS:
Math: {scores[0][1] or 'N/A'}
Science: {scores[1][1] or 'N/A'}
English: {scores[2][1] or 'N/A'}

Certificate: {cert_code or 'N/A'}"""
                
                qr = qrcode.QRCode(version=1, box_size=5, border=1)
                qr.add_data(qr_data)
                qr.make(fit=True)
                qr_img = qr.make_image(fill_color="black", back_color="white")
                qr_img = qr_img.resize((190, 190))
                qr_photo = ImageTk.PhotoImage(qr_img)
                qr_label = tk.Label(qr_frame, image=qr_photo, bg="#ecf0f1")
                qr_label.image = qr_photo  # Keep reference
                qr_label.pack(expand=True)
            except:
                tk.Label(qr_frame, text="QR Error", font=("Arial", 10, "italic"), bg="#ecf0f1", fg="#e74c3c").pack(expand=True)


def main():
    root = tk.Tk()
    app = AwardsProcessingApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
